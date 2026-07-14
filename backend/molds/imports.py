import io
import re
from collections import Counter, defaultdict

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .models import (
    ImportBatch,
    Machine,
    MoldAsset,
    MoldModel,
    MoldMovement,
    Processor,
    RackSlot,
    RackZone,
)
from .services import (
    switch_zone_capacity,
    switch_zone_stacking,
    validate_slot,
    validate_target_machine_assignment,
)


STANDARD_SHEETS = {"机台", "加工方", "模具型号", "模具实体"}


def _bool(value, default=True):
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() not in {"0", "false", "否", "停用", "禁用"}


def _text(value):
    return "" if value is None else str(value).strip()


def _headers(ws):
    return {_text(cell.value): index + 1 for index, cell in enumerate(ws[1]) if _text(cell.value)}


def _sheet_rows(ws):
    headers = _headers(ws)
    for row_no in range(2, ws.max_row + 1):
        row = {name: ws.cell(row_no, col).value for name, col in headers.items()}
        if any(value not in (None, "") for value in row.values()):
            yield row_no, row


def _issue(level, message, *, sheet="", row=None, field=""):
    return {"level": level, "sheet": sheet, "row": row, "field": field, "message": message}


def create_standard_template():
    wb = Workbook()
    wb.remove(wb.active)
    definitions = {
        "机台": ["code", "name", "is_active"],
        "加工方": ["code", "name", "contact", "phone", "is_active"],
        "模具型号": ["code", "product_name", "description", "is_active"],
        "模具实体": [
            "asset_code",
            "model_code",
            "status",
            "rack_code",
            "level_no",
            "zone_code",
            "capacity_mode",
            "position_no",
            "stack_level",
            "machine_code",
            "allows_stacking",
            "notes",
        ],
    }
    for title, headers in definitions.items():
        ws = wb.create_sheet(title)
        ws.append(headers)
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = 18
    wb["模具实体"].append(
        ["MJ-001", "ABC-100", "IN_STOCK", "J01", 1, "A", 2, 1, 1, "", True, "示例行，使用前请删除"]
    )
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _parse_standard(wb):
    issues = []
    masters = {"machines": [], "processors": [], "models": []}
    mapping = [
        ("机台", "machines", ["code", "name"]),
        ("加工方", "processors", ["code", "name"]),
        ("模具型号", "models", ["code", "product_name"]),
    ]
    for sheet, key, required in mapping:
        ws = wb[sheet]
        headers = _headers(ws)
        missing = [field for field in required if field not in headers]
        if missing:
            issues.append(_issue("error", f"缺少列：{', '.join(missing)}", sheet=sheet))
            continue
        for row_no, row in _sheet_rows(ws):
            if any(not _text(row.get(field)) for field in required):
                issues.append(_issue("error", "必填字段为空。", sheet=sheet, row=row_no))
                continue
            record = {name: _text(value) for name, value in row.items()}
            record["is_active"] = _bool(row.get("is_active"))
            masters[key].append(record)

    rows = []
    ws = wb["模具实体"]
    headers = _headers(ws)
    required = ["asset_code", "model_code", "status"]
    missing = [field for field in required if field not in headers]
    if missing:
        issues.append(_issue("error", f"缺少列：{', '.join(missing)}", sheet="模具实体"))
        return {
            "source_type": "standard",
            "masters": masters,
            "rows": rows,
            "capacity_modes": [],
            "stacking_modes": [],
            "issues": issues,
        }

    for row_no, raw in _sheet_rows(ws):
        record = {name: _text(value) for name, value in raw.items()}
        record["row_no"] = row_no
        record["status"] = record.get("status", "").upper()
        record["allows_stacking"] = _bool(raw.get("allows_stacking"), False)
        for integer_field in ["level_no", "capacity_mode", "position_no", "stack_level"]:
            value = raw.get(integer_field)
            if value not in (None, ""):
                try:
                    record[integer_field] = int(value)
                except (TypeError, ValueError):
                    issues.append(_issue("error", f"{integer_field}必须是整数。", sheet="模具实体", row=row_no, field=integer_field))
        if not record.get("asset_code") or not record.get("model_code"):
            issues.append(_issue("error", "asset_code和model_code不能为空。", sheet="模具实体", row=row_no))
        if record["status"] not in MoldAsset.Status.values:
            issues.append(_issue("error", "status必须为IN_STOCK、ON_MACHINE或OUTSOURCED。", sheet="模具实体", row=row_no, field="status"))
        if record["status"] == MoldAsset.Status.IN_STOCK:
            needed = ["rack_code", "level_no", "zone_code", "capacity_mode", "position_no", "stack_level"]
            if any(record.get(field) in (None, "") for field in needed):
                issues.append(_issue("error", "在库模具必须填写完整库位字段。", sheet="模具实体", row=row_no))
        elif record["status"] == MoldAsset.Status.ON_MACHINE and not record.get("machine_code"):
            issues.append(_issue("error", "上机模具必须填写machine_code。", sheet="模具实体", row=row_no))
        rows.append(record)

    duplicate_code_keys = {
        code
        for code, count in Counter(
            row.get("asset_code", "").casefold()
            for row in rows
            if row.get("asset_code")
        ).items()
        if count > 1
    }
    for row in rows:
        if row.get("asset_code", "").casefold() in duplicate_code_keys:
            issues.append(_issue("error", f"模具编号重复：{row['asset_code']}", sheet="模具实体", row=row["row_no"]))
    capacity_modes, stacking_modes = _planned_modes_from_rows(
        rows,
        issues,
        include_capacity=True,
    )
    return {
        "source_type": "standard",
        "masters": masters,
        "rows": rows,
        "capacity_modes": capacity_modes,
        "stacking_modes": stacking_modes,
        "issues": issues,
    }


def _legacy_fixed_rows(ws, rack_code, columns, layers):
    for display_index, layer_no in enumerate(range(layers, 0, -1)):
        row = 5 + display_index
        for position_no in range(1, columns + 1):
            cell = ws.cell(row, 2 + position_no)
            yield {
                "sheet": ws.title,
                "row_no": row,
                "value": _text(cell.value),
                "rack_code": rack_code,
                "level_no": layer_no,
                "zone_code": "F",
                "capacity_mode": columns,
                "position_no": position_no,
                "stack_level": 1,
                "active": True,
            }


def _legacy_j01_rows(ws):
    column_mapping = {
        1: ("A", 1),
        2: ("A", 2),
        3: ("B", 1),
        4: ("B", 2),
    }
    for display_index, layer_no in enumerate(range(6, 0, -1)):
        row = 5 + display_index
        for legacy_position, (zone_code, position_no) in column_mapping.items():
            cell = ws.cell(row, 2 + legacy_position)
            yield {
                "sheet": ws.title,
                "row_no": row,
                "value": _text(cell.value),
                "rack_code": "J01",
                "level_no": layer_no,
                "zone_code": zone_code,
                "capacity_mode": 2,
                "position_no": position_no,
                "stack_level": 1,
                "active": True,
            }


def _legacy_j02_rows(ws):
    for display_index, layer_no in enumerate(range(8, 0, -1)):
        base = 4 + display_index * 4
        selectors = {"A": _text(ws[f"C{base}"].value) or "2位", "B": _text(ws[f"J{base}"].value) or "2位"}
        definitions = {
            ("A", 2): [(f"C{base+1}", 1), (f"F{base+1}", 2)],
            ("A", 3): [(f"C{base+2}", 1), (f"E{base+2}", 2), (f"G{base+2}", 3)],
            ("B", 2): [(f"J{base+1}", 1), (f"M{base+1}", 2)],
            ("B", 3): [(f"J{base+2}", 1), (f"L{base+2}", 2), (f"N{base+2}", 3)],
        }
        for (zone, capacity), cells in definitions.items():
            selected = selectors[zone].startswith(str(capacity))
            for coord, position in cells:
                yield {
                    "sheet": ws.title,
                    "row_no": ws[coord].row,
                    "value": _text(ws[coord].value),
                    "rack_code": "J02",
                    "level_no": layer_no,
                    "zone_code": zone,
                    "capacity_mode": capacity,
                    "position_no": position,
                    "stack_level": 1,
                    "active": selected,
                }


def _legacy_j05_rows(ws):
    starts = {2: ["C", "I"], 3: ["C", "G", "K"], 4: ["C", "F", "I", "L"]}
    starts_right = {2: ["P", "V"], 3: ["P", "T", "X"], 4: ["P", "S", "V", "Y"]}
    for display_index, layer_no in enumerate(range(4, 0, -1)):
        base = 4 + display_index * 8
        selectors = {"A": _text(ws[f"C{base}"].value) or "2位", "B": _text(ws[f"P{base}"].value) or "2位"}
        mode_rows = {2: (base + 1, base + 2), 3: (base + 3, base + 4), 4: (base + 5, base + 6)}
        for zone, columns_by_mode in [("A", starts), ("B", starts_right)]:
            for capacity, columns in columns_by_mode.items():
                selected = selectors[zone].startswith(str(capacity))
                upper_row, lower_row = mode_rows[capacity]
                for position, column in enumerate(columns, 1):
                    for stack_level, row in [(2, upper_row), (1, lower_row)]:
                        cell = ws[f"{column}{row}"]
                        yield {
                            "sheet": ws.title,
                            "row_no": row,
                            "value": _text(cell.value),
                            "rack_code": "J05",
                            "level_no": layer_no,
                            "zone_code": zone,
                            "capacity_mode": capacity,
                            "position_no": position,
                            "stack_level": stack_level,
                            "active": selected,
                        }


def _asset_prefix(model_code):
    prefix = re.sub(r"[^A-Za-z0-9_-]+", "-", model_code).strip("-").upper()
    return prefix or "MOLD"


def _parse_legacy(wb):
    issues = []
    candidates = []
    candidates.extend(_legacy_j01_rows(wb["J01-6层4列"]))
    candidates.extend(_legacy_j02_rows(wb["J02-8层可变"]))
    candidates.extend(_legacy_fixed_rows(wb["J03-6层2列"], "J03", 2, 6))
    candidates.extend(_legacy_fixed_rows(wb["J04-6层2列"], "J04", 2, 6))
    candidates.extend(_legacy_j05_rows(wb["J05-4层可变叠放"]))

    for item in candidates:
        if item["value"] and not item["active"]:
            issues.append(_issue("error", "未启用的灰色格位中仍有数据。", sheet=item["sheet"], row=item["row_no"]))

    capacity_modes = []
    seen_capacity_modes = set()
    for item in candidates:
        if not item["active"]:
            continue
        key = (item["rack_code"], item["level_no"], item["zone_code"])
        if key in seen_capacity_modes:
            continue
        seen_capacity_modes.add(key)
        capacity_modes.append(
            {
                "rack_code": item["rack_code"],
                "level_no": item["level_no"],
                "zone_code": item["zone_code"],
                "capacity_mode": item["capacity_mode"],
            }
        )

    active = [item for item in candidates if item["active"] and item["value"]]
    counters = defaultdict(int)
    existing = {
        code.casefold()
        for code in MoldAsset.objects.filter(is_active=True).values_list(
            "asset_code", flat=True
        )
    }
    rows = []
    by_physical = {}
    for item in active:
        if item["value"] == "禁放":
            item["block_slot"] = True
            rows.append(item)
            continue
        model_code = item["value"]
        prefix = _asset_prefix(model_code)
        while True:
            counters[prefix] += 1
            asset_code = f"{prefix}-{counters[prefix]:02d}"
            if asset_code.casefold() not in existing:
                existing.add(asset_code.casefold())
                break
        row = {
            **item,
            "asset_code": asset_code,
            "model_code": model_code,
            "product_name": model_code,
            "status": MoldAsset.Status.IN_STOCK,
            "allows_stacking": False,
            "notes": "由旧台账导入",
        }
        physical = (item["rack_code"], item["level_no"], item["zone_code"], item["capacity_mode"], item["position_no"], item["stack_level"])
        if physical in by_physical:
            issues.append(_issue("error", "同一库位出现多个模具。", sheet=item["sheet"], row=item["row_no"]))
        by_physical[physical] = row
        rows.append(row)

    # 旧表中的上叠空悬属于阻断错误；已经实际叠放的下层模具视为允许叠放。
    rows_by_physical = {
        (
            row["rack_code"],
            row["level_no"],
            row["zone_code"],
            row["capacity_mode"],
            row["position_no"],
            row["stack_level"],
        ): row
        for row in rows
        if row.get("asset_code")
    }
    occupied = {
        (row["rack_code"], row["level_no"], row["zone_code"], row["capacity_mode"], row["position_no"], row["stack_level"])
        for row in rows
        if row.get("asset_code")
    }
    for row in rows:
        if row.get("asset_code") and row["stack_level"] == 2:
            lower = (row["rack_code"], row["level_no"], row["zone_code"], row["capacity_mode"], row["position_no"], 1)
            if lower in occupied:
                rows_by_physical[lower]["allows_stacking"] = True
    _, stacking_modes = _planned_modes_from_rows(
        rows,
        issues,
        include_capacity=False,
    )
    return {
        "source_type": "legacy",
        "masters": {"machines": [], "processors": [], "models": []},
        "rows": rows,
        "capacity_modes": capacity_modes,
        "stacking_modes": stacking_modes,
        "issues": issues,
    }


def _slot_key(row, stack_level=None):
    required = ["rack_code", "level_no", "zone_code", "capacity_mode", "position_no"]
    if any(row.get(field) in (None, "") for field in required):
        return None
    try:
        return (
            str(row["rack_code"]).strip().upper(),
            int(row["level_no"]),
            str(row.get("zone_code") or "F").strip().upper(),
            int(row["capacity_mode"]),
            int(row["position_no"]),
            int(stack_level if stack_level is not None else row.get("stack_level", 1)),
        )
    except (TypeError, ValueError):
        return None


def _planned_modes_from_rows(rows, issues, *, include_capacity):
    capacity_by_zone = {}
    stacking_by_zone = {}
    for row in rows:
        if not row.get("asset_code") and not row.get("block_slot"):
            continue
        if (
            row.get("status", MoldAsset.Status.IN_STOCK)
            != MoldAsset.Status.IN_STOCK
            and not row.get("block_slot")
        ):
            continue
        key = _slot_key(row)
        if key is None:
            continue
        zone_key = key[:3]
        if include_capacity:
            previous = capacity_by_zone.get(zone_key)
            if previous and previous["capacity_mode"] != key[3]:
                issues.append(
                    _issue(
                        "error",
                        "同一货架区域不能计划多个容量模式。",
                        sheet=row.get("sheet", "模具实体"),
                        row=row.get("row_no"),
                        field="capacity_mode",
                    )
                )
            elif not previous:
                capacity_by_zone[zone_key] = {
                    "rack_code": zone_key[0],
                    "level_no": zone_key[1],
                    "zone_code": zone_key[2],
                    "capacity_mode": key[3],
                    "sheet": row.get("sheet", "模具实体"),
                    "row_no": row.get("row_no"),
                }
        # “禁放”只改变候选槽本身，不应因此开启整个区域的叠放模式。
        if row.get("asset_code") and key[5] == 2 and zone_key not in stacking_by_zone:
            stacking_by_zone[zone_key] = {
                "rack_code": zone_key[0],
                "level_no": zone_key[1],
                "zone_code": zone_key[2],
                "stacking_enabled": True,
                "sheet": row.get("sheet", "模具实体"),
                "row_no": row.get("row_no"),
            }
    return list(capacity_by_zone.values()), list(stacking_by_zone.values())


def _find_slot_by_key(key):
    if key is None:
        return None
    return (
        RackSlot.objects.select_related("zone__level__rack")
        .filter(
            zone__level__rack__code=key[0],
            zone__level__level_no=key[1],
            zone__code=key[2],
            capacity_mode=key[3],
            position_no=key[4],
            stack_level=key[5],
        )
        .first()
    )


def _validate_master_rows(result):
    sheet_names = {"machines": "机台", "processors": "加工方", "models": "模具型号"}
    for kind, records in result["masters"].items():
        duplicates = {
            code
            for code, count in Counter(_text(item.get("code")) for item in records).items()
            if code and count > 1
        }
        for code in sorted(duplicates):
            result["issues"].append(
                _issue("error", f"编号重复：{code}", sheet=sheet_names[kind], field="code")
            )


def _validate_capacity_modes(result):
    seen = {}
    for item in result.get("capacity_modes", []):
        sheet = item.get("sheet", "")
        row_no = item.get("row_no")
        try:
            zone_key = (
                str(item["rack_code"]).strip().upper(),
                int(item["level_no"]),
                str(item["zone_code"]).strip().upper(),
            )
            capacity_mode = int(item["capacity_mode"])
        except (KeyError, TypeError, ValueError):
            result["issues"].append(
                _issue("error", "容量模式计划格式无效。", sheet=sheet, row=row_no)
            )
            continue
        previous = seen.get(zone_key)
        if previous is not None and previous != capacity_mode:
            result["issues"].append(
                _issue(
                    "error",
                    "同一货架区域存在冲突的容量模式计划。",
                    sheet=sheet,
                    row=row_no,
                )
            )
            continue
        seen[zone_key] = capacity_mode
        zone = (
            RackZone.objects.select_related("level__rack")
            .filter(
                level__rack__code=zone_key[0],
                level__level_no=zone_key[1],
                code=zone_key[2],
            )
            .first()
        )
        if not zone:
            result["issues"].append(
                _issue(
                    "error",
                    "台账中的货架层或区域在系统中不存在。",
                    sheet=sheet,
                    row=row_no,
                )
            )
            continue
        if not zone.is_active:
            result["issues"].append(
                _issue(
                    "error",
                    f"{zone} 已停用，不能切换容量。",
                    sheet=sheet,
                    row=row_no,
                )
            )
        elif capacity_mode not in zone.allowed_capacities:
            result["issues"].append(
                _issue(
                    "error",
                    "台账选择的容量不属于该区域允许容量。",
                    sheet=sheet,
                    row=row_no,
                )
            )
        elif zone.capacity_mode != capacity_mode and MoldAsset.objects.filter(
            current_slot__zone=zone
        ).exists():
            result["issues"].append(
                _issue(
                    "error",
                    f"{zone} 仍有模具，不能按台账切换容量。",
                    sheet=sheet,
                    row=row_no,
                )
            )


def _validate_stacking_modes(result):
    seen = {}
    for item in result.get("stacking_modes", []):
        sheet = item.get("sheet", "")
        row_no = item.get("row_no")
        try:
            zone_key = (
                str(item["rack_code"]).strip().upper(),
                int(item["level_no"]),
                str(item["zone_code"]).strip().upper(),
            )
        except (KeyError, TypeError, ValueError):
            result["issues"].append(
                _issue("error", "叠放模式计划格式无效。", sheet=sheet, row=row_no)
            )
            continue
        enabled = item.get("stacking_enabled")
        if not isinstance(enabled, bool):
            result["issues"].append(
                _issue(
                    "error",
                    "stacking_enabled必须为布尔值。",
                    sheet=sheet,
                    row=row_no,
                )
            )
            continue
        previous = seen.get(zone_key)
        if previous is not None and previous != enabled:
            result["issues"].append(
                _issue(
                    "error",
                    "同一货架区域存在冲突的叠放模式计划。",
                    sheet=sheet,
                    row=row_no,
                )
            )
            continue
        seen[zone_key] = enabled
        zone = (
            RackZone.objects.select_related("level__rack")
            .filter(
                level__rack__code=zone_key[0],
                level__level_no=zone_key[1],
                code=zone_key[2],
            )
            .first()
        )
        if not zone:
            result["issues"].append(
                _issue(
                    "error",
                    "叠放计划中的货架层或区域在系统中不存在。",
                    sheet=sheet,
                    row=row_no,
                )
            )
            continue
        if not zone.is_active:
            result["issues"].append(
                _issue(
                    "error",
                    f"{zone} 已停用，不能切换叠放。",
                    sheet=sheet,
                    row=row_no,
                )
            )
        elif enabled and not zone.supports_stacking:
            result["issues"].append(
                _issue(
                    "error",
                    f"{zone} 的物理结构不支持叠放。",
                    sheet=sheet,
                    row=row_no,
                )
            )
        elif not enabled and MoldAsset.objects.filter(
            current_slot__zone=zone,
            current_slot__stack_level=2,
        ).exists():
            result["issues"].append(
                _issue(
                    "error",
                    f"{zone} 的上叠位仍有模具，不能关闭叠放。",
                    sheet=sheet,
                    row=row_no,
                )
            )


def _capacity_plan_map(result):
    plans = {}
    for item in result.get("capacity_modes", []):
        try:
            key = (
                str(item["rack_code"]).strip().upper(),
                int(item["level_no"]),
                str(item["zone_code"]).strip().upper(),
            )
            plans[key] = int(item["capacity_mode"])
        except (KeyError, TypeError, ValueError):
            continue
    return plans


def _stacking_plan_map(result):
    plans = {}
    for item in result.get("stacking_modes", []):
        try:
            key = (
                str(item["rack_code"]).strip().upper(),
                int(item["level_no"]),
                str(item["zone_code"]).strip().upper(),
            )
        except (KeyError, TypeError, ValueError):
            continue
        if isinstance(item.get("stacking_enabled"), bool):
            plans[key] = item["stacking_enabled"]
    return plans


def _validate_business_rules(result):
    _validate_master_rows(result)
    _validate_capacity_modes(result)
    _validate_stacking_modes(result)

    source_type = result["source_type"]
    capacity_plans = _capacity_plan_map(result)
    stacking_plans = _stacking_plan_map(result)
    existing_codes = {
        code.casefold()
        for code in MoldAsset.objects.filter(is_active=True).values_list(
            "asset_code", flat=True
        )
    }
    planned_locations = {}
    imported_rows_by_slot = {}

    for row in result["rows"]:
        code = row.get("asset_code")
        if code and code.casefold() in existing_codes:
            result["issues"].append(
                _issue(
                    "error",
                    f"模具编号已存在：{code}",
                    sheet=row.get("sheet", "模具实体"),
                    row=row.get("row_no"),
                )
            )

        if not code and not row.get("block_slot"):
            continue
        if row.get("status", MoldAsset.Status.IN_STOCK) != MoldAsset.Status.IN_STOCK and not row.get("block_slot"):
            continue

        key = _slot_key(row)
        if key is None:
            continue
        if key in planned_locations:
            result["issues"].append(
                _issue(
                    "error",
                    "同一导入文件中存在库位冲突。",
                    sheet=row.get("sheet", "模具实体"),
                    row=row.get("row_no"),
                )
            )
        else:
            planned_locations[key] = row
        if code:
            imported_rows_by_slot[key] = row

        slot = _find_slot_by_key(key)
        if not slot:
            result["issues"].append(
                _issue(
                    "error",
                    "指定库位不存在。",
                    sheet=row.get("sheet", "模具实体"),
                    row=row.get("row_no"),
                )
            )
            continue
        zone_key = key[:3]
        planned_capacity = capacity_plans.get(zone_key, slot.zone.capacity_mode)
        capacity_enabled = slot.capacity_mode == planned_capacity
        planned_stacking = stacking_plans.get(
            zone_key,
            slot.zone.stacking_enabled,
        )
        stacking_enabled = key[5] == 1 or planned_stacking
        base_enabled = (
            slot.zone.level.rack.is_active
            and slot.zone.level.rack.is_configured
            and slot.zone.is_active
        )
        if row.get("block_slot"):
            # 禁放标记可以直接写入预生成的S2候选槽，不需要开启叠放。
            enabled_for_import = base_enabled and capacity_enabled
        else:
            enabled_for_import = (
                base_enabled
                and capacity_enabled
                and stacking_enabled
                and not slot.is_blocked
            )
        if not enabled_for_import:
            result["issues"].append(
                _issue(
                    "error",
                    "指定库位未启用、容量模式不匹配或已禁放。",
                    sheet=row.get("sheet", "模具实体"),
                    row=row.get("row_no"),
                )
            )
        occupant = getattr(slot, "occupant", None)
        if occupant:
            result["issues"].append(
                _issue(
                    "error",
                    f"指定库位已被模具 {occupant.asset_code} 占用。",
                    sheet=row.get("sheet", "模具实体"),
                    row=row.get("row_no"),
                )
            )
        if key[5] == 2 and not slot.zone.supports_stacking:
            result["issues"].append(
                _issue(
                    "error",
                    "指定区域不支持上叠。",
                    sheet=row.get("sheet", "模具实体"),
                    row=row.get("row_no"),
                )
            )

    # 上叠导入必须有实际下层模具；标准模板还必须明确允许叠放。
    for key, row in imported_rows_by_slot.items():
        if key[5] != 2:
            continue
        lower_key = (*key[:5], 1)
        lower_row = imported_rows_by_slot.get(lower_key)
        lower_slot = _find_slot_by_key(lower_key)
        lower_mold = getattr(lower_slot, "occupant", None) if lower_slot else None
        if not lower_row and not lower_mold:
            result["issues"].append(
                _issue(
                    "error",
                    "上叠位置下方没有模具。",
                    sheet=row.get("sheet", "模具实体"),
                    row=row.get("row_no"),
                )
            )
        elif source_type == "standard":
            lower_allows = (
                bool(lower_row and lower_row.get("allows_stacking"))
                or bool(lower_mold and lower_mold.allows_stacking)
            )
            if not lower_allows:
                result["issues"].append(
                    _issue(
                        "error",
                        "上叠位置的下层模具未标记为允许叠放。",
                        sheet=row.get("sheet", "模具实体"),
                        row=row.get("row_no"),
                    )
                )

    if source_type == "standard":
        model_specs = {item["code"]: item for item in result["masters"]["models"]}
        machine_specs = {item["code"]: item for item in result["masters"]["machines"]}
        existing_models = {item.code: item for item in MoldModel.objects.all()}
        existing_machines = {item.code: item for item in Machine.objects.all()}
        for row in result["rows"]:
            sheet = row.get("sheet", "模具实体")
            row_no = row.get("row_no")
            model_code = row.get("model_code")
            model_spec = model_specs.get(model_code)
            model = existing_models.get(model_code)
            if model_code and not model_spec and not model:
                result["issues"].append(
                    _issue("error", f"模具型号不存在：{model_code}", sheet=sheet, row=row_no, field="model_code")
                )
            elif (model_spec and not model_spec.get("is_active", True)) or (not model_spec and model and not model.is_active):
                result["issues"].append(
                    _issue("error", f"模具型号已停用：{model_code}", sheet=sheet, row=row_no, field="model_code")
                )

            if row.get("status") == MoldAsset.Status.ON_MACHINE:
                code = row.get("machine_code")
                spec = machine_specs.get(code)
                machine = existing_machines.get(code)
                if code and not spec and not machine:
                    result["issues"].append(
                        _issue("error", f"机台不存在：{code}", sheet=sheet, row=row_no, field="machine_code")
                    )
                elif (spec and not spec.get("is_active", True)) or (not spec and machine and not machine.is_active):
                    result["issues"].append(
                        _issue("error", f"机台已停用：{code}", sheet=sheet, row=row_no, field="machine_code")
                    )


def preview_workbook(uploaded_file, user):
    wb = load_workbook(uploaded_file, data_only=False)
    if STANDARD_SHEETS.issubset(set(wb.sheetnames)):
        result = _parse_standard(wb)
        kind = ImportBatch.Kind.STANDARD
    elif {"J01-6层4列", "J02-8层可变", "J05-4层可变叠放"}.issubset(set(wb.sheetnames)):
        result = _parse_legacy(wb)
        kind = ImportBatch.Kind.LEGACY
    else:
        raise ValueError("无法识别Excel格式，请使用系统标准模板或实际布局版台账。")

    for index, row in enumerate(result["rows"], start=1):
        row["preview_key"] = str(index)
    _validate_business_rules(result)

    batch = ImportBatch.objects.create(
        kind=kind,
        original_name=getattr(uploaded_file, "name", "upload.xlsx"),
        payload={
            "source_type": result["source_type"],
            "masters": result["masters"],
            "rows": result["rows"],
            "capacity_modes": result.get("capacity_modes", []),
            "stacking_modes": result.get("stacking_modes", []),
        },
        errors=[issue for issue in result["issues"] if issue["level"] == "error"],
        warnings=[issue for issue in result["issues"] if issue["level"] == "warning"],
        created_by=user,
    )
    preview_rows = [row for row in result["rows"] if row.get("asset_code")]
    return {
        "token": str(batch.id),
        "batch_id": str(batch.id),
        "file_name": batch.original_name,
        "source_type": result["source_type"],
        "total_rows": len(preview_rows),
        "valid_rows": len(preview_rows) if not batch.errors else 0,
        "error_count": len(batch.errors),
        "warning_count": len(batch.warnings),
        "rows": [
            {
                "row_no": row.get("row_no", index + 1),
                "row_key": row.get("preview_key", str(index + 1)),
                "asset_code": row.get("asset_code", ""),
                "model_code": row.get("model_code", ""),
                "product_name": row.get("product_name", ""),
                "status": row.get("status", MoldAsset.Status.IN_STOCK),
                "location": _row_location(row),
                "valid": not batch.errors,
            }
            for index, row in enumerate(preview_rows)
        ],
        "issues": batch.errors + batch.warnings,
    }


def _row_location(row):
    if row.get("status") == MoldAsset.Status.IN_STOCK:
        try:
            level_no = int(row.get("level_no", 0))
            position_no = int(row.get("position_no", 0))
            stack_level = int(row.get("stack_level", 1))
        except (TypeError, ValueError):
            return "库位字段格式无效"
        return f"{row.get('rack_code', '')}-L{level_no:02d}-{row.get('zone_code', '')}-P{position_no:02d}-S{stack_level}"
    if row.get("status") == MoldAsset.Status.ON_MACHINE:
        return row.get("machine_code", "")
    if row.get("status") == MoldAsset.Status.OUTSOURCED:
        return "客户收回"
    return ""


def _find_slot(row):
    return RackSlot.objects.select_related("zone__level__rack").get(
        zone__level__rack__code=row["rack_code"],
        zone__level__level_no=row["level_no"],
        zone__code=row["zone_code"] or "F",
        capacity_mode=row["capacity_mode"],
        position_no=row["position_no"],
        stack_level=row.get("stack_level", 1),
    )


@transaction.atomic
def commit_batch(batch, user, asset_code_updates=None):
    batch = ImportBatch.objects.select_for_update().get(pk=batch.pk)
    if batch.status != ImportBatch.Status.PREVIEWED:
        raise ValueError("该导入批次已经处理。")
    if batch.errors:
        raise ValueError("预检存在错误，不能提交。")
    payload = batch.payload
    if asset_code_updates:
        if batch.kind != ImportBatch.Kind.LEGACY:
            raise ValueError("只有旧台账生成的建议模具编号允许在提交前修改。")
        rows_by_key = {
            str(row.get("preview_key")): row
            for row in payload.get("rows", [])
            if row.get("asset_code")
        }
        for update in asset_code_updates:
            if not isinstance(update, dict):
                raise ValueError("提交内容包含无效的编号修改。")
            key = str(update.get("row_key", "")).strip()
            if key not in rows_by_key:
                raise ValueError("提交内容包含无效的预览行。")
            asset_code = _text(update.get("asset_code"))
            if not asset_code:
                raise ValueError("模具编号不能为空。")
            if len(asset_code) > 100:
                raise ValueError(f"模具编号过长：{asset_code[:20]}…")
            rows_by_key[key]["asset_code"] = asset_code

        imported_codes = [
            row["asset_code"]
            for row in payload.get("rows", [])
            if row.get("asset_code")
        ]
        duplicate_keys = {
            code
            for code, count in Counter(
                item.casefold() for item in imported_codes
            ).items()
            if count > 1
        }
        duplicates = {
            code for code in imported_codes if code.casefold() in duplicate_keys
        }
        if duplicates:
            raise ValueError(f"模具编号重复：{', '.join(sorted(duplicates))}")
        existing_code_keys = {
            code.casefold()
            for code in MoldAsset.objects.filter(is_active=True).values_list(
                "asset_code", flat=True
            )
        }
        existing = {
            code for code in imported_codes if code.casefold() in existing_code_keys
        }
        if existing:
            raise ValueError(f"模具编号已存在：{', '.join(sorted(existing))}")
        batch.payload = payload

    mode_validation = {
        "capacity_modes": payload.get("capacity_modes", []),
        "stacking_modes": payload.get("stacking_modes", []),
        "issues": [],
    }
    _validate_capacity_modes(mode_validation)
    _validate_stacking_modes(mode_validation)
    if mode_validation["issues"]:
        messages = "; ".join(
            issue["message"] for issue in mode_validation["issues"]
        )
        raise ValueError(f"库位模式计划无效：{messages}")

    for item in payload.get("capacity_modes", []):
        zone = RackZone.objects.select_for_update().get(
            level__rack__code=item["rack_code"],
            level__level_no=item["level_no"],
            code=item["zone_code"],
        )
        capacity = int(item["capacity_mode"])
        try:
            switch_zone_capacity(zone, capacity)
        except ValidationError as exc:
            raise ValueError(f"{zone} 容量切换失败：{' '.join(exc.messages)}") from exc
    for item in payload.get("stacking_modes", []):
        zone = RackZone.objects.select_for_update().get(
            level__rack__code=item["rack_code"],
            level__level_no=item["level_no"],
            code=item["zone_code"],
        )
        try:
            switch_zone_stacking(zone, item["stacking_enabled"])
        except ValidationError as exc:
            raise ValueError(f"{zone} 叠放切换失败：{' '.join(exc.messages)}") from exc
    for item in payload.get("masters", {}).get("machines", []):
        Machine.objects.update_or_create(code=item["code"], defaults={"name": item["name"], "is_active": item.get("is_active", True)})
    for item in payload.get("masters", {}).get("processors", []):
        Processor.objects.update_or_create(
            code=item["code"],
            defaults={"name": item["name"], "contact": item.get("contact", ""), "phone": item.get("phone", ""), "is_active": item.get("is_active", True)},
        )
    for item in payload.get("masters", {}).get("models", []):
        MoldModel.objects.update_or_create(
            code=item["code"],
            defaults={"product_name": item["product_name"], "description": item.get("description", ""), "is_active": item.get("is_active", True)},
        )

    imported = 0
    for row in payload.get("rows", []):
        if row.get("block_slot"):
            slot = _find_slot(row)
            slot.is_blocked = True
            slot.blocking_reason = "旧台账标记禁放"
            slot.save(update_fields=["is_blocked", "blocking_reason"])
            continue
        model, _ = MoldModel.objects.get_or_create(
            code=row["model_code"], defaults={"product_name": row.get("product_name") or row["model_code"]}
        )
        kwargs = {
            "asset_code": row["asset_code"],
            "mold_model": model,
            "status": row["status"],
            "allows_stacking": row.get("allows_stacking", False),
            "notes": row.get("notes", ""),
            "status_changed_at": timezone.now(),
        }
        if row["status"] == MoldAsset.Status.IN_STOCK:
            slot = _find_slot(row)
            validate_slot(slot)
            kwargs["current_slot"] = slot
        elif row["status"] == MoldAsset.Status.ON_MACHINE:
            machine = Machine.objects.select_for_update().get(code=row["machine_code"])
            try:
                validate_target_machine_assignment(machine)
            except ValidationError as exc:
                raise ValueError("；".join(exc.messages)) from exc
            kwargs["current_machine"] = machine
        mold = MoldAsset(**kwargs)
        mold.full_clean()
        mold.save()
        MoldMovement.objects.create(
            mold=mold,
            action=MoldMovement.Action.CREATE,
            to_status=mold.status,
            to_slot=mold.current_slot,
            to_machine=mold.current_machine,
            to_processor=mold.current_processor,
            note="Excel批量导入",
            operator=user,
        )
        imported += 1
    batch.status = ImportBatch.Status.COMMITTED
    batch.committed_at = timezone.now()
    batch.save(update_fields=["payload", "status", "committed_at"])
    return imported
