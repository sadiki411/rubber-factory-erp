import io

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook

from molds.models import MoldAsset, MoldModel, RackSlot
from molds.services import seed_default_racks


STANDARD_HEADERS = {
    "机台": ["code", "name", "is_active"],
    "加工方": ["code", "name", "contact", "phone", "is_active"],
    "模具型号": ["code", "product_name", "description", "is_active"],
    "模具实体": [
        "asset_code",
        "model_code",
        "status",
        "rack_code",
        "level_no",
        "zone_code",
        "capacity_mode",
        "position_no",
        "stack_level",
        "machine_code",
        "processor_code",
        "allows_stacking",
        "notes",
    ],
}


class SeededRackMixin:
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.user = get_user_model().objects.create_user(username="shared", password="shared-password")
        seed_default_racks()

    @staticmethod
    def slot(rack, level, zone=None, capacity=None, position=1, stack=1):
        if zone is None:
            zone = "F" if rack in {"J03", "J04"} else "A"
        if capacity is None:
            capacity = 2
        return RackSlot.objects.select_related("zone__level__rack").get(
            zone__level__rack__code=rack,
            zone__level__level_no=level,
            zone__code=zone,
            capacity_mode=capacity,
            position_no=position,
            stack_level=stack,
        )

    @staticmethod
    def create_mold(asset_code, slot, *, model=None, allows_stacking=False):
        model = model or MoldModel.objects.create(code=f"MODEL-{asset_code}", product_name=f"产品 {asset_code}")
        return MoldAsset.objects.create(
            asset_code=asset_code,
            mold_model=model,
            status=MoldAsset.Status.IN_STOCK,
            current_slot=slot,
            allows_stacking=allows_stacking,
        )


def workbook_upload(sheet_rows, name="mold-import.xlsx"):
    """Build a standard import workbook from dictionaries keyed by sheet name."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, headers in STANDARD_HEADERS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for row in sheet_rows.get(sheet_name, []):
            sheet.append([row.get(header, "") for header in headers])
    stream = io.BytesIO()
    workbook.save(stream)
    return SimpleUploadedFile(
        name,
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def legacy_workbook_upload(cell_values=None, name="legacy-ledger.xlsx"):
    """Build the minimum workbook shape understood by the visual-ledger parser."""
    cell_values = cell_values or {}
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in [
        "J01-6层4列",
        "J02-8层可变",
        "J03-6层2列",
        "J04-6层2列",
        "J05-4层可变叠放",
    ]:
        sheet = workbook.create_sheet(sheet_name)
        for coordinate, value in cell_values.get(sheet_name, {}).items():
            sheet[coordinate] = value
    stream = io.BytesIO()
    workbook.save(stream)
    return SimpleUploadedFile(
        name,
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
