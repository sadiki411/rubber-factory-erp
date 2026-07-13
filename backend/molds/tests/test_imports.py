import io

from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from molds.imports import commit_batch, create_standard_template, preview_workbook
from molds.models import ImportBatch, Machine, MoldAsset, MoldModel, MoldMovement, Processor

from .helpers import SeededRackMixin, legacy_workbook_upload, workbook_upload


def standard_rows(*asset_rows):
    return {
        "机台": [{"code": "MC-01", "name": "一号机", "is_active": True}],
        "加工方": [{"code": "OUT-01", "name": "外协一厂", "is_active": True}],
        "模具型号": [
            {"code": "ABC-100", "product_name": "密封圈", "description": "", "is_active": True}
        ],
        "模具实体": list(asset_rows),
    }


def stock_row(asset_code="ABC-100-01", **overrides):
    row = {
        "asset_code": asset_code,
        "model_code": "ABC-100",
        "status": "IN_STOCK",
        "rack_code": "J01",
        "level_no": 1,
        "zone_code": "A",
        "capacity_mode": 2,
        "position_no": 1,
        "stack_level": 1,
        "allows_stacking": False,
        "notes": "测试导入",
    }
    row.update(overrides)
    return row


class StandardTemplateTests(TestCase):
    def test_generated_template_has_all_four_sheets_and_structured_location_columns(self):
        workbook = load_workbook(io.BytesIO(create_standard_template()), data_only=False)
        self.assertEqual(workbook.sheetnames, ["机台", "加工方", "模具型号", "模具实体"])
        entity_headers = [cell.value for cell in workbook["模具实体"][1]]
        for required in [
            "asset_code",
            "model_code",
            "status",
            "rack_code",
            "level_no",
            "zone_code",
            "capacity_mode",
            "position_no",
            "stack_level",
            "machine_code",
        ]:
            self.assertIn(required, entity_headers)
        self.assertNotIn("processor_code", entity_headers)
        self.assertNotIn(None, entity_headers)
        self.assertEqual(workbook["模具实体"].max_column, len(entity_headers))
        example = {
            header: workbook["模具实体"].cell(2, column_no + 1).value
            for column_no, header in enumerate(entity_headers)
        }
        self.assertEqual(
            (example["rack_code"], example["zone_code"], example["capacity_mode"]),
            ("J01", "A", 2),
        )
        self.assertTrue(example["allows_stacking"])
        self.assertEqual(example["notes"], "示例行，使用前请删除")


class StandardImportTests(SeededRackMixin, TestCase):
    def test_preview_and_commit_all_three_states_with_history(self):
        upload = workbook_upload(
            standard_rows(
                stock_row(),
                {
                    "asset_code": "ABC-100-02",
                    "model_code": "ABC-100",
                    "status": "ON_MACHINE",
                    "machine_code": "MC-01",
                },
                {
                    "asset_code": "ABC-100-03",
                    "model_code": "ABC-100",
                    "status": "OUTSOURCED",
                },
            )
        )
        preview = preview_workbook(upload, self.user)
        self.assertEqual(preview["source_type"], "standard")
        self.assertEqual(preview["total_rows"], 3)
        self.assertEqual(preview["error_count"], 0)

        batch = ImportBatch.objects.get(pk=preview["token"])
        self.assertEqual(commit_batch(batch, self.user), 3)
        self.assertEqual(MoldAsset.objects.count(), 3)
        self.assertEqual(MoldMovement.objects.filter(action="CREATE").count(), 3)
        self.assertEqual(Machine.objects.get(code="MC-01").current_molds.count(), 1)
        returned = MoldAsset.objects.get(asset_code="ABC-100-03")
        self.assertIsNone(returned.current_processor)
        self.assertEqual(
            MoldMovement.objects.get(mold=returned, action="CREATE").get_to_status_display(),
            "客户收回",
        )
        batch.refresh_from_db()
        self.assertEqual(batch.status, ImportBatch.Status.COMMITTED)
        self.assertIsNotNone(batch.committed_at)

    def test_duplicate_asset_codes_inside_workbook_are_blocking(self):
        upload = workbook_upload(
            standard_rows(
                stock_row(),
                stock_row("abc-100-01", position_no=2),
            )
        )
        preview = preview_workbook(upload, self.user)
        self.assertGreaterEqual(preview["error_count"], 2)
        self.assertTrue(any("模具编号重复" in issue["message"] for issue in preview["issues"]))
        with self.assertRaises(ValueError):
            commit_batch(ImportBatch.objects.get(pk=preview["token"]), self.user)
        self.assertFalse(MoldAsset.objects.exists())

    def test_existing_asset_code_and_repeated_standard_import_are_blocking(self):
        upload_bytes = workbook_upload(standard_rows(stock_row())).read()
        first = preview_workbook(
            workbook_upload(standard_rows(stock_row()), name="first.xlsx"), self.user
        )
        commit_batch(ImportBatch.objects.get(pk=first["token"]), self.user)

        from django.core.files.uploadedfile import SimpleUploadedFile

        repeated = preview_workbook(
            SimpleUploadedFile(
                "repeated.xlsx",
                upload_bytes,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            self.user,
        )
        self.assertGreater(repeated["error_count"], 0)
        self.assertTrue(any("模具编号已存在" in issue["message"] for issue in repeated["issues"]))

    def test_soft_deleted_asset_code_can_be_reused_by_import(self):
        MoldAsset.objects.create(
            asset_code="ABC-100-01",
            mold_model=MoldModel.objects.create(
                code="ARCHIVED-MODEL",
                product_name="已删除记录",
            ),
            status=MoldAsset.Status.OUTSOURCED,
            is_active=False,
        )

        preview = preview_workbook(
            workbook_upload(
                standard_rows(stock_row("abc-100-01")),
                name="reuse-archived-code.xlsx",
            ),
            self.user,
        )
        self.assertEqual(preview["error_count"], 0, preview["issues"])
        self.assertEqual(
            commit_batch(ImportBatch.objects.get(pk=preview["token"]), self.user),
            1,
        )
        self.assertEqual(
            MoldAsset.objects.filter(asset_code__iexact="ABC-100-01").count(),
            2,
        )
        self.assertEqual(
            MoldAsset.objects.filter(
                asset_code__iexact="ABC-100-01",
                is_active=True,
            ).count(),
            1,
        )

    def test_same_batch_cannot_be_committed_twice(self):
        preview = preview_workbook(workbook_upload(standard_rows(stock_row())), self.user)
        batch = ImportBatch.objects.get(pk=preview["token"])
        self.assertEqual(commit_batch(batch, self.user), 1)
        with self.assertRaises(ValueError):
            commit_batch(batch, self.user)
        self.assertEqual(MoldAsset.objects.filter(asset_code="ABC-100-01").count(), 1)

    def test_commit_is_all_or_nothing_when_a_late_row_fails(self):
        batch = ImportBatch.objects.create(
            kind=ImportBatch.Kind.STANDARD,
            original_name="transaction.xlsx",
            created_by=self.user,
            payload={
                "source_type": "standard",
                "masters": {
                    "machines": [],
                    "processors": [],
                    "models": [{"code": "TX-MODEL", "product_name": "事务测试", "is_active": True}],
                },
                "rows": [
                    {
                        **stock_row("TX-001"),
                        "model_code": "TX-MODEL",
                    },
                    {
                        "asset_code": "TX-002",
                        "model_code": "TX-MODEL",
                        "status": MoldAsset.Status.ON_MACHINE,
                        "machine_code": "MISSING-MACHINE",
                    },
                ],
            },
        )
        with self.assertRaises(Machine.DoesNotExist):
            commit_batch(batch, self.user)

        self.assertFalse(MoldAsset.objects.filter(asset_code__startswith="TX-").exists())
        self.assertFalse(MoldModel.objects.filter(code="TX-MODEL").exists())
        batch.refresh_from_db()
        self.assertEqual(batch.status, ImportBatch.Status.PREVIEWED)

    def test_preflight_blocks_two_assets_targeting_same_slot(self):
        preview = preview_workbook(
            workbook_upload(
                standard_rows(
                    stock_row("SLOT-DUP-01"),
                    stock_row("SLOT-DUP-02"),
                )
            ),
            self.user,
        )
        self.assertGreater(preview["error_count"], 0)
        self.assertTrue(any("库位" in issue["message"] for issue in preview["issues"]))

    def test_preflight_blocks_slot_already_occupied_in_database(self):
        slot = self.slot("J01", 1, position=1)
        self.create_mold("EXISTING-001", slot)
        preview = preview_workbook(
            workbook_upload(standard_rows(stock_row("NEW-001"))),
            self.user,
        )
        self.assertGreater(preview["error_count"], 0)
        self.assertTrue(any("占用" in issue["message"] or "库位冲突" in issue["message"] for issue in preview["issues"]))

    def test_preflight_blocks_unknown_or_disabled_slot(self):
        preview = preview_workbook(
            workbook_upload(
                standard_rows(stock_row("BAD-SLOT-001", rack_code="J99"))
            ),
            self.user,
        )
        self.assertGreater(preview["error_count"], 0)
        self.assertTrue(any("库位" in issue["message"] for issue in preview["issues"]))

    def test_preflight_blocks_orphan_upper_stack(self):
        preview = preview_workbook(
            workbook_upload(
                standard_rows(
                    stock_row(
                        "ORPHAN-UPPER",
                        rack_code="J05",
                        zone_code="A",
                        capacity_mode=2,
                        position_no=1,
                        stack_level=2,
                    )
                )
            ),
            self.user,
        )
        self.assertGreater(preview["error_count"], 0)
        self.assertTrue(any("上叠" in issue["message"] or "下方" in issue["message"] for issue in preview["issues"]))


class LegacyImportTests(SeededRackMixin, TestCase):
    def test_legacy_cell_becomes_suggested_asset_and_commits_to_mapped_slot(self):
        upload = legacy_workbook_upload({"J01-6层4列": {"C5": "ABC-100"}})
        preview = preview_workbook(upload, self.user)
        self.assertEqual(preview["source_type"], "legacy")
        self.assertEqual(preview["error_count"], 0)
        self.assertEqual(preview["total_rows"], 1)
        self.assertEqual(preview["rows"][0]["asset_code"], "ABC-100-01")
        self.assertIn("J01-L06", preview["rows"][0]["location"])

        batch = ImportBatch.objects.get(pk=preview["token"])
        self.assertEqual(commit_batch(batch, self.user), 1)
        mold = MoldAsset.objects.select_related("current_slot__zone__level__rack").get(asset_code="ABC-100-01")
        self.assertEqual(mold.current_slot.zone.level.rack.code, "J01")
        self.assertEqual(mold.current_slot.zone.level.level_no, 6)
        self.assertEqual(mold.current_slot.zone.code, "A")
        self.assertEqual(mold.current_slot.capacity_mode, 2)
        self.assertEqual(mold.current_slot.position_no, 1)

    def test_legacy_j01_four_columns_map_across_two_physical_zones(self):
        preview = preview_workbook(
            legacy_workbook_upload(
                {"J01-6层4列": {"C5": "LEFT-1", "D5": "LEFT-2", "E5": "RIGHT-1", "F5": "RIGHT-2"}}
            ),
            self.user,
        )
        self.assertEqual(preview["error_count"], 0)
        batch = ImportBatch.objects.get(pk=preview["token"])
        self.assertEqual(commit_batch(batch, self.user), 4)
        locations = {
            mold.mold_model.code: (
                mold.current_slot.zone.code,
                mold.current_slot.position_no,
                mold.current_slot.capacity_mode,
            )
            for mold in MoldAsset.objects.select_related("mold_model", "current_slot__zone")
        }
        self.assertEqual(
            locations,
            {
                "LEFT-1": ("A", 1, 2),
                "LEFT-2": ("A", 2, 2),
                "RIGHT-1": ("B", 1, 2),
                "RIGHT-2": ("B", 2, 2),
            },
        )

    def test_legacy_suggested_asset_code_can_be_changed_before_commit(self):
        preview = preview_workbook(
            legacy_workbook_upload({"J01-6层4列": {"C5": "ABC-100"}}),
            self.user,
        )
        batch = ImportBatch.objects.get(pk=preview["token"])
        row = preview["rows"][0]
        self.assertEqual(
            commit_batch(
                batch,
                self.user,
                asset_code_updates=[
                    {"row_key": row["row_key"], "asset_code": "CUSTOM-MOLD-01"}
                ],
            ),
            1,
        )
        self.assertTrue(MoldAsset.objects.filter(asset_code="CUSTOM-MOLD-01").exists())
        batch.refresh_from_db()
        self.assertEqual(batch.payload["rows"][0]["asset_code"], "CUSTOM-MOLD-01")

    def test_legacy_capacity_selection_is_applied_even_when_zone_is_empty(self):
        preview = preview_workbook(
            legacy_workbook_upload(
                {"J02-8层可变": {"C4": "3位", "J4": "2位"}}
            ),
            self.user,
        )
        self.assertEqual(preview["error_count"], 0)
        self.assertEqual(
            commit_batch(ImportBatch.objects.get(pk=preview["token"]), self.user),
            0,
        )
        zone = self.slot("J02", 8, zone="A", capacity=3).zone
        zone.refresh_from_db()
        self.assertEqual(zone.capacity_mode, 3)

    def test_legacy_forbidden_marker_blocks_slot_without_creating_mold(self):
        preview = preview_workbook(
            legacy_workbook_upload({"J01-6层4列": {"C5": "禁放"}}), self.user
        )
        self.assertEqual(preview["error_count"], 0)
        self.assertEqual(preview["total_rows"], 0)
        self.assertEqual(commit_batch(ImportBatch.objects.get(pk=preview["token"]), self.user), 0)
        slot = self.slot("J01", 6, position=1)
        slot.refresh_from_db()
        self.assertTrue(slot.is_blocked)
        self.assertIn("旧台账", slot.blocking_reason)
        self.assertFalse(MoldAsset.objects.exists())

    def test_data_in_inactive_gray_capacity_mode_is_blocking(self):
        preview = preview_workbook(
            legacy_workbook_upload(
                {"J02-8层可变": {"C4": "2位", "J4": "2位", "C6": "残留型号"}}
            ),
            self.user,
        )
        self.assertGreater(preview["error_count"], 0)
        self.assertTrue(any("未启用的灰色格位" in issue["message"] for issue in preview["issues"]))

    def test_orphan_upper_stack_in_legacy_sheet_is_blocking(self):
        preview = preview_workbook(
            legacy_workbook_upload(
                {"J05-4层可变叠放": {"C4": "2位", "P4": "2位", "C5": "只有上层"}}
            ),
            self.user,
        )
        self.assertGreater(preview["error_count"], 0)
        self.assertTrue(any("上叠位置下方没有模具" in issue["message"] for issue in preview["issues"]))


class ImportApiTests(SeededRackMixin, TestCase):
    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def test_template_preview_error_report_and_blocked_commit(self):
        template = self.client.get("/api/imports/template/")
        self.assertEqual(template.status_code, 200)
        self.assertEqual(
            load_workbook(io.BytesIO(template.content)).sheetnames,
            ["机台", "加工方", "模具型号", "模具实体"],
        )

        preview = self.client.post(
            "/api/imports/preview/",
            {"file": workbook_upload(standard_rows(stock_row(rack_code="J99")))},
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200)
        payload = preview.json()
        self.assertGreater(payload["error_count"], 0)

        report = self.client.get(f"/api/imports/{payload['token']}/errors/")
        self.assertEqual(report.status_code, 200)
        report_book = load_workbook(io.BytesIO(report.content))
        self.assertEqual(report_book.active.title, "预检问题")
        self.assertGreater(report_book.active.max_row, 1)

        commit = self.client.post(
            "/api/imports/commit/", {"token": payload["token"]}, format="json"
        )
        self.assertEqual(commit.status_code, 400)

    def test_legacy_code_edit_round_trip_through_api(self):
        preview = self.client.post(
            "/api/imports/preview/",
            {"file": legacy_workbook_upload({"J01-6层4列": {"C5": "ABC-100"}})},
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200)
        row = preview.json()["rows"][0]
        committed = self.client.post(
            "/api/imports/commit/",
            {
                "token": preview.json()["token"],
                "rows": [{"row_key": row["row_key"], "asset_code": "API-CUSTOM-01"}],
            },
            format="json",
        )
        self.assertEqual(committed.status_code, 200)
        self.assertEqual(committed.json()["imported_count"], 1)
        self.assertTrue(MoldAsset.objects.filter(asset_code="API-CUSTOM-01").exists())
