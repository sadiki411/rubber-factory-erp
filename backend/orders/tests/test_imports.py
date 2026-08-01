import io
import tempfile
from datetime import date, datetime
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework.test import APITestCase

from orders.imports import commit_business_batch, preview_business_workbook
from orders.models import (
    BusinessImportBatch,
    BusinessRecordRevision,
    MaterialReceipt,
    ProductInspectionCriterion,
    ProductSpecification,
)
from quality.models import QualityOrder


CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SYNTHETIC_SPEC = "TEST-SPEC-A"
SYNTHETIC_MATERIAL = "SYN-RUBBER-A"
SYNTHETIC_FACTORY_ORDER = "TEST-DEMAND-001"
SYNTHETIC_PROJECT = "TEST-PROJECT-001"
CUSTOM_DISPLAY_TEXT = "9/4"
CUSTOM_RAW_VALUE = 98765


def workbook_bytes(workbook):
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def upload(name, content):
    return SimpleUploadedFile(name, content, content_type=CONTENT_TYPE)


def product_workbook(*, literal_strip=False):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "工作表1"
    sheet.append(
        [
            "规格",
            "材质",
            "料长",
            "切料重",
            "条数",
            "一次加硫条件",
            "二烤条件",
            "总孔数",
            "有效孔数",
            "模具在库",
            "备注",
        ]
    )
    sheet.append(
        [
            SYNTHETIC_SPEC,
            SYNTHETIC_MATERIAL,
            "TEST-LENGTH-A",
            "TEST-CUT-WEIGHT-A",
            CUSTOM_RAW_VALUE,
            "TEST-CURE-A",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    if literal_strip:
        sheet["E2"].number_format = CUSTOM_DISPLAY_TEXT
    return workbook_bytes(workbook)


def internal_order_workbook(
    order_no="ORD-1",
    *,
    duplicate=False,
    corrupt_styles=False,
    sheet_title="2026年订单",
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    sheet.append(
        [
            "订单编号",
            "规格",
            "胶料配方",
            "交期",
            "订单量",
            "成型工时",
            "下单时间",
            "模具尺寸",
            "是否生产",
            "出货日期",
        ]
    )
    row = [
        order_no,
        SYNTHETIC_SPEC,
        SYNTHETIC_MATERIAL,
        date(2026, 8, 20),
        240,
        7.5,
        date(2026, 8, 1),
        "TEST-MOLD-SIZE-A",
        "否",
        "",
    ]
    sheet.append(row)
    if duplicate:
        sheet.append(row)
    sheet.append([None, None, None, None, None, None, None, None, "否", None])
    content = workbook_bytes(workbook)
    if not corrupt_styles:
        return content
    source = ZipFile(io.BytesIO(content))
    output = io.BytesIO()
    with source, ZipFile(output, "w", ZIP_DEFLATED) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "xl/styles.xml":
                text = data.decode("utf-8")
                text = text.replace("</fills>", "<fill/></fills>", 1)
                data = text.encode("utf-8")
            target.writestr(item, data)
    return output.getvalue()


def factory_workbook(
    *,
    order_no=SYNTHETIC_FACTORY_ORDER,
    item_no=1,
    order_quantity=2400,
    required_material_kg=24,
    issued_on=date(2026, 8, 3),
    include_criteria=True,
    project_no=SYNTHETIC_PROJECT,
    specification="TEST-SPEC-B",
    mold_no="TEST-MOLD-01",
    criterion_unit="mm",
    material_length=275,
    primary_curing="TEST-CURE-PRIMARY",
    criterion_lower=9.8,
    criterion_upper=10.2,
):
    workbook = Workbook()
    first = workbook.active
    first.title = "sheet1"
    first.append(["生产工作联络单"])
    first.append(
        [
            "协力商：",
            "",
            "NBR-T3",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "发单时间：",
            issued_on,
        ]
    )
    first.append(
        [
            "独立需求号",
            "项次",
            "材质",
            "规格",
            "订单量",
            "完成日",
            "参考工时",
            "胶料用量（KG）",
            "切料重",
            "料长",
            "一次加硫条件",
            "二次加硫条件",
            "模具号",
            "模具尺寸",
        ]
    )
    first.append(
        [
            order_no,
            item_no,
            SYNTHETIC_MATERIAL,
            specification,
            order_quantity,
            date(2026, 8, 10),
            2.5,
            required_material_kg,
            9.75,
            material_length,
            primary_curing,
            "TEST-CURE-SECONDARY",
            mold_no,
            "TEST-MOLD-SIZE-B",
        ]
    )
    if include_criteria:
        second = workbook.create_sheet("Sheet2")
        second.append(["生产工作联络单二"])
        second.append(["独立需求号", "项次", "项目号", "客户", "类别", "版本", "检验项目", "下限", "上限", "单位"])
        second.append(
            [
                order_no,
                item_no,
                project_no,
                "SYNTHETIC-CUSTOMER",
                "TEST-CATEGORY",
                "V1",
                "TEST-DIMENSION",
                criterion_lower,
                criterion_upper,
                criterion_unit,
            ]
        )
    return workbook_bytes(workbook)


def material_workbook(
    batch_no="TEST-BATCH-001",
    *,
    order_no=SYNTHETIC_FACTORY_ORDER,
    item_no=1,
    specification="TEST-SPEC-B",
    material=SYNTHETIC_MATERIAL,
    weight=24.5,
    printed_at="2026-08-04 09:30:00",
    printed_label="打印日期：",
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(["混料发料清单与制造进料检验记录"])
    sheet.append(["共：", "1支", "", "", "课别：", "NBR-T3", printed_label, printed_at])
    sheet.append(["序号", "项次", "独立需求号", "成品品名", "成品规格", "材质", "批号", "出片尺寸", "重量", "制造时间"])
    sheet.append(
        [
            1,
            item_no,
            order_no,
            SYNTHETIC_PROJECT,
            specification,
            material,
            batch_no,
            "TEST-SHEET-SIZE",
            weight,
            "2026/08/04",
        ]
    )
    return workbook_bytes(workbook)


def internal_total_workbook(
    *,
    order_no="TOTAL-ORDER-001",
    specification=SYNTHETIC_SPEC,
    material=SYNTHETIC_MATERIAL,
    order_quantity=240,
    due_date=date(2026, 8, 20),
    manual_received_material_kg="",
    process_card_text="无",
    production_quantity="100M",
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2026年7-9月订单"
    sheet.append(
        [
            "订单编号",
            "流程卡",
            "规格",
            "胶料配方",
            "交期",
            "订单量",
            "胶料用量",
            "已发胶料",
            "成型工时",
            "下单时间",
            "模具尺寸",
            "是否生产",
            "生产数量",
            "出货日期",
            "出货数量",
        ]
    )
    sheet.append(
        [
            order_no,
            process_card_text,
            specification,
            material,
            due_date,
            order_quantity,
            12.5,
            manual_received_material_kg,
            7.5,
            date(2026, 8, 1),
            "TEST-MOLD-SIZE-A",
            "否",
            production_quantity,
            "8月18日",
            "80M",
        ]
    )
    return workbook_bytes(workbook)


class BusinessImportTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(username="importer", password="test")
        self.media_dir = tempfile.TemporaryDirectory()
        self.override = override_settings(MEDIA_ROOT=Path(self.media_dir.name))
        self.override.enable()

    def tearDown(self):
        self.override.disable()
        self.media_dir.cleanup()

    def preview(self, name, content):
        return preview_business_workbook(upload(name, content), self.user)

    def commit(self, result):
        return commit_business_batch(
            BusinessImportBatch.objects.get(pk=result["token"]), self.user
        )

    def test_literal_custom_number_format_uses_display_text_and_keeps_raw_value(self):
        result = self.preview("products.xlsx", product_workbook(literal_strip=True))
        self.assertEqual(result["error_count"], 0, result["issues"])
        self.commit(result)
        product = ProductSpecification.objects.get()
        self.assertEqual(product.strip_count, CUSTOM_DISPLAY_TEXT)
        self.assertEqual(product.raw_data["条数"]["raw_value"], CUSTOM_RAW_VALUE)
        self.assertEqual(
            product.raw_data["条数"]["number_format"], CUSTOM_DISPLAY_TEXT
        )

    def test_bad_styles_uses_safe_ooxml_dates_and_preserves_identical_rows(self):
        result = self.preview(
            "orders.xlsx",
            internal_order_workbook(duplicate=True, corrupt_styles=True),
        )
        self.assertEqual(result["source_type"], "INTERNAL_ORDERS")
        self.assertEqual(result["total_rows"], 2)
        self.assertEqual(result["error_count"], 0, result["issues"])
        self.assertEqual(
            BusinessImportBatch.objects.get(pk=result["token"]).parser,
            "safe-ooxml-1",
        )
        self.commit(result)
        orders = list(QualityOrder.objects.order_by("source_row"))
        self.assertEqual(len(orders), 2)
        self.assertEqual(orders[0].order_date, date(2026, 8, 1))
        self.assertEqual(orders[0].due_date, date(2026, 8, 20))
        self.assertNotEqual(orders[0].source_key, orders[1].source_key)

    def test_internal_order_links_unique_product_and_warns_on_ambiguous_match(self):
        unique = ProductSpecification.objects.create(
            specification=SYNTHETIC_SPEC, material=SYNTHETIC_MATERIAL
        )
        result = self.preview("order-unique.xlsx", internal_order_workbook("LINK-1"))
        self.commit(result)
        self.assertEqual(QualityOrder.objects.get(order_no="LINK-1").product_specification, unique)

        ProductSpecification.objects.create(
            specification=SYNTHETIC_SPEC, material=SYNTHETIC_MATERIAL
        )
        ambiguous = self.preview(
            "order-ambiguous.xlsx",
            internal_order_workbook("LINK-2", sheet_title="2026年订单-2"),
        )
        self.assertTrue(any("多条相同规格" in item["message"] for item in ambiguous["issues"]))
        self.commit(ambiguous)
        self.assertIsNone(QualityOrder.objects.get(order_no="LINK-2").product_specification)

    def test_factory_work_contact_imports_order_product_and_criterion(self):
        result = self.preview("factory.xlsx", factory_workbook())
        self.assertEqual(
            result["counts"],
            {
                "product_specifications": 1,
                "orders": 1,
                "material_receipts": 0,
                "inspection_criteria": 1,
            },
        )
        self.assertEqual(result["error_count"], 0, result["issues"])
        self.commit(result)
        order = QualityOrder.objects.get(order_no=SYNTHETIC_FACTORY_ORDER)
        criterion = ProductInspectionCriterion.objects.get()
        self.assertEqual(order.product_specification_id, criterion.product_specification_id)
        self.assertEqual(criterion.order_id, order.pk)
        self.assertEqual(criterion.project_no, SYNTHETIC_PROJECT)
        self.assertTrue(order.production_required)

    def test_factory_main_sheet_without_criteria_is_still_an_order_import(self):
        result = self.preview(
            "factory-main-only.xlsx", factory_workbook(include_criteria=False)
        )
        self.assertEqual(result["source_type"], "FACTORY_WORK_CONTACT")
        self.assertEqual(result["counts"]["orders"], 1)
        self.assertEqual(result["counts"]["inspection_criteria"], 0)
        self.assertEqual(result["error_count"], 0, result["issues"])

    def test_factory_main_only_reuses_existing_project_product_and_keeps_criteria_linked(self):
        complete = self.preview(
            "factory-complete.xlsx",
            factory_workbook(issued_on=date(2026, 8, 3)),
        )
        self.commit(complete)
        order = QualityOrder.objects.get(order_no=SYNTHETIC_FACTORY_ORDER)
        product = order.product_specification
        criterion = ProductInspectionCriterion.objects.get()
        product_id = product.pk

        main_only = self.preview(
            "factory-main-only-newer.xlsx",
            factory_workbook(
                issued_on=date(2026, 8, 5),
                include_criteria=False,
                material_length=300,
                primary_curing="MAIN-ONLY-UPDATED-CURE",
            ),
        )
        rows = {row["record_type"]: row for row in main_only["rows"]}
        self.assertEqual(main_only["error_count"], 0, main_only["issues"])
        self.assertEqual(rows["PRODUCT_SPECIFICATION"]["action"], "UPDATE")
        self.assertEqual(rows["PRODUCT_SPECIFICATION"]["match_id"], product_id)
        self.commit(main_only)

        order.refresh_from_db()
        product.refresh_from_db()
        criterion.refresh_from_db()
        self.assertEqual(ProductSpecification.objects.count(), 1)
        self.assertEqual(order.product_specification_id, product_id)
        self.assertEqual(criterion.product_specification_id, product_id)
        self.assertEqual(product.customer_product_no, SYNTHETIC_PROJECT)
        self.assertEqual(product.material_length, "300")
        self.assertEqual(product.primary_curing, "MAIN-ONLY-UPDATED-CURE")

    def test_factory_main_only_blocks_product_identity_change_without_project(self):
        complete = self.preview("factory-complete-base.xlsx", factory_workbook())
        self.commit(complete)
        order = QualityOrder.objects.get(order_no=SYNTHETIC_FACTORY_ORDER)
        product_id = order.product_specification_id
        criterion_id = ProductInspectionCriterion.objects.get().pk

        unsafe = self.preview(
            "factory-main-only-unsafe-identity.xlsx",
            factory_workbook(
                issued_on=date(2026, 8, 5),
                include_criteria=False,
                specification="DIFFERENT-SPEC",
            ),
        )
        rows = {row["record_type"]: row for row in unsafe["rows"]}
        self.assertGreater(unsafe["error_count"], 0)
        self.assertEqual(rows["PRODUCT_SPECIFICATION"]["action"], "SKIP")
        self.assertEqual(rows["ORDER"]["action"], "SKIP")
        self.assertTrue(
            any(
                issue.get("field") == "customer_product_no"
                and "不能安全复用" in issue["message"]
                for issue in unsafe["issues"]
            )
        )
        with self.assertRaises(ValueError):
            self.commit(unsafe)
        order.refresh_from_db()
        self.assertEqual(ProductSpecification.objects.count(), 1)
        self.assertEqual(order.product_specification_id, product_id)
        self.assertEqual(ProductInspectionCriterion.objects.get().pk, criterion_id)

    def test_existing_duplicate_product_identity_reuses_order_link_without_third_product(self):
        first = self.preview(
            "factory-before-product-duplicate.xlsx",
            factory_workbook(issued_on=date(2026, 8, 3)),
        )
        self.commit(first)
        order = QualityOrder.objects.get(order_no=SYNTHETIC_FACTORY_ORDER)
        linked_product = order.product_specification
        ProductSpecification.objects.create(
            customer_product_no=linked_product.customer_product_no,
            specification=linked_product.specification,
            material=linked_product.material,
            mold_no=linked_product.mold_no,
        )

        newer = self.preview(
            "factory-after-product-duplicate.xlsx",
            factory_workbook(
                issued_on=date(2026, 8, 5),
                material_length=300,
            ),
        )
        product_row = next(
            row
            for row in newer["rows"]
            if row["record_type"] == "PRODUCT_SPECIFICATION"
        )
        self.assertEqual(newer["error_count"], 0, newer["issues"])
        self.assertEqual(product_row["match_id"], linked_product.pk)
        self.assertEqual(product_row["action"], "UPDATE")
        self.commit(newer)
        linked_product.refresh_from_db()
        order.refresh_from_db()
        self.assertEqual(ProductSpecification.objects.count(), 2)
        self.assertEqual(order.product_specification_id, linked_product.pk)
        self.assertEqual(linked_product.material_length, "300")

    def test_existing_duplicate_product_identity_without_linked_order_blocks_preview(self):
        for _ in range(2):
            ProductSpecification.objects.create(
                customer_product_no=SYNTHETIC_PROJECT,
                specification="TEST-SPEC-B",
                material=SYNTHETIC_MATERIAL,
                mold_no="TEST-MOLD-01",
            )
        preview = self.preview("factory-ambiguous-products.xlsx", factory_workbook())
        product_row = next(
            row
            for row in preview["rows"]
            if row["record_type"] == "PRODUCT_SPECIFICATION"
        )
        self.assertGreater(preview["error_count"], 0)
        self.assertEqual(product_row["action"], "SKIP")
        self.assertTrue(
            any(
                issue.get("field") == "product_specification"
                and "多条相同产品身份" in issue["message"]
                for issue in preview["issues"]
            )
        )
        self.assertEqual(ProductSpecification.objects.count(), 2)
        self.assertEqual(QualityOrder.objects.count(), 0)

    def test_factory_batch_reuses_identical_product_for_multiple_orders(self):
        workbook = load_workbook(
            io.BytesIO(factory_workbook(include_criteria=False))
        )
        sheet = workbook["sheet1"]
        second_row = [cell.value for cell in sheet[4]]
        second_row[0] = "TEST-DEMAND-002"
        second_row[1] = 2
        sheet.append(second_row)
        preview = self.preview("factory-two-orders.xlsx", workbook_bytes(workbook))
        self.assertEqual(preview["error_count"], 0, preview["issues"])
        self.assertEqual(preview["counts"]["product_specifications"], 1)
        self.assertEqual(preview["counts"]["orders"], 2)
        self.commit(preview)

        self.assertEqual(ProductSpecification.objects.count(), 1)
        self.assertEqual(QualityOrder.objects.count(), 2)
        self.assertEqual(
            len(
                set(
                    QualityOrder.objects.values_list(
                        "product_specification_id", flat=True
                    )
                )
            ),
            1,
        )

    def test_factory_batch_blocks_conflicting_process_for_same_product_identity(self):
        workbook = load_workbook(
            io.BytesIO(factory_workbook(include_criteria=False))
        )
        sheet = workbook["sheet1"]
        second_row = [cell.value for cell in sheet[4]]
        second_row[0] = "TEST-DEMAND-002"
        second_row[1] = 2
        second_row[9] = 999
        sheet.append(second_row)
        preview = self.preview(
            "factory-conflicting-product.xlsx", workbook_bytes(workbook)
        )
        self.assertGreater(preview["error_count"], 0)
        conflict = next(
            issue
            for issue in preview["issues"]
            if issue.get("field") == "product_specification"
        )
        self.assertIn("第4行", conflict["message"])
        self.assertIn("第5行", conflict["message"])

    def test_factory_order_uses_stable_key_to_update_and_preserve_manual_fields(self):
        first = self.preview("factory-first.xlsx", factory_workbook())
        self.commit(first)
        order = QualityOrder.objects.get(order_no=SYNTHETIC_FACTORY_ORDER)
        order.production_required = False
        order.process_card_text = "人工确认：2张"
        order.notes = "手工排产备注"
        order.save()

        changed = self.preview(
            "factory-changed.xlsx",
            factory_workbook(
                order_quantity=2600,
                required_material_kg=26,
                issued_on=date(2026, 8, 5),
            ),
        )
        order_row = next(row for row in changed["rows"] if row["record_type"] == "ORDER")
        self.assertEqual(order_row["action"], "UPDATE")
        self.assertIn("order_quantity", order_row["changes"])
        result = self.commit(changed)
        self.assertEqual(result["updated"]["orders"], 1)

        order.refresh_from_db()
        self.assertEqual(QualityOrder.objects.filter(order_no=SYNTHETIC_FACTORY_ORDER).count(), 1)
        self.assertEqual(order.order_quantity, 2600)
        self.assertEqual(order.required_material_kg, 26)
        self.assertFalse(order.production_required)
        self.assertEqual(order.process_card_text, "人工确认：2张")
        self.assertEqual(order.notes, "手工排产备注")
        self.assertEqual(order.source_system, "NBR-T3")
        self.assertEqual(
            timezone.localtime(order.source_document_at).date(), date(2026, 8, 5)
        )

        business_updated_at = order.updated_at
        imported_at = order.last_imported_at
        repeated = self.preview(
            "factory-changed-copy.xlsx",
            factory_workbook(
                order_quantity=2600,
                required_material_kg=26,
                issued_on=date(2026, 8, 5),
            ),
        )
        self.commit(repeated)
        order.refresh_from_db()
        self.assertEqual(order.updated_at, business_updated_at)
        self.assertGreaterEqual(order.last_imported_at, imported_at)

    def test_older_factory_source_does_not_overwrite_newer_order(self):
        current = self.preview(
            "factory-current.xlsx",
            factory_workbook(order_quantity=2600, issued_on=date(2026, 8, 5)),
        )
        self.commit(current)
        older = self.preview(
            "factory-older.xlsx",
            factory_workbook(order_quantity=2400, issued_on=date(2026, 8, 3)),
        )
        order_row = next(row for row in older["rows"] if row["record_type"] == "ORDER")
        self.assertEqual(order_row["action"], "SKIP")
        self.commit(older)
        self.assertEqual(
            QualityOrder.objects.get(order_no=SYNTHETIC_FACTORY_ORDER).order_quantity,
            2600,
        )

    def test_older_factory_source_skips_order_product_and_all_criteria_as_a_group(self):
        newer = self.preview(
            "factory-newer-process.xlsx",
            factory_workbook(
                issued_on=date(2026, 8, 5),
                order_quantity=2600,
                material_length=300,
                primary_curing="NEW-CURE-PROCESS",
                criterion_lower=9.9,
                criterion_upper=10.1,
            ),
        )
        self.commit(newer)
        order = QualityOrder.objects.get(order_no=SYNTHETIC_FACTORY_ORDER)
        product = order.product_specification
        criterion = ProductInspectionCriterion.objects.get()
        original_ids = (order.pk, product.pk, criterion.pk)

        older = self.preview(
            "factory-older-process.xlsx",
            factory_workbook(
                issued_on=date(2026, 8, 3),
                order_quantity=2400,
                material_length=275,
                primary_curing="OLD-CURE-PROCESS",
                criterion_lower=9.8,
                criterion_upper=10.2,
            ),
        )
        actions = {row["record_type"]: row["action"] for row in older["rows"]}
        self.assertEqual(
            actions,
            {
                "PRODUCT_SPECIFICATION": "SKIP",
                "ORDER": "SKIP",
                "INSPECTION_CRITERION": "SKIP",
            },
        )
        result = self.commit(older)
        self.assertEqual(result["skipped"]["product_specifications"], 1)
        self.assertEqual(result["skipped"]["orders"], 1)
        self.assertEqual(result["skipped"]["inspection_criteria"], 1)

        order.refresh_from_db()
        product.refresh_from_db()
        criterion.refresh_from_db()
        self.assertEqual(QualityOrder.objects.count(), 1)
        self.assertEqual(ProductSpecification.objects.count(), 1)
        self.assertEqual(ProductInspectionCriterion.objects.count(), 1)
        self.assertEqual((order.pk, product.pk, criterion.pk), original_ids)
        self.assertEqual(order.order_quantity, 2600)
        self.assertEqual(product.material_length, "300")
        self.assertEqual(product.primary_curing, "NEW-CURE-PROCESS")
        self.assertEqual(criterion.lower_limit, "9.9")
        self.assertEqual(criterion.upper_limit, "10.1")
        self.assertEqual(order.product_specification_id, product.pk)
        self.assertEqual(criterion.product_specification_id, product.pk)
        self.assertEqual(criterion.order_id, order.pk)

    def test_older_new_order_reuses_shared_product_without_reverting_process(self):
        newer = self.preview(
            "factory-order-a-newer.xlsx",
            factory_workbook(
                order_no="TEST-DEMAND-A",
                item_no=1,
                issued_on=date(2026, 8, 5),
                material_length=300,
                primary_curing="NEW-SHARED-CURE",
            ),
        )
        self.commit(newer)
        product = ProductSpecification.objects.get()
        product_id = product.pk

        older_other_order = self.preview(
            "factory-order-b-older.xlsx",
            factory_workbook(
                order_no="TEST-DEMAND-B",
                item_no=2,
                issued_on=date(2026, 7, 1),
                material_length=275,
                primary_curing="OLD-SHARED-CURE",
            ),
        )
        rows = {
            row["record_type"]: row for row in older_other_order["rows"]
        }
        self.assertEqual(older_other_order["error_count"], 0, older_other_order["issues"])
        self.assertEqual(rows["PRODUCT_SPECIFICATION"]["action"], "SKIP")
        self.assertEqual(rows["PRODUCT_SPECIFICATION"]["match_id"], product_id)
        self.assertEqual(rows["ORDER"]["action"], "CREATE")
        self.assertEqual(rows["INSPECTION_CRITERION"]["action"], "CREATE")
        self.assertTrue(
            any(
                "不会回退共享产品工艺参数" in issue["message"]
                for issue in older_other_order["issues"]
            )
        )
        self.commit(older_other_order)

        product.refresh_from_db()
        order_b = QualityOrder.objects.get(order_no="TEST-DEMAND-B")
        self.assertEqual(ProductSpecification.objects.count(), 1)
        self.assertEqual(QualityOrder.objects.count(), 2)
        self.assertEqual(ProductInspectionCriterion.objects.count(), 2)
        self.assertEqual(order_b.product_specification_id, product_id)
        self.assertEqual(product.material_length, "300")
        self.assertEqual(product.primary_curing, "NEW-SHARED-CURE")

    def test_stale_factory_group_with_unmatched_dependencies_never_creates(self):
        newer = self.preview(
            "factory-newer-identity.xlsx",
            factory_workbook(
                issued_on=date(2026, 8, 5),
                project_no="NEW-PROJECT",
                mold_no="NEW-MOLD",
                criterion_unit="mm",
                material_length=300,
                criterion_lower=9.9,
                criterion_upper=10.1,
            ),
        )
        self.commit(newer)
        order = QualityOrder.objects.get(order_no=SYNTHETIC_FACTORY_ORDER)
        product_id = order.product_specification_id
        criterion = ProductInspectionCriterion.objects.get()

        older = self.preview(
            "factory-older-identity.xlsx",
            factory_workbook(
                issued_on=date(2026, 8, 3),
                project_no="OLD-PROJECT",
                mold_no="OLD-MOLD",
                criterion_unit="cm",
                material_length=275,
                criterion_lower=9.8,
                criterion_upper=10.2,
            ),
        )
        rows = {row["record_type"]: row for row in older["rows"]}
        self.assertEqual(rows["PRODUCT_SPECIFICATION"]["action"], "SKIP")
        self.assertEqual(rows["ORDER"]["action"], "SKIP")
        self.assertEqual(rows["INSPECTION_CRITERION"]["action"], "SKIP")
        self.assertIsNone(rows["PRODUCT_SPECIFICATION"]["match_id"])
        self.assertIsNone(rows["INSPECTION_CRITERION"]["match_id"])
        self.commit(older)

        order.refresh_from_db()
        criterion.refresh_from_db()
        self.assertEqual(QualityOrder.objects.count(), 1)
        self.assertEqual(ProductSpecification.objects.count(), 1)
        self.assertEqual(ProductInspectionCriterion.objects.count(), 1)
        self.assertEqual(order.product_specification_id, product_id)
        self.assertEqual(criterion.product_specification_id, product_id)
        self.assertEqual(criterion.project_no, "NEW-PROJECT")
        self.assertEqual(criterion.unit, "mm")

    def test_missing_factory_source_time_cannot_overwrite_a_versioned_group(self):
        current = self.preview(
            "factory-versioned.xlsx",
            factory_workbook(
                issued_on=date(2026, 8, 5),
                order_quantity=2600,
                material_length=300,
                criterion_lower=9.9,
            ),
        )
        self.commit(current)
        order = QualityOrder.objects.get(order_no=SYNTHETIC_FACTORY_ORDER)
        product = order.product_specification
        criterion = ProductInspectionCriterion.objects.get()

        missing_time = self.preview(
            "factory-without-source-time.xlsx",
            factory_workbook(
                issued_on=None,
                order_quantity=2400,
                material_length=275,
                criterion_lower=9.8,
            ),
        )
        self.assertTrue(
            any("时间缺失或早于" in issue["message"] for issue in missing_time["issues"])
        )
        self.assertEqual(
            {row["record_type"]: row["action"] for row in missing_time["rows"]},
            {
                "PRODUCT_SPECIFICATION": "SKIP",
                "ORDER": "SKIP",
                "INSPECTION_CRITERION": "SKIP",
            },
        )
        self.commit(missing_time)

        order.refresh_from_db()
        product.refresh_from_db()
        criterion.refresh_from_db()
        self.assertEqual(order.order_quantity, 2600)
        self.assertEqual(product.material_length, "300")
        self.assertEqual(criterion.lower_limit, "9.9")
        self.assertEqual(order.product_specification_id, criterion.product_specification_id)

    def test_newer_factory_source_time_updates_lineage_when_business_data_is_same(self):
        first = self.preview(
            "factory-0900.xlsx",
            factory_workbook(issued_on=datetime(2026, 8, 3, 9, 0)),
        )
        self.commit(first)
        order = QualityOrder.objects.get(order_no=SYNTHETIC_FACTORY_ORDER)
        first_batch_id = order.source_batch_id
        self.assertEqual(order.last_source_batch_id, first_batch_id)

        newer = self.preview(
            "factory-1100.xlsx",
            factory_workbook(issued_on=datetime(2026, 8, 3, 11, 0)),
        )
        order_row = next(row for row in newer["rows"] if row["record_type"] == "ORDER")
        self.assertEqual(order_row["action"], "UPDATE")
        self.assertIn("source_document_at", order_row["changes"])
        self.commit(newer)

        order.refresh_from_db()
        self.assertEqual(order.source_batch_id, first_batch_id)
        self.assertEqual(order.last_source_batch_id, BusinessImportBatch.objects.get(pk=newer["token"]).pk)
        self.assertEqual(timezone.localtime(order.source_document_at).hour, 11)

        repeated = self.preview(
            "factory-1100-copy.xlsx",
            factory_workbook(issued_on=datetime(2026, 8, 3, 11, 0)),
        )
        repeated_row = next(
            row for row in repeated["rows"] if row["record_type"] == "ORDER"
        )
        self.assertEqual(repeated_row["action"], "SKIP")
        self.commit(repeated)
        order.refresh_from_db()
        self.assertEqual(order.source_batch_id, first_batch_id)
        self.assertEqual(order.last_source_batch_id, BusinessImportBatch.objects.get(pk=repeated["token"]).pk)

    def test_factory_product_identity_change_updates_order_and_criterion_links(self):
        first = self.preview("factory-original.xlsx", factory_workbook())
        self.commit(first)
        order = QualityOrder.objects.get(order_no=SYNTHETIC_FACTORY_ORDER)
        criterion = ProductInspectionCriterion.objects.get()
        original_product_id = order.product_specification_id
        original_criterion_id = criterion.pk

        changed = self.preview(
            "factory-corrected-product.xlsx",
            factory_workbook(
                project_no="TEST-PROJECT-002",
                mold_no="TEST-MOLD-02",
            ),
        )
        actions = {row["record_type"]: row["action"] for row in changed["rows"]}
        self.assertEqual(actions["PRODUCT_SPECIFICATION"], "CREATE")
        self.assertEqual(actions["ORDER"], "UPDATE")
        self.assertEqual(actions["INSPECTION_CRITERION"], "UPDATE")
        self.commit(changed)

        order.refresh_from_db()
        criterion.refresh_from_db()
        self.assertEqual(ProductSpecification.objects.count(), 2)
        self.assertNotEqual(order.product_specification_id, original_product_id)
        self.assertEqual(criterion.pk, original_criterion_id)
        self.assertEqual(criterion.product_specification_id, order.product_specification_id)
        self.assertEqual(criterion.project_no, "TEST-PROJECT-002")
        self.assertEqual(order.product_specification.mold_no, "TEST-MOLD-02")

    def test_inspection_criterion_matching_keeps_units_distinct(self):
        first = self.preview("criterion-mm.xlsx", factory_workbook(criterion_unit="mm"))
        self.commit(first)
        second = self.preview("criterion-cm.xlsx", factory_workbook(criterion_unit="cm"))
        criterion_row = next(
            row for row in second["rows"] if row["record_type"] == "INSPECTION_CRITERION"
        )
        self.assertEqual(criterion_row["action"], "CREATE")
        self.commit(second)
        self.assertEqual(
            set(ProductInspectionCriterion.objects.values_list("unit", flat=True)),
            {"mm", "cm"},
        )

    def test_factory_deduplicates_identical_criteria_within_one_file(self):
        workbook = load_workbook(io.BytesIO(factory_workbook()))
        criteria_sheet = workbook["Sheet2"]
        criteria_sheet.append([cell.value for cell in criteria_sheet[3]])

        preview = self.preview(
            "factory-identical-criteria.xlsx", workbook_bytes(workbook)
        )
        self.assertEqual(preview["error_count"], 0, preview["issues"])
        self.assertEqual(preview["counts"]["inspection_criteria"], 1)
        self.commit(preview)
        self.assertEqual(ProductInspectionCriterion.objects.count(), 1)

    def test_factory_blocks_conflicting_criteria_within_one_file(self):
        workbook = load_workbook(io.BytesIO(factory_workbook()))
        criteria_sheet = workbook["Sheet2"]
        duplicate = [cell.value for cell in criteria_sheet[3]]
        duplicate[7] = 9.7
        criteria_sheet.append(duplicate)

        preview = self.preview(
            "factory-conflicting-criteria.xlsx", workbook_bytes(workbook)
        )
        self.assertGreater(preview["error_count"], 0)
        issue = next(
            item
            for item in preview["issues"]
            if item.get("field") == "inspection_criterion"
        )
        self.assertIn("第3行", issue["message"])
        self.assertIn("第4行", issue["message"])
        with self.assertRaises(ValueError):
            self.commit(preview)
        self.assertEqual(ProductInspectionCriterion.objects.count(), 0)

    def test_factory_blocks_ambiguous_existing_criterion_identity(self):
        first = self.preview("factory-criterion-base.xlsx", factory_workbook())
        self.commit(first)
        existing = ProductInspectionCriterion.objects.get()
        ProductInspectionCriterion.objects.create(
            product_specification=existing.product_specification,
            order=existing.order,
            item_no=existing.item_no,
            project_no=existing.project_no,
            customer=existing.customer,
            category=existing.category,
            version=existing.version,
            inspection_item=existing.inspection_item,
            lower_limit=existing.lower_limit,
            upper_limit=existing.upper_limit,
            unit=existing.unit,
        )

        newer = self.preview(
            "factory-criterion-ambiguous.xlsx",
            factory_workbook(
                issued_on=date(2026, 8, 5),
                criterion_lower=9.9,
            ),
        )
        criterion_row = next(
            row
            for row in newer["rows"]
            if row["record_type"] == "INSPECTION_CRITERION"
        )
        self.assertGreater(newer["error_count"], 0)
        self.assertEqual(criterion_row["action"], "SKIP")
        self.assertTrue(
            any(
                issue.get("field") == "inspection_criterion"
                and "多条相同检验标准业务键" in issue["message"]
                for issue in newer["issues"]
            )
        )
        with self.assertRaises(ValueError):
            self.commit(newer)
        self.assertEqual(ProductInspectionCriterion.objects.count(), 2)

    def test_internal_total_fifteen_columns_update_same_row_and_keep_text(self):
        first = self.preview(
            "2026-total.xlsx",
            internal_total_workbook(manual_received_material_kg=6.25),
        )
        self.commit(first)
        order = QualityOrder.objects.get(order_no="TOTAL-ORDER-001")
        self.assertEqual(order.process_card_text, "无")
        self.assertEqual(order.production_quantity, "100M")
        self.assertEqual(order.shipment_date, "8月18日")
        self.assertEqual(order.shipped_quantity, "80M")
        self.assertEqual(order.required_material_kg, 12.5)
        self.assertEqual(order.manual_received_material_kg, 6.25)
        self.assertEqual(order.order_date, date(2026, 8, 1))
        self.assertEqual(order.due_date, date(2026, 8, 20))
        order.notes = "在线手工备注"
        order.manual_received_material_kg = 7.5
        order.save()
        receipt = MaterialReceipt.objects.create(
            order=order,
            order_no=order.order_no,
            specification=order.specification,
            material=order.material,
            weight_kg=2.5,
        )

        changed = self.preview(
            "2026-total-updated.xlsx",
            internal_total_workbook(order_quantity=260, production_quantity="120M"),
        )
        order_row = next(row for row in changed["rows"] if row["record_type"] == "ORDER")
        self.assertEqual(order_row["action"], "UPDATE")
        result = self.commit(changed)
        self.assertEqual(result["updated"]["orders"], 1)
        order.refresh_from_db()
        self.assertEqual(QualityOrder.objects.filter(order_no="TOTAL-ORDER-001").count(), 1)
        self.assertEqual(order.order_quantity, 260)
        self.assertEqual(order.production_quantity, "120M")
        self.assertEqual(order.manual_received_material_kg, 7.5)
        self.assertTrue(MaterialReceipt.objects.filter(pk=receipt.pk, order=order).exists())
        self.assertEqual(order.notes, "在线手工备注")

    def test_internal_total_row_identity_allows_order_number_correction(self):
        first = self.preview(
            "total-order-number-first.xlsx",
            internal_total_workbook(order_no="TOTAL-OLD-NO"),
        )
        self.commit(first)
        corrected = self.preview(
            "total-order-number-corrected.xlsx",
            internal_total_workbook(order_no="TOTAL-NEW-NO"),
        )
        order_row = next(
            row for row in corrected["rows"] if row["record_type"] == "ORDER"
        )
        self.assertEqual(order_row["action"], "UPDATE")
        self.assertIn("order_no", order_row["changes"])
        self.commit(corrected)
        self.assertEqual(QualityOrder.objects.count(), 1)
        self.assertEqual(QualityOrder.objects.get().order_no, "TOTAL-NEW-NO")

    def test_ambiguous_internal_total_row_skips_without_creating_third_order(self):
        for index in range(2):
            QualityOrder.objects.create(
                order_no=f"AMBIGUOUS-{index}",
                specification=SYNTHETIC_SPEC,
                material=SYNTHETIC_MATERIAL,
                order_quantity=240,
                source_sheet="AMBIGUOUS-TOTAL",
                source_row=2,
                source_key=f"legacy-ambiguous-{index}",
                created_by=self.user,
            )
        preview = self.preview(
            "ambiguous-total.xlsx",
            internal_order_workbook(
                "AMBIGUOUS-CORRECTION", sheet_title="AMBIGUOUS-TOTAL"
            ),
        )
        order_row = next(
            row for row in preview["rows"] if row["record_type"] == "ORDER"
        )
        self.assertEqual(order_row["action"], "SKIP")
        self.assertTrue(
            any("多条旧订单" in issue["message"] for issue in preview["issues"])
        )
        result = self.commit(preview)
        self.assertEqual(result["skipped"]["orders"], 1)
        self.assertEqual(QualityOrder.objects.count(), 2)

    def test_internal_total_manual_received_material_warns_once_per_workbook(self):
        workbook = load_workbook(
            io.BytesIO(internal_total_workbook(manual_received_material_kg=3.5))
        )
        sheet = workbook.active
        sheet.append([cell.value for cell in sheet[2]])
        result = self.preview("total-with-manual-material.xlsx", workbook_bytes(workbook))
        warnings = [
            issue
            for issue in result["issues"]
            if issue.get("field") == "manual_received_material_kg"
            and issue["level"] == "warning"
        ]
        self.assertEqual(len(warnings), 1)
        self.assertIn("与后续导入的发料明细累计相加", warnings[0]["message"])

    def test_material_receipt_links_unique_order_and_reimport_skips(self):
        order = QualityOrder.objects.create(
            order_no=SYNTHETIC_FACTORY_ORDER,
            item_no="1",
            specification="TEST-SPEC-B",
            order_quantity=2400,
            created_by=self.user,
        )
        content = material_workbook()
        first = self.preview("material.xlsx", content)
        self.commit(first)
        receipt = MaterialReceipt.objects.get()
        self.assertEqual(receipt.order_id, order.pk)

        repeated = self.preview("material.xlsx", content)
        self.assertEqual(repeated["warning_count"], 1)
        result = self.commit(repeated)
        self.assertEqual(result["skipped"]["material_receipts"], 1)
        self.assertEqual(MaterialReceipt.objects.count(), 1)

    def test_material_batch_uses_stable_key_and_updates_changed_weight(self):
        order = QualityOrder.objects.create(
            order_no=SYNTHETIC_FACTORY_ORDER,
            item_no="1",
            specification="TEST-SPEC-B",
            order_quantity=2400,
            created_by=self.user,
        )
        first = self.preview("material-first.xlsx", material_workbook(weight=24.5))
        self.commit(first)
        changed = self.preview(
            "material-corrected.xlsx",
            material_workbook(weight=25.25, printed_at="2026-08-05 09:30:00"),
        )
        receipt_row = next(
            row for row in changed["rows"] if row["record_type"] == "MATERIAL_RECEIPT"
        )
        self.assertEqual(receipt_row["action"], "UPDATE")
        result = self.commit(changed)
        self.assertEqual(result["updated"]["material_receipts"], 1)
        receipt = MaterialReceipt.objects.get()
        self.assertEqual(receipt.order_id, order.pk)
        self.assertEqual(receipt.weight_kg, 25.25)
        self.assertEqual(MaterialReceipt.objects.count(), 1)

    def test_newer_material_print_time_updates_lineage_without_duplicate_receipt(self):
        first = self.preview(
            "material-print-time.xlsx",
            material_workbook(
                printed_at="2026-08-04 09:30:00",
                printed_label="打印时间：",
            ),
        )
        self.commit(first)
        receipt = MaterialReceipt.objects.get()
        first_batch_id = receipt.source_batch_id
        self.assertEqual(receipt.last_source_batch_id, first_batch_id)

        newer = self.preview(
            "material-newer-print-time.xlsx",
            material_workbook(
                printed_at="2026-08-05 09:30:00",
                printed_label="打印时间：",
            ),
        )
        receipt_row = next(
            row for row in newer["rows"] if row["record_type"] == "MATERIAL_RECEIPT"
        )
        self.assertEqual(receipt_row["action"], "UPDATE")
        self.assertIn("source_document_at", receipt_row["changes"])
        self.commit(newer)

        receipt.refresh_from_db()
        self.assertEqual(MaterialReceipt.objects.count(), 1)
        self.assertEqual(receipt.source_batch_id, first_batch_id)
        self.assertEqual(
            receipt.last_source_batch_id,
            BusinessImportBatch.objects.get(pk=newer["token"]).pk,
        )
        self.assertEqual(
            timezone.localtime(receipt.source_document_at).date(), date(2026, 8, 5)
        )

    def test_missing_material_source_time_cannot_overwrite_versioned_receipt(self):
        first = self.preview(
            "material-versioned.xlsx",
            material_workbook(weight=24.5, printed_at="2026-08-05 09:30:00"),
        )
        self.commit(first)
        receipt = MaterialReceipt.objects.get()

        missing_time = self.preview(
            "material-without-source-time.xlsx",
            material_workbook(weight=99, printed_at=""),
        )
        receipt_row = next(
            row
            for row in missing_time["rows"]
            if row["record_type"] == "MATERIAL_RECEIPT"
        )
        self.assertEqual(receipt_row["action"], "SKIP")
        self.assertTrue(
            any("时间缺失或早于" in issue["message"] for issue in missing_time["issues"])
        )
        self.commit(missing_time)
        receipt.refresh_from_db()
        self.assertEqual(MaterialReceipt.objects.count(), 1)
        self.assertEqual(receipt.weight_kg, 24.5)

    def test_material_receipt_claims_only_a_unique_legacy_empty_item_order(self):
        unique_order = QualityOrder.objects.create(
            order_no="LEGACY-UNIQUE",
            item_no="",
            specification="TEST-SPEC-B",
            material=SYNTHETIC_MATERIAL,
            order_quantity=2400,
            source_system="INTERNAL_TOTAL",
            created_by=self.user,
        )
        unique = self.preview(
            "material-unique-legacy.xlsx",
            material_workbook(order_no="LEGACY-UNIQUE", batch_no="LEGACY-BATCH-1"),
        )
        self.assertTrue(any("唯一匹配" in issue["message"] for issue in unique["issues"]))
        self.commit(unique)
        unique_order.refresh_from_db()
        unique_receipt = MaterialReceipt.objects.get(batch_no="LEGACY-BATCH-1")
        self.assertEqual(unique_order.item_no, "1")
        self.assertEqual(unique_receipt.order_id, unique_order.pk)

        ambiguous_orders = [
            QualityOrder.objects.create(
                order_no="LEGACY-AMBIGUOUS",
                item_no="",
                specification="TEST-SPEC-B",
                material=SYNTHETIC_MATERIAL,
                order_quantity=2400,
                source_system="INTERNAL_TOTAL",
                created_by=self.user,
            )
            for _ in range(2)
        ]
        ambiguous = self.preview(
            "material-ambiguous-legacy.xlsx",
            material_workbook(
                order_no="LEGACY-AMBIGUOUS",
                batch_no="LEGACY-BATCH-2",
            ),
        )
        self.assertTrue(any("多条可能对应" in issue["message"] for issue in ambiguous["issues"]))
        self.commit(ambiguous)
        ambiguous_receipt = MaterialReceipt.objects.get(batch_no="LEGACY-BATCH-2")
        self.assertIsNone(ambiguous_receipt.order_id)
        for order in ambiguous_orders:
            order.refresh_from_db()
            self.assertEqual(order.item_no, "")

    def test_factory_order_claims_internal_total_after_material_fills_item_number(self):
        total = self.preview(
            "2026-total-before-factory.xlsx",
            internal_total_workbook(
                order_no=SYNTHETIC_FACTORY_ORDER,
                specification="TEST-SPEC-B",
                material=SYNTHETIC_MATERIAL,
                order_quantity=2400,
                due_date=date(2026, 8, 10),
            ),
        )
        self.commit(total)
        order = QualityOrder.objects.get(order_no=SYNTHETIC_FACTORY_ORDER)
        original_order_id = order.pk
        original_source_batch_id = order.source_batch_id

        material = self.preview("material-before-factory.xlsx", material_workbook())
        self.commit(material)
        order.refresh_from_db()
        self.assertEqual(order.item_no, "1")
        self.assertEqual(MaterialReceipt.objects.get().order_id, order.pk)

        factory = self.preview("factory-after-total-and-material.xlsx", factory_workbook())
        order_row = next(row for row in factory["rows"] if row["record_type"] == "ORDER")
        self.assertEqual(order_row["action"], "UPDATE")
        self.assertEqual(order_row["match_id"], original_order_id)
        self.commit(factory)

        order.refresh_from_db()
        self.assertEqual(QualityOrder.objects.filter(order_no=SYNTHETIC_FACTORY_ORDER).count(), 1)
        self.assertEqual(order.pk, original_order_id)
        self.assertEqual(order.source_batch_id, original_source_batch_id)
        self.assertEqual(
            order.last_source_batch_id,
            BusinessImportBatch.objects.get(pk=factory["token"]).pk,
        )
        self.assertEqual(order.source_system, "NBR-T3")
        self.assertTrue(order.external_key.startswith("factory-order|"))

    def test_order_import_relinks_previous_unlinked_material_by_unique_item(self):
        material = self.preview("material-first.xlsx", material_workbook())
        self.commit(material)
        receipt = MaterialReceipt.objects.get()
        self.assertIsNone(receipt.order_id)

        order_batch = self.preview("factory-after-material.xlsx", factory_workbook())
        self.commit(order_batch)
        receipt.refresh_from_db()
        self.assertEqual(receipt.order.order_no, SYNTHETIC_FACTORY_ORDER)
        self.assertEqual(receipt.order.item_no, "1")
        self.assertGreaterEqual(
            BusinessRecordRevision.objects.filter(
                record_type=BusinessRecordRevision.RecordType.MATERIAL_RECEIPT,
                record_id=receipt.pk,
            ).count(),
            2,
        )

    def test_material_receipt_preview_warns_when_order_is_missing_or_ambiguous(self):
        missing = self.preview("material-missing.xlsx", material_workbook())
        self.assertTrue(any("未找到对应订单" in item["message"] for item in missing["issues"]))
        self.commit(missing)
        self.assertIsNone(MaterialReceipt.objects.get().order_id)

        MaterialReceipt.objects.all().delete()
        for _ in range(2):
            QualityOrder.objects.create(
                order_no=SYNTHETIC_FACTORY_ORDER,
                item_no="1",
                specification="TEST-SPEC-B",
                material=SYNTHETIC_MATERIAL,
                order_quantity=2400,
                created_by=self.user,
            )
        ambiguous_content = material_workbook("TEST-BATCH-002")
        ambiguous = self.preview("material-ambiguous.xlsx", ambiguous_content)
        self.assertTrue(any("多条可能对应" in item["message"] for item in ambiguous["issues"]))
        self.commit(ambiguous)
        receipt = MaterialReceipt.objects.order_by("-id").first()
        self.assertIsNone(receipt.order_id)

    def test_duplicate_factory_and_material_business_keys_block_preview(self):
        factory = load_workbook(
            io.BytesIO(factory_workbook(include_criteria=False))
        )
        factory_sheet = factory["sheet1"]
        factory_sheet.append([cell.value for cell in factory_sheet[4]])
        factory_preview = self.preview(
            "factory-duplicate-key.xlsx", workbook_bytes(factory)
        )
        self.assertGreater(factory_preview["error_count"], 0)
        factory_issue = next(
            issue
            for issue in factory_preview["issues"]
            if issue.get("field") == "external_key"
        )
        self.assertIn("第4行", factory_issue["message"])
        self.assertIn("第5行", factory_issue["message"])

        material = load_workbook(io.BytesIO(material_workbook()))
        material_sheet = material["sheet1"]
        duplicate_material_row = [cell.value for cell in material_sheet[4]]
        duplicate_material_row[0] = 2
        material_sheet.append(duplicate_material_row)
        material_preview = self.preview(
            "material-duplicate-key.xlsx", workbook_bytes(material)
        )
        self.assertGreater(material_preview["error_count"], 0)
        material_issue = next(
            issue
            for issue in material_preview["issues"]
            if issue.get("field") == "external_key"
        )
        self.assertIn("第4行", material_issue["message"])
        self.assertIn("第5行", material_issue["message"])

    def test_factory_and_material_require_stable_business_key_fields(self):
        factory = load_workbook(
            io.BytesIO(factory_workbook(include_criteria=False))
        )
        factory_sheet = factory["sheet1"]
        factory_sheet.cell(2, 3).value = None
        factory_sheet.cell(4, 1).value = None
        factory_sheet.cell(4, 2).value = None
        factory_sheet.cell(4, 4).value = None
        factory_preview = self.preview(
            "factory-missing-key-fields.xlsx", workbook_bytes(factory)
        )
        factory_error_fields = {
            issue.get("field")
            for issue in factory_preview["issues"]
            if issue["level"] == "error"
        }
        self.assertTrue(
            {"source_system", "order_no", "item_no", "specification"}.issubset(
                factory_error_fields
            )
        )

        material = load_workbook(io.BytesIO(material_workbook()))
        material_sheet = material["sheet1"]
        material_sheet.cell(2, 6).value = None
        material_sheet.cell(4, 2).value = None
        material_sheet.cell(4, 3).value = None
        material_sheet.cell(4, 7).value = None
        material_preview = self.preview(
            "material-missing-key-fields.xlsx", workbook_bytes(material)
        )
        material_error_fields = {
            issue.get("field")
            for issue in material_preview["issues"]
            if issue["level"] == "error"
        }
        self.assertTrue(
            {"source_system", "order_no", "item_no", "batch_no"}.issubset(
                material_error_fields
            )
        )

    def test_commit_is_transactional_when_payload_is_changed_after_preview(self):
        product = self.preview("products.xlsx", product_workbook())
        batch = BusinessImportBatch.objects.get(pk=product["token"])
        bad_order = {
            "row_key": "extra",
            "record_type": "ORDER",
            "sheet": "extra",
            "row": 99,
            "source_key": f"{batch.sha256}:extra:99:ORDER",
            "raw_data": {},
            "order_no": "BAD",
            "specification": "",
            "material": "",
            "order_quantity": 0,
            "status": "OPEN",
        }
        payload = batch.payload
        payload["rows"].append(bad_order)
        batch.payload = payload
        batch.save(update_fields=["payload"])
        with self.assertRaises(ValueError):
            commit_business_batch(batch, self.user)
        self.assertEqual(ProductSpecification.objects.count(), 0)
        self.assertEqual(QualityOrder.objects.count(), 0)
        batch.refresh_from_db()
        self.assertEqual(batch.status, BusinessImportBatch.Status.PREVIEWED)


class BusinessImportApiTests(APITestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(username="import-api", password="test")
        self.media_dir = tempfile.TemporaryDirectory()
        self.override = override_settings(MEDIA_ROOT=Path(self.media_dir.name))
        self.override.enable()

    def tearDown(self):
        self.override.disable()
        self.media_dir.cleanup()

    def test_template_preview_commit_error_report_and_original_file_backup(self):
        self.assertIn(
            self.client.get("/api/orders/product-specifications/").status_code,
            (401, 403),
        )
        self.client.force_authenticate(self.user)
        template = self.client.get(
            "/api/orders/imports/template/?type=product_specifications"
        )
        self.assertEqual(template.status_code, 200, template.content)
        order_template = self.client.get("/api/orders/imports/template/?type=orders")
        self.assertEqual(order_template.status_code, 200, order_template.content)
        workbook = load_workbook(io.BytesIO(order_template.content), read_only=True)
        self.assertIn("已发胶料", [cell.value for cell in workbook.active[1]])
        workbook.close()

        preview = self.client.post(
            "/api/orders/imports/preview/",
            {"file": upload("products.xlsx", product_workbook(literal_strip=True))},
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200, preview.content)
        payload = preview.json()
        self.assertEqual(payload["counts"]["product_specifications"], 1)
        batch = BusinessImportBatch.objects.get(pk=payload["token"])
        self.assertTrue(batch.original_file.name.startswith("business-imports/"))
        self.assertTrue(batch.original_file.storage.exists(batch.original_file.name))

        report = self.client.get(f"/api/orders/imports/{batch.pk}/errors/")
        self.assertEqual(report.status_code, 200, report.content)
        committed = self.client.post(
            "/api/orders/imports/commit/",
            {"token": payload["token"]},
            format="json",
        )
        self.assertEqual(committed.status_code, 200, committed.content)
        self.assertEqual(committed.json()["imported"]["product_specifications"], 1)

    def test_material_template_round_trip_requires_row_source_system(self):
        self.client.force_authenticate(self.user)
        template = self.client.get(
            "/api/orders/imports/template/?type=material_receipts"
        )
        self.assertEqual(template.status_code, 200, template.content)
        workbook = load_workbook(io.BytesIO(template.content))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        self.assertIn("课别", headers)
        values = {
            "序号": 1,
            "课别": "NBR-T3",
            "项次": 1,
            "独立需求号": "TEMPLATE-ORDER-001",
            "成品品名": "TEMPLATE-PRODUCT",
            "成品规格": "TEMPLATE-SPEC",
            "材质": "TEMPLATE-MATERIAL",
            "批号": "TEMPLATE-BATCH-001",
            "出片尺寸": "300/3",
            "重量": 1.25,
            "制造时间": date(2026, 8, 1),
        }
        sheet.append([values.get(header, "") for header in headers])
        preview = self.client.post(
            "/api/orders/imports/preview/",
            {"file": upload("material-template.xlsx", workbook_bytes(workbook))},
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200, preview.content)
        self.assertEqual(preview.json()["source_type"], "MATERIAL_ISSUE")
        self.assertEqual(preview.json()["error_count"], 0, preview.json()["issues"])

        blank_source = load_workbook(io.BytesIO(template.content))
        blank_sheet = blank_source.active
        blank_headers = [cell.value for cell in blank_sheet[1]]
        values["课别"] = ""
        blank_sheet.append([values.get(header, "") for header in blank_headers])
        invalid = self.client.post(
            "/api/orders/imports/preview/",
            {
                "file": upload(
                    "material-template-no-source.xlsx",
                    workbook_bytes(blank_source),
                )
            },
            format="multipart",
        )
        self.assertEqual(invalid.status_code, 200, invalid.content)
        self.assertTrue(
            any(
                issue.get("field") == "source_system"
                and issue["level"] == "error"
                for issue in invalid.json()["issues"]
            )
        )
