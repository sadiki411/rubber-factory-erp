import hashlib
import io
import math
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import PurePosixPath
from posixpath import dirname, join, normpath
from uuid import UUID
from zipfile import BadZipFile, ZipFile
from xml.etree import ElementTree

from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.files.base import ContentFile
from django.db import IntegrityError, OperationalError, transaction
from django.db.models import Max, Q
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.numbers import BUILTIN_FORMATS, is_date_format
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import CALENDAR_MAC_1904, CALENDAR_WINDOWS_1900, from_excel

from quality.models import QualityOrder

from .models import (
    BusinessImportBatch,
    BusinessRecordRevision,
    MaterialReceipt,
    ProductInspectionCriterion,
    ProductSpecification,
    normalize_product_key,
)
from .services import json_safe, model_snapshot, record_revision


MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_ARCHIVE_FILES = 1000
MAX_WORKSHEETS = 100
MAX_ROWS_PER_SHEET = 5000
MAX_COLUMNS_PER_SHEET = 100
MAX_WORKBOOK_CELLS = 300_000

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
NS = {"m": MAIN_NS, "r": REL_NS, "pr": PACKAGE_REL_NS}

RECORD_PRODUCT = "PRODUCT_SPECIFICATION"
RECORD_ORDER = "ORDER"
RECORD_RECEIPT = "MATERIAL_RECEIPT"
RECORD_CRITERION = "INSPECTION_CRITERION"
RECORD_TYPES = (RECORD_PRODUCT, RECORD_ORDER, RECORD_RECEIPT, RECORD_CRITERION)
INTERNAL_SOURCE_SYSTEM = "INTERNAL_TOTAL"


@dataclass
class CellData:
    raw_value: object = None
    display_text: str = ""
    number_format: str = "General"

    def raw_payload(self):
        return {
            "raw_value": _json_cell_value(self.raw_value),
            "display_text": self.display_text,
            "number_format": self.number_format,
        }


@dataclass
class SheetData:
    name: str
    rows: list


def _json_cell_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _issue(level, message, *, sheet="", row=None, field=""):
    return {
        "level": level,
        "sheet": sheet,
        "row": row,
        "field": field,
        "message": message,
    }


def _clean_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _literal_number_format(number_format):
    text = str(number_format or "").split(";", 1)[0].strip()
    if not text or text.lower() == "general" or is_date_format(text):
        return ""
    if re.search(r"[0#?]", text):
        return ""
    text = re.sub(r"\[[^\]]+\]", "", text)
    text = re.sub(r"_.", "", text)
    text = re.sub(r"\*.", "", text)
    text = re.sub(r'"([^"]*)"', r"\1", text)
    text = re.sub(r"\\(.)", r"\1", text)
    return text.strip()


def _display_value(value, number_format="General", *, date_cell=False):
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time().replace(microsecond=0) == datetime.min.time():
            return value.date().isoformat()
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    literal = _literal_number_format(number_format)
    if literal and isinstance(value, (int, float, Decimal)) and not date_cell:
        return literal
    return _clean_text(value)


def _validate_archive(data):
    if len(data) > MAX_UPLOAD_BYTES:
        raise ValueError(f"Excel文件不能超过{MAX_UPLOAD_BYTES // 1024 // 1024}MB。")
    try:
        with ZipFile(io.BytesIO(data)) as archive:
            members = archive.infolist()
            if len(members) > MAX_ARCHIVE_FILES:
                raise ValueError("Excel压缩包包含的文件数量过多。")
            if sum(item.file_size for item in members) > MAX_UNCOMPRESSED_BYTES:
                raise ValueError("Excel解压后的数据量过大。")
            for item in members:
                path = PurePosixPath(item.filename)
                if path.is_absolute() or ".." in path.parts:
                    raise ValueError("Excel包含不安全的压缩包路径。")
    except BadZipFile as exc:
        raise ValueError("文件不是有效的.xlsx工作簿。") from exc


def _validate_dimensions(sheets):
    if len(sheets) > MAX_WORKSHEETS:
        raise ValueError(f"工作表不能超过{MAX_WORKSHEETS}个。")
    total = 0
    for sheet in sheets:
        if len(sheet.rows) > MAX_ROWS_PER_SHEET:
            raise ValueError(f"工作表“{sheet.name}”不能超过{MAX_ROWS_PER_SHEET}行。")
        max_columns = max((len(cells) for _, cells in sheet.rows), default=0)
        if max_columns > MAX_COLUMNS_PER_SHEET:
            raise ValueError(f"工作表“{sheet.name}”不能超过{MAX_COLUMNS_PER_SHEET}列。")
        total += len(sheet.rows) * max_columns
        if total > MAX_WORKBOOK_CELLS:
            raise ValueError("工作簿单元格数量过多。")


def _openpyxl_sheets(data):
    workbook = load_workbook(io.BytesIO(data), read_only=False, data_only=False)
    try:
        sheets = []
        for worksheet in workbook.worksheets:
            rows = []
            for row_no, row in enumerate(worksheet.iter_rows(), start=1):
                cells = []
                last_nonempty = -1
                for index, cell in enumerate(row):
                    value = getattr(cell, "value", None)
                    date_cell = bool(getattr(cell, "is_date", False))
                    number_format = getattr(cell, "number_format", "General") or "General"
                    cells.append(
                        CellData(
                            raw_value=value,
                            display_text=_display_value(
                                value, number_format, date_cell=date_cell
                            ),
                            number_format=number_format,
                        )
                    )
                    if value not in (None, "") or cells[-1].display_text:
                        last_nonempty = index
                if last_nonempty >= 0:
                    rows.append((row_no, cells[: last_nonempty + 1]))
            sheets.append(SheetData(worksheet.title, rows))
        return sheets
    finally:
        workbook.close()


def _relationship_target(base_path, target):
    target = target.replace("\\", "/")
    if target.startswith("/"):
        return target.lstrip("/")
    return normpath(join(dirname(base_path), target))


def _parse_styles(archive):
    formats = dict(BUILTIN_FORMATS)
    style_formats = []
    if "xl/styles.xml" not in archive.namelist():
        return formats, style_formats
    root = ElementTree.fromstring(archive.read("xl/styles.xml"))
    num_fmts = root.find("m:numFmts", NS)
    if num_fmts is not None:
        for item in num_fmts:
            try:
                formats[int(item.attrib["numFmtId"])] = item.attrib.get("formatCode", "General")
            except (KeyError, ValueError):
                continue
    cell_xfs = root.find("m:cellXfs", NS)
    if cell_xfs is not None:
        for xf in cell_xfs:
            try:
                num_fmt_id = int(xf.attrib.get("numFmtId", "0"))
            except ValueError:
                num_fmt_id = 0
            style_formats.append(formats.get(num_fmt_id, "General"))
    return formats, style_formats


def _xml_cell_value(cell, shared_strings, style_formats, epoch):
    cell_type = cell.attrib.get("t", "n")
    style_id = int(cell.attrib.get("s", "0") or 0)
    number_format = (
        style_formats[style_id] if 0 <= style_id < len(style_formats) else "General"
    )
    value_node = cell.find("m:v", NS)
    formula_node = cell.find("m:f", NS)
    raw_text = value_node.text if value_node is not None else None
    raw_value = None
    if cell_type == "s" and raw_text not in (None, ""):
        raw_value = shared_strings[int(raw_text)]
    elif cell_type == "inlineStr":
        inline = cell.find("m:is", NS)
        raw_value = (
            "".join(item.text or "" for item in inline.iter(f"{{{MAIN_NS}}}t"))
            if inline is not None
            else ""
        )
    elif cell_type in {"str", "e"}:
        raw_value = raw_text or ""
    elif cell_type == "b":
        raw_value = raw_text == "1"
    elif cell_type == "d":
        raw_value = parse_datetime(raw_text or "") or parse_date(raw_text or "") or raw_text
    elif raw_text not in (None, ""):
        try:
            numeric = float(raw_text)
            raw_value = int(numeric) if numeric.is_integer() else numeric
        except ValueError:
            raw_value = raw_text
    elif formula_node is not None:
        raw_value = f"={formula_node.text or ''}"

    date_cell = bool(isinstance(raw_value, (int, float)) and is_date_format(number_format))
    display_value = raw_value
    if date_cell:
        try:
            display_value = from_excel(raw_value, epoch=epoch)
        except (ValueError, TypeError, OverflowError):
            display_value = raw_value
            date_cell = False
    return CellData(
        raw_value=raw_value,
        display_text=_display_value(display_value, number_format, date_cell=date_cell),
        number_format=number_format,
    )


def _ooxml_sheets(data):
    with ZipFile(io.BytesIO(data)) as archive:
        shared_strings = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in shared_root.findall("m:si", NS):
                shared_strings.append(
                    "".join(node.text or "" for node in item.iter(f"{{{MAIN_NS}}}t"))
                )
        _, style_formats = _parse_styles(archive)
        workbook_path = "xl/workbook.xml"
        workbook_root = ElementTree.fromstring(archive.read(workbook_path))
        properties = workbook_root.find("m:workbookPr", NS)
        date_1904 = properties is not None and properties.attrib.get("date1904") in {
            "1",
            "true",
            "True",
        }
        epoch = CALENDAR_MAC_1904 if date_1904 else CALENDAR_WINDOWS_1900
        relationship_path = "xl/_rels/workbook.xml.rels"
        relationships = ElementTree.fromstring(archive.read(relationship_path))
        targets = {
            item.attrib["Id"]: _relationship_target(workbook_path, item.attrib["Target"])
            for item in relationships
        }
        sheets = []
        for sheet in workbook_root.find("m:sheets", NS):
            name = sheet.attrib.get("name", "Sheet")
            target = targets[sheet.attrib[f"{{{REL_NS}}}id"]]
            root = ElementTree.fromstring(archive.read(target))
            rows = []
            for row in root.findall(".//m:sheetData/m:row", NS):
                row_no = int(row.attrib.get("r", "0") or 0)
                cells_by_index = {}
                max_index = -1
                for cell in row.findall("m:c", NS):
                    reference = cell.attrib.get("r", "A1")
                    letters = re.match(r"[A-Z]+", reference)
                    if not letters:
                        continue
                    index = 0
                    for letter in letters.group(0):
                        index = index * 26 + ord(letter) - 64
                    index -= 1
                    cells_by_index[index] = _xml_cell_value(
                        cell, shared_strings, style_formats, epoch
                    )
                    max_index = max(max_index, index)
                if max_index >= 0:
                    rows.append(
                        (
                            row_no,
                            [cells_by_index.get(index, CellData()) for index in range(max_index + 1)],
                        )
                    )
            sheets.append(SheetData(name, rows))
        return sheets


def read_business_workbook(data):
    _validate_archive(data)
    try:
        sheets = _openpyxl_sheets(data)
        parser = "openpyxl-3.1"
    except (TypeError, ValueError, KeyError, IndexError, AttributeError, OSError):
        sheets = _ooxml_sheets(data)
        parser = "safe-ooxml-1"
    _validate_dimensions(sheets)
    return sheets, parser


def _cell(cells, index):
    return cells[index] if 0 <= index < len(cells) else CellData()


def _normalized_header(value):
    return re.sub(r"[\s（）()：:]", "", str(value or "")).casefold()


def _header_map(cells):
    result = {}
    for index, cell in enumerate(cells):
        key = _normalized_header(cell.display_text)
        if key:
            result[key] = index
    return result


def _find_header(sheet, required, max_scan=20):
    normalized_required = {_normalized_header(item) for item in required}
    for row_no, cells in sheet.rows[:max_scan]:
        mapping = _header_map(cells)
        if normalized_required.issubset(mapping):
            return row_no, mapping
    return None, None


def _sheet_row_iter(sheet, header_row):
    for row_no, cells in sheet.rows:
        if row_no > header_row:
            yield row_no, cells


def _mapped_cell(cells, mapping, *names):
    for name in names:
        index = mapping.get(_normalized_header(name))
        if index is not None:
            return _cell(cells, index)
    return CellData()


def _mapping_has(mapping, *names):
    return any(_normalized_header(name) in mapping for name in names)


def _raw_row(cells, mapping):
    result = {}
    for header, index in mapping.items():
        result[header] = _cell(cells, index).raw_payload()
    return result


def _date_value(cell):
    value = cell.raw_value
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and 1 <= value <= 100_000:
        # Some legacy workbooks have a damaged stylesheet and store date cells
        # as General. Date parsing is only called for known date columns, so a
        # plausible Excel serial is safe to recover here.
        try:
            converted = from_excel(value, epoch=CALENDAR_WINDOWS_1900)
            return converted.date().isoformat() if isinstance(converted, datetime) else converted.isoformat()
        except (ValueError, TypeError, OverflowError):
            pass
    text = cell.display_text.strip()
    parsed = parse_date(text)
    if parsed:
        return parsed.isoformat()
    for fmt in ("%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _datetime_value(cell):
    value = cell.raw_value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).isoformat()
    text = cell.display_text.strip()
    parsed = parse_datetime(text)
    if parsed:
        return parsed.isoformat()
    parsed_date = parse_date(text)
    if parsed_date:
        return datetime.combine(parsed_date, datetime.min.time()).isoformat()
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).isoformat()
        except ValueError:
            continue
    return None


def _decimal_text(cell, issues, *, sheet, row, field, required=False):
    text = cell.display_text.strip()
    if not text:
        if required:
            issues.append(_issue("error", f"{field}不能为空。", sheet=sheet, row=row, field=field))
        return None
    try:
        value = Decimal(text.replace(",", ""))
    except InvalidOperation:
        issues.append(_issue("error", f"{field}不是有效数字：{text}", sheet=sheet, row=row, field=field))
        return None
    if value < 0:
        issues.append(_issue("error", f"{field}不能小于0。", sheet=sheet, row=row, field=field))
        return None
    return format(value, "f")


def _positive_integer(cell, issues, *, sheet, row, field):
    text = cell.display_text.replace(",", "").strip()
    try:
        value = Decimal(text)
    except InvalidOperation:
        issues.append(_issue("error", f"{field}不是有效整数：{text}", sheet=sheet, row=row, field=field))
        return None
    if value != value.to_integral_value() or value < 1:
        issues.append(_issue("error", f"{field}必须是大于0的整数。", sheet=sheet, row=row, field=field))
        return None
    return int(value)


def _boolean_value(cell):
    text = cell.display_text.strip().casefold()
    if text in {"是", "yes", "y", "true", "1", "生产", "完成"}:
        return True
    if text in {"否", "no", "n", "false", "0", "不生产"}:
        return False
    return None


def _source_key(sha256, sheet, row, record_type):
    safe_sheet = re.sub(r"\s+", " ", sheet).strip()
    return f"{sha256}:{safe_sheet}:{row}:{record_type}"


def _stable_component(value):
    return " ".join(str(value or "").split()).casefold()


def _external_key(kind, *values):
    components = [_stable_component(value) for value in values]
    if not all(components):
        return ""
    value = "|".join((kind, *components))
    if len(value) <= 500:
        return value
    return f"{kind}|sha256:{hashlib.sha256(value.encode('utf-8')).hexdigest()}"


def _record_base(sha256, sheet, row, record_type, raw_data):
    return {
        "row_key": f"{sheet}:{row}:{record_type}",
        "record_type": record_type,
        "sheet": sheet,
        "row": row,
        "source_key": _source_key(sha256, sheet, row, record_type),
        "raw_data": raw_data,
    }


def _parse_product_specifications(sheet, header_row, mapping, sha256, issues):
    records = []
    for row_no, cells in _sheet_row_iter(sheet, header_row):
        important = [
            _mapped_cell(cells, mapping, "规格"),
            _mapped_cell(cells, mapping, "材质"),
            _mapped_cell(cells, mapping, "料长"),
            _mapped_cell(cells, mapping, "切料重"),
            _mapped_cell(cells, mapping, "一次加硫条件"),
        ]
        if not any(cell.display_text for cell in important):
            continue
        record = _record_base(
            sha256, sheet.name, row_no, RECORD_PRODUCT, _raw_row(cells, mapping)
        )
        record.update(
            {
                "product_name": _mapped_cell(cells, mapping, "产品名称").display_text,
                "customer_product_no": _mapped_cell(
                    cells, mapping, "客户产品号", "产品编号"
                ).display_text,
                "specification": _mapped_cell(cells, mapping, "规格").display_text,
                "material": _mapped_cell(cells, mapping, "材质").display_text,
                "material_length": _mapped_cell(cells, mapping, "料长").display_text,
                "cut_weight": _mapped_cell(cells, mapping, "切料重").display_text,
                "strip_count": _mapped_cell(cells, mapping, "条数").display_text,
                "primary_curing": _mapped_cell(cells, mapping, "一次加硫条件").display_text,
                "secondary_curing": _mapped_cell(
                    cells, mapping, "二烤条件", "二次加硫条件"
                ).display_text,
                "total_cavities": _mapped_cell(cells, mapping, "总孔数").display_text,
                "effective_cavities": _mapped_cell(cells, mapping, "有效孔数").display_text,
                "mold_in_stock": _mapped_cell(cells, mapping, "模具在库").display_text,
                "mold_no": _mapped_cell(cells, mapping, "模具号").display_text,
                "mold_size": _mapped_cell(cells, mapping, "模具尺寸").display_text,
                "standard_hours": _mapped_cell(cells, mapping, "标准工时").display_text,
                "notes": _mapped_cell(cells, mapping, "备注").display_text,
            }
        )
        if not any(
            (record["product_name"], record["customer_product_no"], record["specification"])
        ):
            issues.append(
                _issue(
                    "error",
                    "产品名称、客户产品号和规格至少填写一项。",
                    sheet=sheet.name,
                    row=row_no,
                    field="specification",
                )
            )
        record["normalized_key"] = normalize_product_key(
            record["product_name"],
            record["customer_product_no"],
            record["specification"],
            record["material"],
            record["mold_no"],
        )
        record["source_fields"] = [
            "product_name",
            "customer_product_no",
            "specification",
            "material",
            "material_length",
            "cut_weight",
            "strip_count",
            "primary_curing",
            "secondary_curing",
            "total_cavities",
            "effective_cavities",
            "mold_in_stock",
            "mold_no",
            "mold_size",
            "standard_hours",
            "notes",
        ]
        records.append(record)
    return records


def _parse_internal_orders(sheets, sha256, issues):
    records = []
    manual_received_warning_added = False
    for sheet in sheets:
        header_row, mapping = _find_header(
            sheet, ["订单编号", "规格", "胶料配方", "交期", "订单量"]
        )
        if not mapping:
            continue
        for row_no, cells in _sheet_row_iter(sheet, header_row):
            core_names = (
                "订单编号",
                "规格",
                "胶料配方",
                "交期",
                "订单量",
                "成型工时",
                "下单时间",
                "模具尺寸",
                "出货日期",
            )
            if not any(_mapped_cell(cells, mapping, name).display_text for name in core_names):
                continue
            order_no = _mapped_cell(cells, mapping, "订单编号").display_text
            specification = _mapped_cell(cells, mapping, "规格").display_text
            quantity = _positive_integer(
                _mapped_cell(cells, mapping, "订单量"),
                issues,
                sheet=sheet.name,
                row=row_no,
                field="order_quantity",
            )
            if not order_no:
                issues.append(_issue("error", "订单号不能为空。", sheet=sheet.name, row=row_no, field="order_no"))
            if not specification:
                issues.append(_issue("error", "规格不能为空。", sheet=sheet.name, row=row_no, field="specification"))
            due_cell = _mapped_cell(cells, mapping, "交期")
            order_date_cell = _mapped_cell(cells, mapping, "下单时间")
            completion_cell = _mapped_cell(cells, mapping, "完成")
            completion_value = _boolean_value(completion_cell)
            manual_received_cell = _mapped_cell(cells, mapping, "已发胶料")
            record = _record_base(
                sha256, sheet.name, row_no, RECORD_ORDER, _raw_row(cells, mapping)
            )
            record.update(
                {
                    "order_no": order_no,
                    "item_no": _mapped_cell(cells, mapping, "项次").display_text,
                    "batch_no": "",
                    "product_code": "",
                    "product_name": _mapped_cell(cells, mapping, "产品名称").display_text,
                    "specification": specification,
                    "material": _mapped_cell(cells, mapping, "胶料配方", "材质").display_text,
                    "order_quantity": quantity,
                    "order_date": _date_value(order_date_cell),
                    "due_date": _date_value(due_cell),
                    "mold_size": _mapped_cell(cells, mapping, "模具尺寸").display_text,
                    "forming_hours": _decimal_text(
                        _mapped_cell(cells, mapping, "成型工时"),
                        issues,
                        sheet=sheet.name,
                        row=row_no,
                        field="forming_hours",
                    )
                    if _mapped_cell(cells, mapping, "成型工时").display_text
                    else None,
                    "production_required": _boolean_value(
                        _mapped_cell(cells, mapping, "是否生产", "完成")
                    ),
                    "legacy_shipment_text": _mapped_cell(
                        cells, mapping, "出货日期", "出货信息"
                    ).display_text,
                    "required_material_kg": _decimal_text(
                        _mapped_cell(cells, mapping, "所需胶料", "胶料用量", "胶料用量KG"),
                        issues,
                        sheet=sheet.name,
                        row=row_no,
                        field="required_material_kg",
                    )
                    if _mapped_cell(cells, mapping, "所需胶料", "胶料用量", "胶料用量KG").display_text
                    else None,
                    "manual_received_material_kg": _decimal_text(
                        manual_received_cell,
                        issues,
                        sheet=sheet.name,
                        row=row_no,
                        field="manual_received_material_kg",
                    )
                    if manual_received_cell.display_text
                    else None,
                    "process_card_text": _mapped_cell(cells, mapping, "流程卡").display_text,
                    "production_quantity": _mapped_cell(
                        cells, mapping, "生产数量"
                    ).display_text,
                    "shipment_date": _mapped_cell(cells, mapping, "出货日期").display_text,
                    "shipped_quantity": _mapped_cell(cells, mapping, "出货数量").display_text,
                    "source_system": INTERNAL_SOURCE_SYSTEM,
                    "source_document_at": None,
                    "external_key": _external_key(
                        "internal-order",
                        INTERNAL_SOURCE_SYSTEM,
                        sheet.name,
                        row_no,
                    ),
                    "status": (
                        QualityOrder.Status.COMPLETED
                        if completion_value is True
                        else QualityOrder.Status.OPEN
                    ),
                    "notes": "",
                }
            )
            record["source_fields"] = [
                "order_no",
                "specification",
                "material",
                "order_quantity",
                "due_date",
            ]
            optional_source_fields = [
                ("item_no", ("项次",)),
                ("product_name", ("产品名称",)),
                ("order_date", ("下单时间",)),
                ("mold_size", ("模具尺寸",)),
                ("forming_hours", ("成型工时",)),
                ("production_required", ("是否生产", "完成")),
                ("legacy_shipment_text", ("出货日期", "出货信息")),
                ("required_material_kg", ("所需胶料", "胶料用量", "胶料用量KG")),
                ("process_card_text", ("流程卡",)),
                ("production_quantity", ("生产数量",)),
                ("shipment_date", ("出货日期",)),
                ("shipped_quantity", ("出货数量",)),
            ]
            record["source_fields"].extend(
                field
                for field, aliases in optional_source_fields
                if _mapping_has(mapping, *aliases)
            )
            # 空白“已发胶料”表示总表没有提供本次值，不能清空线上手工值；
            # 发料清单累计值来自独立收料记录，本身也不会被订单导入覆盖。
            if manual_received_cell.display_text:
                record["source_fields"].append("manual_received_material_kg")
                if not manual_received_warning_added:
                    issues.append(
                        _issue(
                            "warning",
                            "总表“已发胶料”将作为手工补充量，并与后续导入的发料明细累计相加；若同批胶料已包含在该数值中，请先清空此列或导入后在线校正，避免重复计算。",
                            sheet=sheet.name,
                            row=row_no,
                            field="manual_received_material_kg",
                        )
                    )
                    manual_received_warning_added = True
            if _normalized_header("完成") in mapping:
                record["source_fields"].append("status")
            records.append(record)
    return records


def _link_existing_product_specifications(records, issues):
    candidates = {}
    for product in ProductSpecification.objects.filter(is_active=True).only(
        "id", "specification", "material"
    ):
        key = (
            " ".join(str(product.specification or "").split()).casefold(),
            " ".join(str(product.material or "").split()).casefold(),
        )
        candidates.setdefault(key, []).append(product.pk)
    for record in records:
        if record.get("record_type") != RECORD_ORDER:
            continue
        key = (
            " ".join(str(record.get("specification") or "").split()).casefold(),
            " ".join(str(record.get("material") or "").split()).casefold(),
        )
        matches = candidates.get(key, [])
        if len(matches) == 1:
            record["product_spec_id"] = matches[0]
        elif len(matches) > 1:
            issues.append(
                _issue(
                    "warning",
                    "存在多条相同规格和材质的产品规格资料，订单暂不自动关联，请在线确认。",
                    sheet=record["sheet"],
                    row=record["row"],
                    field="product_specification",
                )
            )


def _metadata_date(sheet, label):
    normalized = _normalized_header(label)
    for _, cells in sheet.rows[:10]:
        for index, cell in enumerate(cells):
            if normalized in _normalized_header(cell.display_text):
                for candidate in cells[index + 1 :]:
                    if candidate.display_text:
                        return _date_value(candidate)
    return None


def _metadata_text(sheet, label):
    normalized = _normalized_header(label)
    for _, cells in sheet.rows[:10]:
        for index, cell in enumerate(cells):
            if normalized in _normalized_header(cell.display_text):
                for candidate in cells[index + 1 : index + 4]:
                    if candidate.display_text:
                        value = candidate.display_text.strip()
                        if value.endswith(("：", ":")):
                            return ""
                        return value
    return ""


def _metadata_datetime(sheet, *labels):
    normalized_labels = {_normalized_header(label) for label in labels}
    for _, cells in sheet.rows[:10]:
        for index, cell in enumerate(cells):
            header = _normalized_header(cell.display_text)
            if any(normalized in header for normalized in normalized_labels):
                for candidate in cells[index + 1 :]:
                    if candidate.display_text:
                        return _datetime_value(candidate)
    return None


FACTORY_PRODUCT_COMPARE_FIELDS = (
    "product_name",
    "customer_product_no",
    "specification",
    "material",
    "material_length",
    "cut_weight",
    "strip_count",
    "primary_curing",
    "secondary_curing",
    "total_cavities",
    "effective_cavities",
    "mold_in_stock",
    "mold_no",
    "mold_size",
    "standard_hours",
    "notes",
)


def _factory_product_signature(record):
    return tuple(
        str(record.get(field, "") or "").strip()
        for field in FACTORY_PRODUCT_COMPARE_FIELDS
    )


def _deduplicate_factory_products(records, issues):
    canonical_by_key = {}
    duplicate_to_canonical = {}
    duplicate_source_keys = set()
    for record in records:
        if record["record_type"] != RECORD_PRODUCT:
            continue
        normalized_key = record.get("normalized_key", "")
        canonical = canonical_by_key.get(normalized_key)
        if canonical is None:
            canonical_by_key[normalized_key] = record
            continue
        if _factory_product_signature(canonical) != _factory_product_signature(record):
            issues.append(
                _issue(
                    "error",
                    f"同一文件中产品身份重复但工艺字段冲突；首次位于工作表“{canonical['sheet']}”第{canonical['row']}行，当前冲突位于第{record['row']}行。",
                    sheet=record["sheet"],
                    row=record["row"],
                    field="product_specification",
                )
            )
            continue
        duplicate_to_canonical[record["source_key"]] = canonical["source_key"]
        duplicate_source_keys.add(record["source_key"])

    if not duplicate_to_canonical:
        return
    for record in records:
        source_key = record.get("product_spec_source_key")
        if source_key in duplicate_to_canonical:
            record["product_spec_source_key"] = duplicate_to_canonical[source_key]
    records[:] = [
        record
        for record in records
        if not (
            record["record_type"] == RECORD_PRODUCT
            and record["source_key"] in duplicate_source_keys
        )
    ]


def _factory_criterion_key(record):
    order_key = record.get("order_source_key") or _external_key(
        "criterion-order", record.get("order_no"), record.get("item_no")
    )
    return tuple(
        _stable_component(value)
        for value in (
            order_key,
            record.get("item_no"),
            record.get("project_no"),
            record.get("category"),
            record.get("version"),
            record.get("inspection_item"),
            record.get("unit"),
        )
    )


def _factory_criterion_signature(record):
    return tuple(
        str(record.get(field, "") or "").strip()
        for field in ("customer", "lower_limit", "upper_limit")
    )


def _deduplicate_factory_criteria(records, issues):
    canonical_by_key = {}
    duplicate_source_keys = set()
    for record in records:
        if record["record_type"] != RECORD_CRITERION:
            continue
        key = _factory_criterion_key(record)
        canonical = canonical_by_key.get(key)
        if canonical is None:
            canonical_by_key[key] = record
            continue
        if _factory_criterion_signature(canonical) != _factory_criterion_signature(
            record
        ):
            issues.append(
                _issue(
                    "error",
                    f"同一文件中检验标准业务键重复但上下限或客户等字段冲突；首次位于工作表“{canonical['sheet']}”第{canonical['row']}行，当前冲突位于第{record['row']}行。",
                    sheet=record["sheet"],
                    row=record["row"],
                    field="inspection_criterion",
                )
            )
            continue
        duplicate_source_keys.add(record["source_key"])
    if duplicate_source_keys:
        records[:] = [
            record
            for record in records
            if not (
                record["record_type"] == RECORD_CRITERION
                and record["source_key"] in duplicate_source_keys
            )
        ]


def _parse_factory_work_contact(sheets, sha256, issues):
    main_sheet = None
    main_header = None
    main_mapping = None
    criteria_sheet = None
    criteria_header = None
    criteria_mapping = None
    for sheet in sheets:
        header_row, mapping = _find_header(
            sheet, ["独立需求号", "项次", "材质", "规格", "订单量"]
        )
        if mapping and "检验项目" not in mapping:
            main_sheet, main_header, main_mapping = sheet, header_row, mapping
        header_row, mapping = _find_header(
            sheet, ["独立需求号", "项次", "项目号", "检验项目", "下限", "上限"]
        )
        if mapping:
            criteria_sheet, criteria_header, criteria_mapping = sheet, header_row, mapping
    if main_sheet is None:
        raise ValueError("生产工作联络单缺少订单工作表。")

    records = []
    product_keys = {}
    products_by_source_key = {}
    order_keys = {}
    order_date_value = _metadata_date(main_sheet, "发单时间")
    source_system = _metadata_text(main_sheet, "协力商")
    source_document_at = _metadata_datetime(main_sheet, "发单时间")
    for row_no, cells in _sheet_row_iter(main_sheet, main_header):
        order_no = _mapped_cell(cells, main_mapping, "独立需求号").display_text
        specification = _mapped_cell(cells, main_mapping, "规格").display_text
        if not any((order_no, specification, _mapped_cell(cells, main_mapping, "订单量").display_text)):
            continue
        item_no = _mapped_cell(cells, main_mapping, "项次").display_text
        product_record = _record_base(
            sha256,
            main_sheet.name,
            row_no,
            RECORD_PRODUCT,
            _raw_row(cells, main_mapping),
        )
        product_record.update(
            {
                "product_name": "",
                "customer_product_no": "",
                "specification": specification,
                "material": _mapped_cell(cells, main_mapping, "材质").display_text,
                "material_length": _mapped_cell(cells, main_mapping, "料长").display_text,
                "cut_weight": _mapped_cell(cells, main_mapping, "切料重").display_text,
                "strip_count": "",
                "primary_curing": _mapped_cell(cells, main_mapping, "一次加硫条件").display_text,
                "secondary_curing": _mapped_cell(cells, main_mapping, "二次加硫条件").display_text,
                "total_cavities": "",
                "effective_cavities": "",
                "mold_in_stock": "",
                "mold_no": _mapped_cell(cells, main_mapping, "模具号").display_text,
                "mold_size": _mapped_cell(cells, main_mapping, "模具尺寸").display_text,
                "standard_hours": _mapped_cell(cells, main_mapping, "参考工时", "标准工时").display_text,
                "notes": "",
                "source_system": source_system,
                "source_document_at": source_document_at,
            }
        )
        product_record["normalized_key"] = normalize_product_key(
            "",
            "",
            specification,
            product_record["material"],
            product_record["mold_no"],
        )
        product_record["source_fields"] = [
            "customer_product_no",
            "specification",
            "material",
            "material_length",
            "cut_weight",
            "primary_curing",
            "secondary_curing",
            "mold_no",
            "mold_size",
            "standard_hours",
        ]
        records.append(product_record)
        products_by_source_key[product_record["source_key"]] = product_record
        product_keys[(order_no.casefold(), item_no.casefold())] = product_record["source_key"]

        quantity = _positive_integer(
            _mapped_cell(cells, main_mapping, "订单量"),
            issues,
            sheet=main_sheet.name,
            row=row_no,
            field="order_quantity",
        )
        order_record = _record_base(
            sha256, main_sheet.name, row_no, RECORD_ORDER, _raw_row(cells, main_mapping)
        )
        order_record.update(
            {
                "order_no": order_no,
                "item_no": item_no,
                "batch_no": "",
                "product_code": "",
                "product_name": "",
                "specification": specification,
                "material": product_record["material"],
                "order_quantity": quantity,
                "order_date": order_date_value,
                "due_date": _date_value(_mapped_cell(cells, main_mapping, "完成日", "交期")),
                "mold_size": product_record["mold_size"],
                "forming_hours": _decimal_text(
                    _mapped_cell(cells, main_mapping, "参考工时", "标准工时"),
                    issues,
                    sheet=main_sheet.name,
                    row=row_no,
                    field="forming_hours",
                ),
                "production_required": True,
                "legacy_shipment_text": "",
                "required_material_kg": _decimal_text(
                    _mapped_cell(cells, main_mapping, "胶料用量KG", "胶料用量（KG）"),
                    issues,
                    sheet=main_sheet.name,
                    row=row_no,
                    field="required_material_kg",
                ),
                "status": QualityOrder.Status.OPEN,
                "notes": "",
                "product_spec_source_key": product_record["source_key"],
                "source_system": source_system,
                "source_document_at": source_document_at,
                "external_key": _external_key(
                    "factory-order", source_system, order_no, item_no
                ),
            }
        )
        order_record["source_fields"] = [
            "order_no",
            "item_no",
            "specification",
            "material",
            "order_quantity",
            "order_date",
            "due_date",
            "mold_size",
            "forming_hours",
            "required_material_kg",
        ]
        if not order_no:
            issues.append(_issue("error", "订单号不能为空。", sheet=main_sheet.name, row=row_no, field="order_no"))
        if not source_system:
            issues.append(
                _issue(
                    "error",
                    "协力商不能为空。",
                    sheet=main_sheet.name,
                    row=row_no,
                    field="source_system",
                )
            )
        if not item_no:
            issues.append(
                _issue(
                    "error",
                    "项次不能为空。",
                    sheet=main_sheet.name,
                    row=row_no,
                    field="item_no",
                )
            )
        if not specification:
            issues.append(
                _issue(
                    "error",
                    "规格不能为空。",
                    sheet=main_sheet.name,
                    row=row_no,
                    field="specification",
                )
            )
        records.append(order_record)
        order_keys[(order_no.casefold(), item_no.casefold())] = order_record["source_key"]

    if criteria_sheet and criteria_mapping:
        for row_no, cells in _sheet_row_iter(criteria_sheet, criteria_header):
            order_no = _mapped_cell(cells, criteria_mapping, "独立需求号").display_text
            item_no = _mapped_cell(cells, criteria_mapping, "项次").display_text
            inspection_item = _mapped_cell(cells, criteria_mapping, "检验项目").display_text
            if not any((order_no, item_no, inspection_item)):
                continue
            link_key = (order_no.casefold(), item_no.casefold())
            product_source_key = product_keys.get(link_key)
            if not product_source_key:
                issues.append(
                    _issue(
                        "error",
                        "检验标准找不到对应的订单产品规格。",
                        sheet=criteria_sheet.name,
                        row=row_no,
                        field="item_no",
                    )
                )
            project_no = _mapped_cell(cells, criteria_mapping, "项目号").display_text
            linked_product = products_by_source_key.get(product_source_key)
            if linked_product is not None and project_no and not linked_product["customer_product_no"]:
                linked_product["customer_product_no"] = project_no
                linked_product["normalized_key"] = normalize_product_key(
                    linked_product["product_name"],
                    linked_product["customer_product_no"],
                    linked_product["specification"],
                    linked_product["material"],
                    linked_product["mold_no"],
                )
            record = _record_base(
                sha256,
                criteria_sheet.name,
                row_no,
                RECORD_CRITERION,
                _raw_row(cells, criteria_mapping),
            )
            record.update(
                {
                    "product_spec_source_key": product_source_key,
                    "order_source_key": order_keys.get(link_key),
                    "item_no": item_no,
                    "customer": _mapped_cell(cells, criteria_mapping, "客户").display_text,
                    "category": _mapped_cell(cells, criteria_mapping, "类别").display_text,
                    "version": _mapped_cell(cells, criteria_mapping, "版本").display_text,
                    "inspection_item": inspection_item,
                    "lower_limit": _mapped_cell(cells, criteria_mapping, "下限").display_text,
                    "upper_limit": _mapped_cell(cells, criteria_mapping, "上限").display_text,
                    "unit": _mapped_cell(cells, criteria_mapping, "单位").display_text,
                    "project_no": project_no,
                    "order_no": order_no,
                }
            )
            record["source_fields"] = [
                "item_no",
                "project_no",
                "customer",
                "category",
                "version",
                "inspection_item",
                "lower_limit",
                "upper_limit",
                "unit",
            ]
            records.append(record)
    for product_record in products_by_source_key.values():
        if (
            not product_record.get("customer_product_no")
            and "customer_product_no" in product_record["source_fields"]
        ):
            product_record["source_fields"].remove("customer_product_no")
    _deduplicate_factory_products(records, issues)
    _deduplicate_factory_criteria(records, issues)
    return records


def _parse_material_issue(sheet, header_row, mapping, sha256, issues):
    records = []
    metadata_source_system = (
        ""
        if _mapping_has(mapping, "课别", "来源单位")
        else _metadata_text(sheet, "课别")
    )
    source_document_at = _metadata_datetime(sheet, "打印日期", "打印时间")
    for row_no, cells in _sheet_row_iter(sheet, header_row):
        order_no = _mapped_cell(cells, mapping, "独立需求号", "订单号").display_text
        source_system = (
            _mapped_cell(cells, mapping, "课别", "来源单位").display_text
            or metadata_source_system
        )
        if not any(
            (
                order_no,
                _mapped_cell(cells, mapping, "成品品名").display_text,
                _mapped_cell(cells, mapping, "重量").display_text,
            )
        ):
            continue
        weight = _decimal_text(
            _mapped_cell(cells, mapping, "重量", "重量kg"),
            issues,
            sheet=sheet.name,
            row=row_no,
            field="weight_kg",
            required=True,
        )
        if not order_no:
            issues.append(_issue("error", "订单号不能为空。", sheet=sheet.name, row=row_no, field="order_no"))
        item_no = _mapped_cell(cells, mapping, "项次").display_text
        batch_no = _mapped_cell(cells, mapping, "批号").display_text
        if not source_system:
            issues.append(
                _issue(
                    "error",
                    "课别不能为空。",
                    sheet=sheet.name,
                    row=row_no,
                    field="source_system",
                )
            )
        if not item_no:
            issues.append(
                _issue(
                    "error",
                    "项次不能为空。",
                    sheet=sheet.name,
                    row=row_no,
                    field="item_no",
                )
            )
        if not batch_no:
            issues.append(
                _issue(
                    "error",
                    "批号不能为空。",
                    sheet=sheet.name,
                    row=row_no,
                    field="batch_no",
                )
            )
        record = _record_base(
            sha256, sheet.name, row_no, RECORD_RECEIPT, _raw_row(cells, mapping)
        )
        record.update(
            {
                "order_no": order_no,
                "item_no": item_no,
                "finished_product_name": _mapped_cell(cells, mapping, "成品品名").display_text,
                "specification": _mapped_cell(cells, mapping, "成品规格", "规格").display_text,
                "material": _mapped_cell(cells, mapping, "材质").display_text,
                "batch_no": batch_no,
                "sheet_size": _mapped_cell(cells, mapping, "出片尺寸").display_text,
                "weight_kg": weight,
                "manufactured_on": _date_value(_mapped_cell(cells, mapping, "制造时间", "制造日期")),
                "source_system": source_system,
                "source_document_at": source_document_at,
            }
        )
        record["external_key"] = _external_key(
            "material-receipt",
            source_system,
            record["order_no"],
            record["item_no"],
            record["batch_no"],
        )
        record["source_fields"] = [
            "order_no",
            "item_no",
            "finished_product_name",
            "specification",
            "material",
            "batch_no",
            "sheet_size",
            "weight_kg",
            "manufactured_on",
        ]
        records.append(record)
    return records


def _link_material_receipt_orders(records, issues):
    for record in records:
        if record.get("record_type") != RECORD_RECEIPT:
            continue
        queryset = QualityOrder.objects.filter(order_no=record.get("order_no", ""))
        if record.get("item_no"):
            queryset = queryset.filter(item_no=record["item_no"])
        matches = list(queryset.order_by("id")[:3])
        used_legacy_fallback = False
        if not matches and record.get("item_no") and record.get("specification"):
            legacy = QualityOrder.objects.filter(
                order_no=record.get("order_no", ""),
                item_no="",
                specification__iexact=record["specification"],
            )
            if record.get("material"):
                legacy = legacy.filter(material__iexact=record["material"])
            matches = list(legacy.order_by("id")[:3])
            used_legacy_fallback = bool(matches)
            if len(matches) == 1:
                record["claim_order_item"] = True
                issues.append(
                    _issue(
                        "warning",
                        "按订单号、规格和材质唯一匹配到旧总表订单，提交时将安全补齐项次。",
                        sheet=record["sheet"],
                        row=record["row"],
                        field="item_no",
                    )
                )
        if (
            len(matches) > 1
            and record.get("specification")
            and not used_legacy_fallback
        ):
            refined = queryset.filter(specification__iexact=record["specification"])
            if record.get("material"):
                refined = refined.filter(material__iexact=record["material"])
            matches = list(refined.order_by("id")[:3])
        if len(matches) == 1:
            record["order_id"] = matches[0].pk
        elif not matches:
            issues.append(
                _issue(
                    "warning",
                    "未找到对应订单明细，收料记录仍会保留，但暂不计入具体订单的已收胶料。",
                    sheet=record["sheet"],
                    row=record["row"],
                    field="order",
                )
            )
        else:
            issues.append(
                _issue(
                    "warning",
                    "存在多条可能对应的订单明细，收料记录暂不自动关联，请在线确认。",
                    sheet=record["sheet"],
                    row=record["row"],
                    field="order",
                )
            )


def _validate_duplicate_business_keys(records, issues):
    seen = {}
    for record in records:
        if record["record_type"] not in {RECORD_ORDER, RECORD_RECEIPT}:
            continue
        external_key = record.get("external_key", "")
        if not external_key:
            continue
        first = seen.get((record["record_type"], external_key))
        if first is None:
            seen[(record["record_type"], external_key)] = record
            continue
        label = "客户订单" if record["record_type"] == RECORD_ORDER else "发料记录"
        issues.append(
            _issue(
                "error",
                f"同一文件内存在重复{label}业务键；首次位于工作表“{first['sheet']}”第{first['row']}行，当前重复位于第{record['row']}行。",
                sheet=record["sheet"],
                row=record["row"],
                field="external_key",
            )
        )


def parse_business_sheets(sheets, sha256):
    issues = []
    material_match = None
    product_match = None
    internal_matches = []
    factory_main = False
    factory_criteria = False
    for sheet in sheets:
        header_row, mapping = _find_header(
            sheet, ["独立需求号", "成品品名", "成品规格", "材质", "重量"]
        )
        if mapping:
            material_match = (sheet, header_row, mapping)
        header_row, mapping = _find_header(
            sheet, ["规格", "材质", "料长", "切料重", "一次加硫条件"]
        )
        if mapping and "订单量" not in mapping:
            product_match = (sheet, header_row, mapping)
        header_row, mapping = _find_header(
            sheet, ["订单编号", "规格", "胶料配方", "交期", "订单量"]
        )
        if mapping:
            internal_matches.append((sheet, header_row, mapping))
        if _find_header(sheet, ["独立需求号", "项次", "材质", "规格", "订单量"])[1]:
            factory_main = True
        if _find_header(sheet, ["独立需求号", "项次", "项目号", "检验项目"])[1]:
            factory_criteria = True

    if material_match:
        sheet, header_row, mapping = material_match
        records = _parse_material_issue(sheet, header_row, mapping, sha256, issues)
        _link_material_receipt_orders(records, issues)
        source_type = BusinessImportBatch.SourceType.MATERIAL_ISSUE
    elif factory_main:
        records = _parse_factory_work_contact(sheets, sha256, issues)
        source_type = BusinessImportBatch.SourceType.FACTORY_WORK_CONTACT
    elif product_match:
        sheet, header_row, mapping = product_match
        records = _parse_product_specifications(sheet, header_row, mapping, sha256, issues)
        source_type = BusinessImportBatch.SourceType.PRODUCT_SPECIFICATIONS
    elif internal_matches:
        records = _parse_internal_orders(sheets, sha256, issues)
        _link_existing_product_specifications(records, issues)
        source_type = BusinessImportBatch.SourceType.INTERNAL_ORDERS
    else:
        raise ValueError("无法识别Excel格式，请使用产品规格、内部订单、生产工作联络单或发料清单。")
    _validate_duplicate_business_keys(records, issues)
    if not records:
        issues.append(_issue("error", "工作簿中没有可导入的有效业务行。"))
    return source_type, records, issues


def _existing_source_keys(records):
    by_type = {record_type: [] for record_type in RECORD_TYPES}
    for record in records:
        by_type[record["record_type"]].append(record["source_key"])
    return {
        RECORD_PRODUCT: set(
            ProductSpecification.objects.filter(source_key__in=by_type[RECORD_PRODUCT]).values_list(
                "source_key", flat=True
            )
        ),
        RECORD_ORDER: set(
            QualityOrder.objects.filter(source_key__in=by_type[RECORD_ORDER]).values_list(
                "source_key", flat=True
            )
        ),
        RECORD_RECEIPT: set(
            MaterialReceipt.objects.filter(source_key__in=by_type[RECORD_RECEIPT]).values_list(
                "source_key", flat=True
            )
        ),
        RECORD_CRITERION: set(
            ProductInspectionCriterion.objects.filter(
                source_key__in=by_type[RECORD_CRITERION]
            ).values_list("source_key", flat=True)
        ),
    }


def _datetime_from_payload(value):
    if not value:
        return None
    parsed = parse_datetime(str(value))
    if parsed is None:
        parsed_date = parse_date(str(value))
        if parsed_date:
            parsed = datetime.combine(parsed_date, datetime.min.time())
    if parsed is not None and timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed


def _model_record_value(record, field):
    value = record.get(field)
    if field in {"order_date", "due_date", "manufactured_on"}:
        return _date_from_payload(value)
    if field in {
        "forming_hours",
        "required_material_kg",
        "manual_received_material_kg",
        "weight_kg",
    }:
        return _decimal_from_payload(value)
    return value


def _changes_for_record(instance, record, *, extra=None):
    changes = {}
    for field in record.get("source_fields", []):
        incoming = _model_record_value(record, field)
        current = getattr(instance, field, None)
        if current != incoming:
            changes[field] = {
                "from": json_safe(current),
                "to": json_safe(incoming),
            }
    for field, incoming in (extra or {}).items():
        current = getattr(instance, field, None)
        if current != incoming:
            changes[field] = {"from": json_safe(current), "to": json_safe(incoming)}
    return changes


def _unique_match(queryset):
    matches = list(queryset.order_by("id")[:2])
    return matches[0] if len(matches) == 1 else None, len(matches) > 1


def _order_match(record, issues, *, lock=False):
    queryset = QualityOrder.objects.all()
    if lock:
        queryset = queryset.select_for_update()
    match = queryset.filter(source_key=record["source_key"]).first()
    if match:
        return match
    external_key = record.get("external_key", "")
    if external_key:
        match = queryset.filter(external_key=external_key).first()
        if match:
            return match

    source_type = record.get("source_type")
    if source_type == BusinessImportBatch.SourceType.INTERNAL_ORDERS:
        match, ambiguous = _unique_match(
            queryset.filter(
                source_sheet=record["sheet"],
                source_row=record["row"],
                source_key__gt="",
            )
        )
        if ambiguous:
            record["ambiguous_internal_row"] = True
            issues.append(
                _issue(
                    "warning",
                    "同一总表行对应多条旧订单，已跳过自动更新，请在线核对。",
                    sheet=record["sheet"],
                    row=record["row"],
                    field="order",
                )
            )
        return match

    order_no = record.get("order_no", "")
    item_no = record.get("item_no", "")
    source_system = record.get("source_system", "")
    if order_no and item_no:
        match, ambiguous = _unique_match(
            queryset.filter(order_no=order_no, item_no=item_no).filter(
                Q(source_system="")
                | Q(source_system=source_system)
                | Q(source_system=INTERNAL_SOURCE_SYSTEM)
            )
        )
        if match:
            return match
        if ambiguous:
            issues.append(
                _issue(
                    "warning",
                    "订单号和项次对应多条旧订单，不能自动认领旧记录。",
                    sheet=record["sheet"],
                    row=record["row"],
                    field="item_no",
                )
            )
            return None

    legacy = queryset.filter(
        order_no=order_no,
        item_no="",
        specification__iexact=record.get("specification", ""),
        material__iexact=record.get("material", ""),
        order_quantity=record.get("order_quantity"),
    )
    due_date = _date_from_payload(record.get("due_date"))
    exact, exact_ambiguous = _unique_match(legacy.filter(due_date=due_date))
    if exact:
        return exact
    if not exact_ambiguous and due_date is not None:
        missing_due, missing_due_ambiguous = _unique_match(legacy.filter(due_date__isnull=True))
        if missing_due:
            issues.append(
                _issue(
                    "warning",
                    "旧总表订单交期为空，已按其余完整字段唯一匹配并准备补齐。",
                    sheet=record["sheet"],
                    row=record["row"],
                    field="due_date",
                )
            )
            return missing_due
        exact_ambiguous = missing_due_ambiguous
    if exact_ambiguous:
        issues.append(
            _issue(
                "warning",
                "旧总表存在多条相同订单明细，不能自动认领；将保留为独立客户订单行。",
                sheet=record["sheet"],
                row=record["row"],
                field="order",
            )
        )
    return None


def _receipt_match(record, *, lock=False):
    queryset = MaterialReceipt.objects.all()
    if lock:
        queryset = queryset.select_for_update()
    match = queryset.filter(source_key=record["source_key"]).first()
    if match:
        return match
    external_key = record.get("external_key", "")
    if external_key:
        match = queryset.filter(external_key=external_key).first()
        if match:
            return match
    if all(record.get(field) for field in ("order_no", "item_no", "batch_no")):
        match, _ = _unique_match(
            queryset.filter(
                order_no=record["order_no"],
                item_no=record["item_no"],
                batch_no=record["batch_no"],
            ).filter(Q(source_system="") | Q(source_system=record.get("source_system", "")))
        )
        if match:
            return match
    match, _ = _unique_match(
        queryset.filter(
            source_sheet=record["sheet"],
            source_row=record["row"],
            source_key__gt="",
            batch_no=record.get("batch_no", ""),
        )
    )
    return match


def _product_match(record, *, lock=False):
    for key in (
        "ambiguous_product_identity",
        "ambiguous_product_ids",
        "ambiguous_product_resolved",
        "ambiguous_product_error_added",
        "unsafe_missing_identity",
    ):
        record.pop(key, None)
    queryset = ProductSpecification.objects.all()
    if lock:
        queryset = queryset.select_for_update()
    match = queryset.filter(source_key=record["source_key"]).first()
    if match:
        return match
    normalized_key = record.get("normalized_key", "")
    if normalized_key:
        matches = list(
            queryset.filter(normalized_key=normalized_key, is_active=True)
            .order_by("id")[:3]
        )
        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            record["ambiguous_product_identity"] = True
            record["ambiguous_product_ids"] = [item.pk for item in matches]
    return None


def _criterion_match(record, order_id=None, *, lock=False):
    for key in ("ambiguous_criterion_identity", "ambiguous_criterion_ids"):
        record.pop(key, None)
    queryset = ProductInspectionCriterion.objects.all()
    if lock:
        queryset = queryset.select_for_update()
    match = queryset.filter(source_key=record["source_key"]).first()
    if match or not order_id:
        return match
    exact_matches = list(
        queryset.filter(
            order_id=order_id,
            item_no=record.get("item_no", ""),
            project_no=record.get("project_no", ""),
            category=record.get("category", ""),
            version=record.get("version", ""),
            inspection_item=record.get("inspection_item", ""),
            unit=record.get("unit", ""),
        ).order_by("id")[:3]
    )
    if len(exact_matches) == 1:
        return exact_matches[0]
    if len(exact_matches) > 1:
        record["ambiguous_criterion_identity"] = True
        record["ambiguous_criterion_ids"] = [item.pk for item in exact_matches]
        return None
    # 项目号是产品资料的一部分，客户修正项目号时仍应更新同一条检验标准。
    # 回退匹配继续包含单位，避免同名检验项目的不同单位被错误合并。
    fallback_matches = list(
        queryset.filter(
            order_id=order_id,
            item_no=record.get("item_no", ""),
            category=record.get("category", ""),
            version=record.get("version", ""),
            inspection_item=record.get("inspection_item", ""),
            unit=record.get("unit", ""),
        ).order_by("id")[:3]
    )
    if len(fallback_matches) == 1:
        return fallback_matches[0]
    if len(fallback_matches) > 1:
        record["ambiguous_criterion_identity"] = True
        record["ambiguous_criterion_ids"] = [item.pk for item in fallback_matches]
    return None


def _incoming_is_older(instance, record):
    incoming = _datetime_from_payload(record.get("source_document_at"))
    return bool(
        instance.source_document_at
        and (incoming is None or incoming < instance.source_document_at)
    )


def _newer_source_document(instance, record):
    incoming = _datetime_from_payload(record.get("source_document_at"))
    if incoming and (
        instance.source_document_at is None or incoming > instance.source_document_at
    ):
        return incoming
    return None


def _product_dependency_change(instance, record, product_records):
    source_key = record.get("product_spec_source_key")
    product_record = product_records.get(source_key)
    if product_record is not None:
        match_id = product_record.get("match_id")
        if product_record.get("action") == "CREATE":
            return {
                "from": instance.product_specification_id,
                "to": f"NEW:{source_key}",
            }
        if match_id and instance.product_specification_id != match_id:
            return {"from": instance.product_specification_id, "to": match_id}
    product_spec_id = record.get("product_spec_id")
    if product_spec_id and instance.product_specification_id != product_spec_id:
        return {"from": instance.product_specification_id, "to": product_spec_id}
    return None


def _prepare_record_actions(records, source_type, issues, *, lock=False):
    order_records = {record["source_key"]: record for record in records if record["record_type"] == RECORD_ORDER}
    product_records = {
        record["source_key"]: record for record in records if record["record_type"] == RECORD_PRODUCT
    }

    for record in records:
        record["source_type"] = source_type
        if record["record_type"] != RECORD_PRODUCT:
            continue
        match = _product_match(record, lock=lock)
        record["match_id"] = match.pk if match else None
        if not match:
            record["action"] = (
                "SKIP" if record.get("ambiguous_product_identity") else "CREATE"
            )
            record["changes"] = {}
        else:
            incoming_source_at = _datetime_from_payload(
                record.get("source_document_at")
            )
            latest_linked_order_at = None
            if source_type == BusinessImportBatch.SourceType.FACTORY_WORK_CONTACT:
                latest_linked_order_at = QualityOrder.objects.filter(
                    product_specification_id=match.pk,
                    source_document_at__isnull=False,
                ).aggregate(value=Max("source_document_at"))["value"]
            if latest_linked_order_at and (
                incoming_source_at is None
                or incoming_source_at < latest_linked_order_at
            ):
                record["action"] = "SKIP"
                record["changes"] = {}
                record["protected_shared_product_version"] = True
                issues.append(
                    _issue(
                        "warning",
                        "该产品已关联时间更新的客户订单，旧文件仍会创建或更新当前订单，但不会回退共享产品工艺参数。",
                        sheet=record["sheet"],
                        row=record["row"],
                        field="source_document_at",
                    )
                )
                continue
            changes = _changes_for_record(match, record)
            record["changes"] = changes
            record["action"] = "UPDATE" if changes else "SKIP"

    reused_linked_product_ids = {}
    for record in records:
        if record["record_type"] != RECORD_ORDER:
            continue
        record.pop("unsafe_product_identity", None)
        match = _order_match(record, issues, lock=lock)
        record["match_id"] = match.pk if match else None
        if not match:
            record["action"] = (
                "SKIP"
                if record.get("stale_source_version")
                or record.get("ambiguous_internal_row")
                else "CREATE"
            )
            record["changes"] = {}
            continue
        if (
            source_type == BusinessImportBatch.SourceType.INTERNAL_ORDERS
            and match.external_key.startswith("factory-order|")
        ):
            record["action"] = "SKIP"
            record["changes"] = {}
            issues.append(
                _issue(
                    "warning",
                    "该总表行已由客户订单认领，旧总表不会覆盖客户来源数据。",
                    sheet=record["sheet"],
                    row=record["row"],
                    field="order",
                )
            )
            continue
        if _incoming_is_older(match, record):
            record["action"] = "SKIP"
            record["changes"] = {}
            record["stale_source_version"] = True
            issues.append(
                _issue(
                    "warning",
                    "来源文件时间缺失或早于当前订单版本，已跳过业务字段覆盖。",
                    sheet=record["sheet"],
                    row=record["row"],
                    field="source_document_at",
                )
            )
            continue
        product_record = product_records.get(record.get("product_spec_source_key"))
        should_reuse_linked_product = (
            source_type == BusinessImportBatch.SourceType.FACTORY_WORK_CONTACT
            and product_record is not None
            and match.product_specification_id
            and (
                not product_record.get("customer_product_no")
                or product_record.get("ambiguous_product_identity")
            )
        )
        if should_reuse_linked_product:
            source_key = product_record["source_key"]
            previous_product_id = reused_linked_product_ids.get(source_key)
            if previous_product_id not in (None, match.product_specification_id):
                issues.append(
                    _issue(
                        "error",
                        "同一批产品对应多个现有产品，不能安全自动复用，请先在线整理产品资料。",
                        sheet=record["sheet"],
                        row=record["row"],
                        field="product_specification",
                    )
                )
            else:
                product_queryset = ProductSpecification.objects.all()
                if lock:
                    product_queryset = product_queryset.select_for_update()
                linked_product = product_queryset.filter(
                    pk=match.product_specification_id
                ).first()
                if linked_product is not None:
                    can_reuse = True
                    if product_record.get("ambiguous_product_identity") and (
                        linked_product.pk
                        not in product_record.get("ambiguous_product_ids", [])
                    ):
                        can_reuse = False
                    if not product_record.get("customer_product_no"):
                        identity_mismatches = [
                            field
                            for field in ("specification", "material", "mold_no")
                            if _stable_component(product_record.get(field))
                            != _stable_component(getattr(linked_product, field, ""))
                        ]
                        if identity_mismatches:
                            can_reuse = False
                            product_record["unsafe_missing_identity"] = True
                            product_record["action"] = "SKIP"
                            product_record["changes"] = {}
                            record["unsafe_product_identity"] = True
                            issues.append(
                                _issue(
                                    "error",
                                    "缺少项目号且规格、材质或模具号与现有产品不一致，不能安全复用；请上传包含检验标准和项目号的完整客户文件。",
                                    sheet=record["sheet"],
                                    row=record["row"],
                                    field="customer_product_no",
                                )
                            )
                    if can_reuse:
                        reused_linked_product_ids[source_key] = linked_product.pk
                        product_record["match_id"] = linked_product.pk
                        product_record["ambiguous_product_resolved"] = True
                        if not product_record.get(
                            "protected_shared_product_version"
                        ):
                            product_changes = _changes_for_record(
                                linked_product, product_record
                            )
                            product_record["changes"] = product_changes
                            product_record["action"] = (
                                "UPDATE" if product_changes else "SKIP"
                            )
                    elif product_record.get("ambiguous_product_identity"):
                        product_record["ambiguous_product_error_added"] = True
                        issues.append(
                            _issue(
                                "error",
                                "数据库中存在多条相同产品身份，且当前订单不能唯一指明可复用产品；请先在线整理产品资料。",
                                sheet=record["sheet"],
                                row=record["row"],
                                field="product_specification",
                            )
                        )
        if record.get("unsafe_product_identity"):
            record["action"] = "SKIP"
            record["changes"] = {}
            continue
        extra = {}
        if record.get("external_key") and match.external_key != record["external_key"]:
            extra["external_key"] = record["external_key"]
        if record.get("source_system") and match.source_system != record["source_system"]:
            extra["source_system"] = record["source_system"]
        newer_source_document = _newer_source_document(match, record)
        if newer_source_document:
            extra["source_document_at"] = newer_source_document
        changes = _changes_for_record(match, record, extra=extra)
        product_dependency = _product_dependency_change(match, record, product_records)
        if product_dependency:
            changes["product_specification_id"] = product_dependency
        record["changes"] = changes
        record["action"] = "UPDATE" if changes else "SKIP"

    for product_record in product_records.values():
        if (
            product_record.get("ambiguous_product_identity")
            and not product_record.get("ambiguous_product_resolved")
            and not product_record.get("ambiguous_product_error_added")
        ):
            product_record["action"] = "SKIP"
            product_record["changes"] = {}
            product_record["ambiguous_product_error_added"] = True
            issues.append(
                _issue(
                    "error",
                    "数据库中存在多条相同产品身份，无法安全自动选择；请先在线整理产品资料。",
                    sheet=product_record["sheet"],
                    row=product_record["row"],
                    field="product_specification",
                )
            )

    order_match_ids = {
        key: value.get("match_id") for key, value in order_records.items() if value.get("match_id")
    }
    for record in records:
        if record["record_type"] == RECORD_RECEIPT:
            match = _receipt_match(record, lock=lock)
            record["match_id"] = match.pk if match else None
            extra = {}
            if match and record.get("external_key") and match.external_key != record["external_key"]:
                extra["external_key"] = record["external_key"]
            if match and record.get("source_system") and match.source_system != record["source_system"]:
                extra["source_system"] = record["source_system"]
            if match and record.get("order_id") and match.order_id != record["order_id"]:
                extra["order_id"] = record["order_id"]
            if match:
                newer_source_document = _newer_source_document(match, record)
                if newer_source_document:
                    extra["source_document_at"] = newer_source_document
            if not match:
                record["action"] = (
                    "SKIP" if record.get("stale_source_version") else "CREATE"
                )
                record["changes"] = {}
            elif _incoming_is_older(match, record):
                record["action"] = "SKIP"
                record["changes"] = {}
                record["stale_source_version"] = True
                issues.append(
                    _issue(
                        "warning",
                        "来源文件时间缺失或早于当前发料版本，已跳过业务字段覆盖。",
                        sheet=record["sheet"],
                        row=record["row"],
                        field="source_document_at",
                    )
                )
            else:
                changes = _changes_for_record(match, record, extra=extra)
                record["changes"] = changes
                record["action"] = "UPDATE" if changes else "SKIP"
        elif record["record_type"] == RECORD_CRITERION:
            order_id = order_match_ids.get(record.get("order_source_key"))
            match = _criterion_match(record, order_id, lock=lock)
            record["match_id"] = match.pk if match else None
            if not match:
                record["action"] = (
                    "SKIP"
                    if record.get("ambiguous_criterion_identity")
                    else "CREATE"
                )
                record["changes"] = {}
                linked_order_record = order_records.get(
                    record.get("order_source_key")
                )
                if (
                    record.get("ambiguous_criterion_identity")
                    and not (
                        linked_order_record
                        and linked_order_record.get("stale_source_version")
                    )
                ):
                    issues.append(
                        _issue(
                            "error",
                            "数据库中存在多条相同检验标准业务键，无法安全自动选择；请先在线整理检验标准。",
                            sheet=record["sheet"],
                            row=record["row"],
                            field="inspection_criterion",
                        )
                    )
            else:
                changes = _changes_for_record(match, record)
                product_dependency = _product_dependency_change(
                    match, record, product_records
                )
                if product_dependency:
                    changes["product_specification_id"] = product_dependency
                record["changes"] = changes
                record["action"] = "UPDATE" if changes else "SKIP"

    stale_order_keys = {
        record["source_key"]
        for record in records
        if record["record_type"] == RECORD_ORDER
        and record.get("stale_source_version")
    }
    stale_product_keys = {
        record.get("product_spec_source_key")
        for record in records
        if record["record_type"] == RECORD_ORDER
        and record["source_key"] in stale_order_keys
        and record.get("product_spec_source_key")
    }
    for record in records:
        belongs_to_stale_order = (
            record["record_type"] == RECORD_PRODUCT
            and record["source_key"] in stale_product_keys
        ) or (
            record["record_type"] == RECORD_CRITERION
            and record.get("order_source_key") in stale_order_keys
        )
        if belongs_to_stale_order:
            record["action"] = "SKIP"
            record["changes"] = {}
            record["stale_source_version"] = True

    # Make matched products available to order/criterion commit loops.
    for record in records:
        source_key = record.get("product_spec_source_key")
        if source_key and source_key in product_records:
            record["product_spec_match_id"] = product_records[source_key].get("match_id")


def _preview_summary(record):
    if record["record_type"] == RECORD_PRODUCT:
        return " / ".join(
            item
            for item in (record.get("specification"), record.get("material"), record.get("mold_no"))
            if item
        )
    if record["record_type"] == RECORD_ORDER:
        return f"{record.get('order_no', '')} / {record.get('specification', '')} / {record.get('order_quantity') or ''}"
    if record["record_type"] == RECORD_RECEIPT:
        return f"{record.get('order_no', '')} / {record.get('batch_no', '')} / {record.get('weight_kg') or ''}kg"
    return f"{record.get('order_no', '')} / {record.get('inspection_item', '')}"


def preview_business_workbook(uploaded_file, user):
    data = uploaded_file.read()
    uploaded_file.seek(0)
    sha256 = hashlib.sha256(data).hexdigest()
    sheets, parser = read_business_workbook(data)
    source_type, records, issues = parse_business_sheets(sheets, sha256)
    _prepare_record_actions(records, source_type, issues)
    for record in records:
        if record["action"] == "SKIP":
            issues.append(
                _issue(
                    "warning",
                    "该业务行与现有数据相同或来源版本较旧，将跳过业务字段更新。",
                    sheet=record["sheet"],
                    row=record["row"],
                    field="source_key",
                )
            )
    errors = [item for item in issues if item["level"] == "error"]
    warnings = [item for item in issues if item["level"] == "warning"]
    batch = BusinessImportBatch(
        source_type=source_type,
        parser=parser,
        original_name=str(getattr(uploaded_file, "name", "upload.xlsx"))[:255],
        sha256=sha256,
        payload={"rows": json_safe(records), "parser": parser},
        errors=errors,
        warnings=warnings,
        created_by=user,
    )
    batch.original_file.save(batch.original_name, ContentFile(data), save=False)
    batch.save()
    counts = {
        "product_specifications": sum(r["record_type"] == RECORD_PRODUCT for r in records),
        "orders": sum(r["record_type"] == RECORD_ORDER for r in records),
        "material_receipts": sum(r["record_type"] == RECORD_RECEIPT for r in records),
        "inspection_criteria": sum(r["record_type"] == RECORD_CRITERION for r in records),
    }
    error_rows = {(item.get("sheet"), item.get("row")) for item in errors}
    return {
        "token": str(batch.pk),
        "source_type": source_type,
        "total_rows": len(records),
        "counts": counts,
        "error_count": len(errors),
        "warning_count": len(warnings),
        "rows": [
            {
                "row_key": record["row_key"],
                "record_type": record["record_type"],
                "sheet": record["sheet"],
                "row": record["row"],
                "action": record["action"],
                "match_id": record.get("match_id"),
                "changes": record.get("changes", {}),
                "order_no": record.get("order_no", ""),
                "item_no": record.get("item_no", ""),
                "specification": record.get("specification", ""),
                "material": record.get("material", ""),
                "summary": _preview_summary(record),
                "valid": (record["sheet"], record["row"]) not in error_rows,
            }
            for record in records
        ],
        "issues": issues,
    }


def _date_from_payload(value):
    return parse_date(value) if value else None


def _decimal_from_payload(value):
    return Decimal(value) if value not in (None, "") else None


def _find_order_for_receipt(record, imported_orders):
    if record.get("order_id"):
        order = QualityOrder.objects.filter(pk=record["order_id"]).first()
        if (
            order is not None
            and order.order_no == record.get("order_no", "")
            and (
                not record.get("item_no")
                or order.item_no == record.get("item_no", "")
            )
        ):
            return order
    order_source_key = record.get("order_source_key")
    if order_source_key and order_source_key in imported_orders:
        return imported_orders[order_source_key]
    queryset = QualityOrder.objects.filter(order_no=record["order_no"])
    if record.get("item_no"):
        queryset = queryset.filter(item_no=record["item_no"])
    matches = list(queryset.order_by("id")[:2])
    return matches[0] if len(matches) == 1 else None


def _claim_order_for_receipt(record, user, source_batch):
    if not record.get("claim_order_item") or not record.get("order_id"):
        return None
    exact, ambiguous = _unique_match(
        QualityOrder.objects.select_for_update().filter(
            order_no=record.get("order_no", ""),
            item_no=record.get("item_no", ""),
        )
    )
    if exact:
        return exact
    if ambiguous:
        return None
    order = QualityOrder.objects.select_for_update().filter(pk=record["order_id"]).first()
    if order is None or order.item_no or order.order_no != record.get("order_no", ""):
        return None
    candidates = QualityOrder.objects.select_for_update().filter(
        order_no=record.get("order_no", ""),
        item_no="",
        specification__iexact=record.get("specification", ""),
    )
    if record.get("material"):
        candidates = candidates.filter(material__iexact=record["material"])
    candidate, ambiguous = _unique_match(candidates)
    if ambiguous or candidate is None or candidate.pk != order.pk:
        return None
    before = model_snapshot(order)
    order.item_no = record["item_no"]
    order.save(update_fields=["item_no", "updated_at"])
    record_revision(
        order,
        user,
        BusinessRecordRevision.Action.IMPORT,
        source_batch=source_batch,
        before=before,
    )
    return order


def _apply_source_fields(instance, record, source_batch):
    for field in record.get("source_fields", []):
        setattr(instance, field, _model_record_value(record, field))
    if record.get("source_system"):
        instance.source_system = record["source_system"]
    if record.get("external_key"):
        instance.external_key = record["external_key"]
    source_document_at = _datetime_from_payload(record.get("source_document_at"))
    if source_document_at and (
        instance.source_document_at is None or source_document_at >= instance.source_document_at
    ):
        instance.source_document_at = source_document_at
    instance.last_source_batch = source_batch
    instance.last_imported_at = timezone.now()
    instance.raw_data = record.get("raw_data", {})


def _relink_unlinked_receipts(order, user, source_batch):
    if not order.item_no:
        return 0
    if QualityOrder.objects.filter(order_no=order.order_no, item_no=order.item_no).count() != 1:
        return 0
    receipts = list(
        MaterialReceipt.objects.select_for_update().filter(
            order__isnull=True,
            order_no=order.order_no,
            item_no=order.item_no,
        )
    )
    for receipt in receipts:
        before = model_snapshot(receipt)
        receipt.order = order
        receipt.save(update_fields=["order", "updated_at"])
        record_revision(
            receipt,
            user,
            BusinessRecordRevision.Action.IMPORT,
            source_batch=source_batch,
            before=before,
        )
    return len(receipts)


def commit_business_batch(batch, user):
    try:
        with transaction.atomic():
            current = BusinessImportBatch.objects.filter(pk=batch.pk).first()
            if current is None or current.created_by_id != user.pk:
                raise ValueError("无权提交该导入批次。")
            if current.status != BusinessImportBatch.Status.PREVIEWED:
                raise ValueError("该导入批次已经提交或正在提交，不能重复导入。")
            if current.errors:
                raise ValueError("预检存在错误，请修正Excel后重新上传。")
            claimed = BusinessImportBatch.objects.filter(
                pk=current.pk,
                created_by=user,
                status=BusinessImportBatch.Status.PREVIEWED,
            ).update(status=BusinessImportBatch.Status.COMMITTING)
            if claimed != 1:
                raise ValueError("该导入批次已经提交或正在提交，不能重复导入。")
            current.refresh_from_db()
            records = current.payload.get("rows", [])
            commit_issues = []
            _prepare_record_actions(
                records, current.source_type, commit_issues, lock=True
            )
            commit_errors = [
                issue for issue in commit_issues if issue.get("level") == "error"
            ]
            if commit_errors:
                raise ValueError(
                    f"提交前复检失败：{commit_errors[0].get('message', '存在数据冲突。')}"
                )
            counter_names = (
                "product_specifications",
                "orders",
                "material_receipts",
                "inspection_criteria",
            )
            created = {key: 0 for key in counter_names}
            updated = dict(created)
            skipped = dict(created)
            product_specs = {}
            imported_orders = {}

            for record in records:
                if record["record_type"] != RECORD_PRODUCT:
                    continue
                product = None
                if record.get("match_id"):
                    product = ProductSpecification.objects.select_for_update().filter(
                        pk=record["match_id"]
                    ).first()
                if record["action"] == "SKIP" and (
                    product is not None
                    or record.get("stale_source_version")
                    or record.get("ambiguous_product_identity")
                    or record.get("unsafe_missing_identity")
                ):
                    skipped["product_specifications"] += 1
                    if product is not None:
                        product_specs[record["source_key"]] = product
                    continue
                before = model_snapshot(product) if product is not None else None
                if product is None:
                    product = ProductSpecification(
                        source_batch=current,
                        source_sheet=record["sheet"],
                        source_row=record["row"],
                        source_key=record["source_key"],
                    )
                for field in record.get("source_fields", []):
                    setattr(product, field, record.get(field, ""))
                product.raw_data = record.get("raw_data", {})
                product.save()
                record_revision(
                    product,
                    user,
                    BusinessRecordRevision.Action.IMPORT,
                    source_batch=current,
                    before=before,
                )
                product_specs[record["source_key"]] = product
                if before is None:
                    created["product_specifications"] += 1
                else:
                    updated["product_specifications"] += 1

            for record in records:
                if record["record_type"] != RECORD_ORDER:
                    continue
                product_spec = product_specs.get(record.get("product_spec_source_key"))
                if product_spec is None and record.get("product_spec_id"):
                    product_spec = ProductSpecification.objects.filter(
                        pk=record["product_spec_id"], is_active=True
                    ).first()
                order = None
                if record.get("match_id"):
                    order = QualityOrder.objects.select_for_update().filter(
                        pk=record["match_id"]
                    ).first()
                if record["action"] == "SKIP" and (
                    order is not None
                    or record.get("stale_source_version")
                    or record.get("ambiguous_internal_row")
                ):
                    skipped["orders"] += 1
                    if order is not None:
                        imported_at = timezone.now()
                        QualityOrder.objects.filter(pk=order.pk).update(
                            last_imported_at=imported_at,
                            last_source_batch=current,
                        )
                        order.last_imported_at = imported_at
                        order.last_source_batch = current
                        imported_orders[record["source_key"]] = order
                        _relink_unlinked_receipts(order, user, current)
                    continue
                before = model_snapshot(order) if order is not None else None
                if order is None:
                    order = QualityOrder(
                        batch_no=record.get("batch_no", ""),
                        product_code=record.get("product_code", ""),
                        product_name=record.get("product_name", ""),
                        product_specification=product_spec,
                        production_required=record.get("production_required"),
                        status=record.get("status", QualityOrder.Status.OPEN),
                        notes=record.get("notes", ""),
                        source_batch=current,
                        source_sheet=record["sheet"],
                        source_row=record["row"],
                        source_key=record["source_key"],
                        created_by=user,
                    )
                elif (
                    product_spec is not None
                    and order.product_specification_id != product_spec.pk
                ):
                    order.product_specification = product_spec
                _apply_source_fields(order, record, current)
                order.save()
                record_revision(
                    order,
                    user,
                    BusinessRecordRevision.Action.IMPORT,
                    source_batch=current,
                    before=before,
                )
                imported_orders[record["source_key"]] = order
                if before is None:
                    created["orders"] += 1
                else:
                    updated["orders"] += 1
                _relink_unlinked_receipts(order, user, current)

            for record in records:
                if record["record_type"] == RECORD_CRITERION:
                    criterion = None
                    if record.get("match_id"):
                        criterion = ProductInspectionCriterion.objects.select_for_update().filter(
                            pk=record["match_id"]
                        ).first()
                    if record["action"] == "SKIP" and (
                        criterion is not None or record.get("stale_source_version")
                        or record.get("ambiguous_criterion_identity")
                    ):
                        skipped["inspection_criteria"] += 1
                        continue
                    product = product_specs.get(record.get("product_spec_source_key"))
                    if product is None:
                        raise ValueError("检验标准对应的产品规格不存在，请重新预检。")
                    order = imported_orders.get(record.get("order_source_key"))
                    before = model_snapshot(criterion) if criterion is not None else None
                    if criterion is None:
                        criterion = ProductInspectionCriterion(
                            product_specification=product,
                            order=order,
                            source_batch=current,
                            source_sheet=record["sheet"],
                            source_row=record["row"],
                            source_key=record["source_key"],
                        )
                    else:
                        criterion.product_specification = product
                        if order is not None:
                            criterion.order = order
                    for field in record.get("source_fields", []):
                        setattr(criterion, field, record.get(field, ""))
                    criterion.raw_data = record.get("raw_data", {})
                    criterion.save()
                    record_revision(
                        criterion,
                        user,
                        BusinessRecordRevision.Action.IMPORT,
                        source_batch=current,
                        before=before,
                    )
                    if before is None:
                        created["inspection_criteria"] += 1
                    else:
                        updated["inspection_criteria"] += 1
                elif record["record_type"] == RECORD_RECEIPT:
                    receipt = None
                    if record.get("match_id"):
                        receipt = MaterialReceipt.objects.select_for_update().filter(
                            pk=record["match_id"]
                        ).first()
                    if record["action"] == "SKIP" and (
                        receipt is not None or record.get("stale_source_version")
                    ):
                        skipped["material_receipts"] += 1
                        if receipt is not None:
                            imported_at = timezone.now()
                            MaterialReceipt.objects.filter(pk=receipt.pk).update(
                                last_imported_at=imported_at,
                                last_source_batch=current,
                            )
                            receipt.last_imported_at = imported_at
                            receipt.last_source_batch = current
                        continue
                    linked_order = _claim_order_for_receipt(record, user, current)
                    if linked_order is None:
                        linked_order = _find_order_for_receipt(record, imported_orders)
                    before = model_snapshot(receipt) if receipt is not None else None
                    if receipt is None:
                        receipt = MaterialReceipt(
                            order=linked_order,
                            source_batch=current,
                            source_sheet=record["sheet"],
                            source_row=record["row"],
                            source_key=record["source_key"],
                        )
                    elif linked_order is not None:
                        receipt.order = linked_order
                    _apply_source_fields(receipt, record, current)
                    receipt.save()
                    record_revision(
                        receipt,
                        user,
                        BusinessRecordRevision.Action.IMPORT,
                        source_batch=current,
                        before=before,
                    )
                    if before is None:
                        created["material_receipts"] += 1
                    else:
                        updated["material_receipts"] += 1

            committed_at = timezone.now()
            payload = current.payload
            payload["rows"] = json_safe(records)
            changed = BusinessImportBatch.objects.filter(
                pk=current.pk, status=BusinessImportBatch.Status.COMMITTING
            ).update(
                status=BusinessImportBatch.Status.COMMITTED,
                committed_at=committed_at,
                payload=payload,
            )
            if changed != 1:
                raise ValueError("导入批次状态已变化，请刷新后确认结果。")
            imported = {key: created[key] + updated[key] for key in counter_names}
            return {
                "created": created,
                "updated": updated,
                "imported": imported,
                "skipped": skipped,
            }
    except ValueError:
        raise
    except DjangoValidationError as exc:
        if hasattr(exc, "message_dict"):
            message = "；".join(
                f"{field}：{'；'.join(messages)}" for field, messages in exc.message_dict.items()
            )
        else:
            message = "；".join(exc.messages)
        raise ValueError(f"提交数据校验失败：{message}") from exc
    except IntegrityError as exc:
        raise ValueError("数据已被其他操作导入或占用，请重新预检。") from exc
    except OperationalError as exc:
        raise ValueError("数据库正忙，导入未执行，请稍后重试。") from exc


TEMPLATE_HEADERS = {
    "product_specifications": [
        "产品名称",
        "客户产品号",
        "规格",
        "材质",
        "料长",
        "切料重",
        "条数",
        "一次加硫条件",
        "二次加硫条件",
        "总孔数",
        "有效孔数",
        "模具在库",
        "模具号",
        "模具尺寸",
        "标准工时",
        "备注",
    ],
    "orders": [
        "订单编号",
        "项次",
        "产品名称",
        "流程卡",
        "规格",
        "胶料配方",
        "交期",
        "订单量",
        "胶料用量",
        "已发胶料",
        "成型工时",
        "下单时间",
        "模具尺寸",
        "是否生产",
        "生产数量",
        "出货日期",
        "出货数量",
    ],
    "material_receipts": [
        "序号",
        "课别",
        "项次",
        "独立需求号",
        "成品品名",
        "成品规格",
        "材质",
        "批号",
        "出片尺寸",
        "重量",
        "制造时间",
    ],
}


def create_business_template(kind="product_specifications"):
    if kind not in TEMPLATE_HEADERS:
        raise ValueError("无效的模板类型。")
    workbook = Workbook()
    sheet = workbook.active
    titles = {
        "product_specifications": "产品规格",
        "orders": "订单",
        "material_receipts": "胶料收料",
    }
    sheet.title = titles[kind]
    sheet.append(TEMPLATE_HEADERS[kind])
    header_fill = PatternFill("solid", fgColor="1677FF")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(TEMPLATE_HEADERS[kind]))}1"
    for index, header in enumerate(TEMPLATE_HEADERS[kind], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) * 2 + 4, 12), 28)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _safe_excel_text(value):
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def create_business_error_report(batch):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "预检问题"
    sheet.append(["级别", "工作表", "行号", "字段", "说明"])
    for issue in [*batch.errors, *batch.warnings]:
        sheet.append(
            [
                _safe_excel_text(issue.get("level", "")),
                _safe_excel_text(issue.get("sheet", "")),
                issue.get("row", ""),
                _safe_excel_text(issue.get("field", "")),
                _safe_excel_text(issue.get("message", "")),
            ]
        )
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1677FF")
    sheet.freeze_panes = "A2"
    for index, width in enumerate((12, 24, 10, 24, 70), start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
