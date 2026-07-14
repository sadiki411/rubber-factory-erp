import io
import re
from datetime import timedelta
from zipfile import ZipFile

from django.test import TestCase
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from molds.models import Machine
from production.imports import (
    DAILY_HEADERS,
    FIELD_CELLS,
    LEGACY_DAILY_HEADERS,
    SETTLEMENT_CELLS,
    create_production_template,
)
from production.models import ProductionImportBatch, ProductionRun, ProductionStation

from .helpers import ProductionTestMixin, production_workbook_upload


class ProductionTemplateTests(TestCase):
    def test_template_preserves_order_card_concept_and_daily_table(self):
        workbook = load_workbook(io.BytesIO(create_production_template()))
        sheet = workbook.active
        self.assertEqual(sheet.title, "订单卡-001")
        self.assertIn("生产订单统计表", sheet["A1"].value)
        self.assertEqual(sheet["A2"].value, "机台编号")
        self.assertEqual(sheet["I3"].value, "填写系统中已启用的机台编号")
        station_validation = next(
            validation
            for validation in sheet.data_validations.dataValidation
            if "B2" in validation.sqref
        )
        self.assertEqual(station_validation.type, "custom")
        self.assertEqual(station_validation.formula1, "=LEN(TRIM(B2))>0")
        self.assertEqual(sheet["A6"].value, "上模时间")
        self.assertEqual(sheet["C6"].value, "预计换模时间")
        self.assertEqual(sheet["E6"].value, "下机时间")
        self.assertIn("订单完工结算", sheet["A8"].value)
        for _, (label_cell, value_cell, label) in SETTLEMENT_CELLS.items():
            self.assertEqual(sheet[label_cell].value, label)
            self.assertIsNone(sheet[value_cell].value)
        for column, (_, label) in enumerate(DAILY_HEADERS, 1):
            self.assertEqual(sheet.cell(13, column).value, label)


class ProductionImportTests(ProductionTestMixin, TestCase):
    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def test_copied_order_cards_preview_and_commit_without_changing_mold_state(self):
        mold = self.create_mold()
        original_machine_id = mold.current_machine_id
        now = timezone.now().replace(microsecond=0)
        upload = production_workbook_upload(
            [
                {
                    "sheet": "订单卡-历史001",
                    "station_code": "A01",
                    "order_no": "HISTORY-001",
                    "status": "COMPLETED",
                    "mold_code": mold.asset_code,
                    "loaded_at": now - timedelta(hours=3),
                    "unloaded_at": now - timedelta(hours=1),
                    "actual_good_quantity": 590,
                    "actual_defective_quantity": 10,
                    "total_material_kg": 20,
                    "labor_cost": 100,
                    "energy_cost": 30,
                    "other_cost": 0,
                    "settlement_notes": "历史订单结算",
                    "daily_logs": [
                        {
                            "date": timezone.localdate(),
                            "produced_mold_count": 100,
                            "notes": "历史订单日报",
                        }
                    ],
                },
                {
                    "sheet": "订单卡-计划002",
                    "station_code": "4",
                    "order_no": "PLAN-002",
                    "status": "PLANNED",
                    "mold_code": "",
                    "loaded_at": None,
                    "unloaded_at": None,
                    "daily_logs": [],
                },
            ]
        )
        preview = self.client.post(
            "/api/production/imports/preview/", {"file": upload}, format="multipart"
        )
        self.assertEqual(preview.status_code, 200, preview.content)
        payload = preview.json()
        self.assertEqual(payload["error_count"], 0, payload["issues"])
        self.assertEqual(payload["total_rows"], 2)
        self.assertEqual(payload["daily_log_count"], 1)

        committed = self.client.post(
            "/api/production/imports/commit/",
            {"token": payload["token"]},
            format="json",
        )
        self.assertEqual(committed.status_code, 200, committed.content)
        self.assertEqual(
            committed.json(),
            {"imported_count": 2, "log_count": 1, "settled_count": 1},
        )
        self.assertEqual(ProductionRun.objects.count(), 2)
        history_run = ProductionRun.objects.get(order_no="HISTORY-001")
        self.assertEqual(history_run.station.code, "1")
        self.assertEqual(history_run.daily_logs.count(), 1)
        self.assertTrue(history_run.is_settled)
        self.assertEqual(history_run.actual_good_quantity, 590)
        self.assertEqual(history_run.actual_defective_quantity, 10)
        self.assertEqual(history_run.total_material_kg, 20)
        self.assertEqual(history_run.labor_cost, 100)
        self.assertEqual(history_run.energy_cost, 30)
        self.assertEqual(history_run.other_cost, 0)
        self.assertEqual(history_run.settlement_notes, "历史订单结算")
        history_log = history_run.daily_logs.get()
        self.assertEqual(history_log.produced_mold_count, 100)
        self.assertEqual(history_log.notes, "历史订单日报")
        mold.refresh_from_db()
        self.assertEqual(mold.status, mold.Status.ON_MACHINE)
        self.assertEqual(mold.current_machine_id, original_machine_id)

        repeated = self.client.post(
            "/api/production/imports/commit/",
            {"token": payload["token"]},
            format="json",
        )
        self.assertEqual(repeated.status_code, 400)

    def test_preview_and_commit_accept_active_custom_station_code(self):
        machine = Machine.objects.create(
            code="D01", name="D组1号机台", is_active=True
        )
        station = ProductionStation.objects.create(
            code="D01",
            group="D",
            position_no=1,
            machine=machine,
            is_active=True,
        )
        preview = self.client.post(
            "/api/production/imports/preview/",
            {
                "file": production_workbook_upload(
                    [
                        {
                            "station_code": "d01",
                            "order_no": "CUSTOM-STATION-IMPORT",
                            "status": "PLANNED",
                        }
                    ]
                )
            },
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200, preview.content)
        self.assertEqual(preview.json()["error_count"], 0, preview.json()["issues"])

        committed = self.client.post(
            "/api/production/imports/commit/",
            {"token": preview.json()["token"]},
            format="json",
        )
        self.assertEqual(committed.status_code, 200, committed.content)
        run = ProductionRun.objects.get(order_no="CUSTOM-STATION-IMPORT")
        self.assertEqual(run.station_id, station.pk)

    def test_preview_reports_station_and_active_resource_conflicts_and_exports_report(self):
        ProductionRun.objects.create(
            station=ProductionStation.objects.get(code="1"),
            order_no="ACTIVE-001",
            specification="在产件",
            order_quantity=60,
            cavities=6,
            planned_mold_count=10,
            status=ProductionRun.Status.PLANNED,
            created_by=self.user,
        )
        preview = self.client.post(
            "/api/production/imports/preview/",
            {
                "file": production_workbook_upload(
                    [
                        {
                            "station_code": "1",
                            "order_no": "CONFLICT-001",
                            "status": "PLANNED",
                        },
                        {
                            "station_code": "MISSING-01",
                            "order_no": "BAD-STATION",
                            "status": "COMPLETED",
                            "loaded_at": timezone.now() - timedelta(hours=2),
                            "unloaded_at": timezone.now() - timedelta(hours=1),
                        },
                    ]
                )
            },
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200)
        payload = preview.json()
        self.assertGreaterEqual(payload["error_count"], 2)
        self.assertTrue(any("已有待上机" in item["message"] for item in payload["issues"]))
        self.assertTrue(any("机台不存在" in item["message"] for item in payload["issues"]))

        report = self.client.get(
            f"/api/production/imports/{payload['token']}/errors/"
        )
        self.assertEqual(report.status_code, 200)
        report_book = load_workbook(io.BytesIO(report.content))
        self.assertEqual(report_book.active.title, "生产导入预检问题")
        self.assertGreater(report_book.active.max_row, 2)

        blocked = self.client.post(
            "/api/production/imports/commit/",
            {"token": payload["token"]},
            format="json",
        )
        self.assertEqual(blocked.status_code, 400)
        batch = ProductionImportBatch.objects.get(pk=payload["token"])
        self.assertEqual(batch.status, ProductionImportBatch.Status.PREVIEWED)

    def test_new_template_allows_multiple_operators_on_same_day(self):
        now = timezone.now().replace(microsecond=0)
        mold = self.create_mold(
            asset_code="MULTI-OP-IMPORT-MOLD",
            machine_code="5",
        )
        preview = self.client.post(
            "/api/production/imports/preview/",
            {
                "file": production_workbook_upload(
                    [
                        {
                            "station_code": "5",
                            "order_no": "MULTI-OP-IMPORT",
                            "status": "RUNNING",
                            "mold_code": mold.asset_code,
                            "loaded_at": now - timedelta(hours=1),
                            "daily_logs": [
                                {
                                    "date": timezone.localdate(),
                                    "operator": "张三",
                                    "produced_mold_count": 5,
                                },
                                {
                                    "date": timezone.localdate(),
                                    "operator": "李四",
                                    "produced_mold_count": 7,
                                },
                            ],
                        }
                    ]
                )
            },
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200, preview.content)
        self.assertEqual(preview.json()["error_count"], 0, preview.json()["issues"])
        committed = self.client.post(
            "/api/production/imports/commit/",
            {"token": preview.json()["token"]},
            format="json",
        )
        self.assertEqual(committed.status_code, 200, committed.content)
        run = ProductionRun.objects.get(order_no="MULTI-OP-IMPORT")
        self.assertEqual(run.daily_logs.count(), 2)
        self.assertEqual(run.produced_mold_count, 12)

    def test_partial_settlement_block_is_rejected(self):
        now = timezone.now().replace(microsecond=0)
        preview = self.client.post(
            "/api/production/imports/preview/",
            {
                "file": production_workbook_upload(
                    [
                        {
                            "station_code": "6",
                            "order_no": "PARTIAL-SETTLEMENT",
                            "status": "COMPLETED",
                            "loaded_at": now - timedelta(hours=2),
                            "unloaded_at": now - timedelta(hours=1),
                            "actual_good_quantity": 60,
                            "daily_logs": [
                                {
                                    "date": timezone.localdate(),
                                    "operator": "张三",
                                    "produced_mold_count": 10,
                                }
                            ],
                        }
                    ]
                )
            },
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200, preview.content)
        self.assertGreater(preview.json()["error_count"], 0)
        fields = {item["field"] for item in preview.json()["issues"]}
        self.assertIn("actual_defective_quantity", fields)
        self.assertIn("labor_cost", fields)

    def test_legacy_daily_accounting_template_is_aggregated_to_settlement(self):
        now = timezone.now().replace(microsecond=0)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "旧版订单卡"
        for field, (label_cell, value_cell, label) in FIELD_CELLS.items():
            sheet[label_cell] = label
            values = {
                "station_code": "5",
                "order_no": "LEGACY-SETTLEMENT",
                "status": "COMPLETED",
                "specification": "旧模板",
                "order_quantity": 60,
                "cavities": 6,
                "planned_mold_count": 10,
                "material_unit_price": 10,
                "loaded_at": now - timedelta(hours=2),
                "unloaded_at": now - timedelta(hours=1),
            }
            value = values.get(field, "")
            if hasattr(value, "tzinfo") and value.tzinfo is not None:
                value = value.replace(tzinfo=None)
            sheet[value_cell] = value
        sheet["A2"] = "站位编号"
        for column, (_, label) in enumerate(LEGACY_DAILY_HEADERS, 1):
            sheet.cell(10, column, label)
        legacy_values = [
            timezone.localdate(),
            "张三",
            10,
            58,
            2,
            3,
            20,
            5,
            1,
            "旧日报",
        ]
        for column, value in enumerate(legacy_values, 1):
            sheet.cell(11, column, value)
        output = io.BytesIO()
        workbook.save(output)
        upload = SimpleUploadedFile(
            "legacy-production.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        preview = self.client.post(
            "/api/production/imports/preview/", {"file": upload}, format="multipart"
        )
        self.assertEqual(preview.status_code, 200, preview.content)
        self.assertEqual(preview.json()["error_count"], 0, preview.json()["issues"])
        self.assertTrue(
            any("旧版日报成本字段" in item["message"] for item in preview.json()["issues"])
        )
        committed = self.client.post(
            "/api/production/imports/commit/",
            {"token": preview.json()["token"]},
            format="json",
        )
        self.assertEqual(committed.status_code, 200, committed.content)
        run = ProductionRun.objects.get(order_no="LEGACY-SETTLEMENT")
        self.assertTrue(run.is_settled)
        self.assertEqual(run.actual_good_quantity, 58)
        self.assertEqual(run.total_cost, 56)

    def test_commit_is_transactional_when_master_data_changes_after_preview(self):
        now = timezone.now().replace(microsecond=0)
        preview = self.client.post(
            "/api/production/imports/preview/",
            {
                "file": production_workbook_upload(
                    [
                        {
                            "station_code": "1",
                            "order_no": "TX-001",
                            "status": "COMPLETED",
                            "loaded_at": now - timedelta(hours=2),
                            "unloaded_at": now - timedelta(hours=1),
                        },
                        {
                            "station_code": "2",
                            "order_no": "TX-002",
                            "status": "COMPLETED",
                            "loaded_at": now - timedelta(hours=2),
                            "unloaded_at": now - timedelta(hours=1),
                        },
                    ]
                )
            },
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.json()["error_count"], 0)
        ProductionStation.objects.filter(code="2").update(is_active=False)

        commit = self.client.post(
            "/api/production/imports/commit/",
            {"token": preview.json()["token"]},
            format="json",
        )
        self.assertEqual(commit.status_code, 400)
        self.assertFalse(ProductionRun.objects.filter(order_no__startswith="TX-").exists())
        batch = ProductionImportBatch.objects.get(pk=preview.json()["token"])
        self.assertEqual(batch.status, ProductionImportBatch.Status.PREVIEWED)

    def test_excel_status_matrix_nan_lengths_decimals_and_settlement_balance_are_validated(self):
        now = timezone.now().replace(microsecond=0)
        preview = self.client.post(
            "/api/production/imports/preview/",
            {
                "file": production_workbook_upload(
                    [
                        {
                            "sheet": "状态错误",
                            "station_code": "1",
                            "order_no": "BAD-PLANNED",
                            "status": "PLANNED",
                            "loaded_at": now,
                        },
                        {
                            "sheet": "数值错误",
                            "station_code": "2",
                            "order_no": "X" * 101,
                            "status": "COMPLETED",
                            "loaded_at": now - timedelta(hours=2),
                            "unloaded_at": now - timedelta(hours=1),
                            "strip_weight_kg": "1.2345",
                            "unit_price": "NaN",
                            "actual_good_quantity": 999,
                            "actual_defective_quantity": 0,
                            "total_material_kg": 0,
                            "labor_cost": 0,
                            "energy_cost": 0,
                            "other_cost": 0,
                            "settlement_notes": "无效结算数据",
                            "daily_logs": [
                                {
                                    "date": timezone.localdate(),
                                    "produced_mold_count": 1,
                                }
                            ],
                        },
                    ]
                )
            },
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200, preview.content)
        messages = [item["message"] for item in preview.json()["issues"]]
        self.assertTrue(any("待上机订单不能填写上模时间" in item for item in messages))
        self.assertTrue(any("不能超过100个字符" in item for item in messages))
        self.assertTrue(any("数字必须为有限值" in item for item in messages))
        self.assertTrue(any("小数位不能超过3位" in item for item in messages))
        self.assertTrue(any("实际良品与实际不良之和" in item for item in messages))

    def test_excel_completed_order_settlement_is_imported(self):
        now = timezone.now().replace(microsecond=0)
        preview = self.client.post(
            "/api/production/imports/preview/",
            {
                "file": production_workbook_upload(
                    [
                        {
                            "station_code": "3",
                            "order_no": "AUTO-GOOD",
                            "status": "COMPLETED",
                            "loaded_at": now - timedelta(hours=2),
                            "unloaded_at": now - timedelta(hours=1),
                            "actual_good_quantity": 59,
                            "actual_defective_quantity": 1,
                            "total_material_kg": 0,
                            "labor_cost": 0,
                            "energy_cost": 0,
                            "other_cost": 0,
                            "settlement_notes": "完工结算",
                            "daily_logs": [
                                {
                                    "date": timezone.localdate(),
                                    "produced_mold_count": 10,
                                }
                            ],
                        }
                    ]
                )
            },
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200, preview.content)
        self.assertEqual(preview.json()["error_count"], 0, preview.json()["issues"])
        committed = self.client.post(
            "/api/production/imports/commit/",
            {"token": preview.json()["token"]},
            format="json",
        )
        self.assertEqual(committed.status_code, 200, committed.content)
        run = ProductionRun.objects.get(order_no="AUTO-GOOD")
        self.assertEqual(run.actual_good_quantity, 59)
        self.assertEqual(run.actual_defective_quantity, 1)
        self.assertEqual(run.settlement_notes, "完工结算")
        self.assertTrue(run.is_settled)

    def test_commit_rechecks_duplicates_created_after_preview(self):
        now = timezone.now().replace(microsecond=0)
        preview = self.client.post(
            "/api/production/imports/preview/",
            {
                "file": production_workbook_upload(
                    [
                        {
                            "station_code": "3",
                            "order_no": "RACE-ORDER",
                            "status": "COMPLETED",
                            "loaded_at": now - timedelta(hours=3),
                            "unloaded_at": now - timedelta(hours=2),
                        }
                    ]
                )
            },
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200, preview.content)
        self.assertEqual(preview.json()["error_count"], 0, preview.json()["issues"])
        ProductionRun.objects.create(
            station=ProductionStation.objects.get(code="3"),
            order_no="RACE-ORDER",
            specification="并发插入",
            order_quantity=60,
            cavities=6,
            planned_mold_count=10,
            loaded_at=now - timedelta(hours=5),
            unloaded_at=now - timedelta(hours=4),
            status=ProductionRun.Status.COMPLETED,
            created_by=self.user,
        )
        committed = self.client.post(
            "/api/production/imports/commit/",
            {"token": preview.json()["token"]},
            format="json",
        )
        self.assertEqual(committed.status_code, 400, committed.content)
        self.assertEqual(
            ProductionRun.objects.filter(order_no="RACE-ORDER").count(), 1
        )
        batch = ProductionImportBatch.objects.get(pk=preview.json()["token"])
        self.assertEqual(batch.status, ProductionImportBatch.Status.PREVIEWED)

    def test_error_report_escapes_formula_like_text(self):
        preview = self.client.post(
            "/api/production/imports/preview/",
            {
                "file": production_workbook_upload(
                    [
                        {
                            "sheet": "=1+1",
                            "station_code": "D01",
                            "order_no": "FORMULA-SAFE",
                            "status": "PLANNED",
                        }
                    ]
                )
            },
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200, preview.content)
        self.assertGreater(preview.json()["error_count"], 0)
        report = self.client.get(
            f"/api/production/imports/{preview.json()['token']}/errors/"
        )
        report_book = load_workbook(io.BytesIO(report.content), data_only=False)
        sheet_cell = report_book.active["B2"]
        self.assertEqual(sheet_cell.data_type, "s")
        self.assertTrue(sheet_cell.value.startswith("'="), sheet_cell.value)

    def test_workbook_dimension_limit_is_rejected_before_parsing(self):
        workbook = load_workbook(io.BytesIO(create_production_template()))
        workbook.active.cell(2001, 1, "超出安全行数")
        output = io.BytesIO()
        workbook.save(output)
        upload = SimpleUploadedFile(
            "oversized.xlsx",
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response = self.client.post(
            "/api/production/imports/preview/", {"file": upload}, format="multipart"
        )
        self.assertEqual(response.status_code, 400, response.content)
        self.assertIn("不能超过2000行", str(response.json()))

    def test_preview_rejects_overlapping_history_on_the_same_station(self):
        now = timezone.now().replace(microsecond=0)
        preview = self.client.post(
            "/api/production/imports/preview/",
            {
                "file": production_workbook_upload(
                    [
                        {
                            "sheet": "历史一",
                            "station_code": "5",
                            "order_no": "HISTORY-A",
                            "status": "COMPLETED",
                            "loaded_at": now - timedelta(hours=4),
                            "unloaded_at": now - timedelta(hours=2),
                        },
                        {
                            "sheet": "历史二",
                            "station_code": "5",
                            "order_no": "HISTORY-B",
                            "status": "COMPLETED",
                            "loaded_at": now - timedelta(hours=3),
                            "unloaded_at": now - timedelta(hours=1),
                        },
                    ]
                )
            },
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200, preview.content)
        self.assertTrue(
            any(
                "生产时段重叠" in item["message"]
                for item in preview.json()["issues"]
            )
        )

    def test_import_daily_logs_follow_run_lifecycle_rules(self):
        today_log = {
            "date": timezone.localdate(),
            "produced_mold_count": 1,
        }
        blocked = self.client.post(
            "/api/production/imports/preview/",
            {
                "file": production_workbook_upload(
                    [
                        {
                            "sheet": "待上机带日报",
                            "station_code": "5",
                            "order_no": "PLAN-WITH-LOG",
                            "status": "PLANNED",
                            "daily_logs": [today_log],
                        },
                        {
                            "sheet": "未上模取消带日报",
                            "station_code": "6",
                            "order_no": "CANCEL-WITHOUT-LOAD",
                            "status": "CANCELLED",
                            "daily_logs": [today_log],
                        },
                    ]
                )
            },
            format="multipart",
        )
        self.assertEqual(blocked.status_code, 200, blocked.content)
        messages = [item["message"] for item in blocked.json()["issues"]]
        self.assertTrue(any("待上机订单不能填写生产日报" in item for item in messages))
        self.assertTrue(
            any("未上模即取消的订单不能填写生产日报" in item for item in messages)
        )

        now = timezone.now().replace(microsecond=0)
        allowed = self.client.post(
            "/api/production/imports/preview/",
            {
                "file": production_workbook_upload(
                    [
                        {
                            "station_code": "5",
                            "order_no": "CANCELLED-HISTORY-IMPORT",
                            "status": "CANCELLED",
                            "loaded_at": now - timedelta(hours=2),
                            "unloaded_at": now - timedelta(hours=1),
                            "daily_logs": [today_log],
                        }
                    ]
                )
            },
            format="multipart",
        )
        self.assertEqual(allowed.status_code, 200, allowed.content)
        self.assertEqual(allowed.json()["error_count"], 0, allowed.json()["issues"])
        committed = self.client.post(
            "/api/production/imports/commit/",
            {"token": allowed.json()["token"]},
            format="json",
        )
        self.assertEqual(committed.status_code, 200, committed.content)
        imported = ProductionRun.objects.get(order_no="CANCELLED-HISTORY-IMPORT")
        self.assertEqual(imported.daily_logs.count(), 1)

    def test_import_rejects_daily_dates_outside_production_window(self):
        now = timezone.now().replace(microsecond=0)
        today = timezone.localdate()
        preview = self.client.post(
            "/api/production/imports/preview/",
            {
                "file": production_workbook_upload(
                    [
                        {
                            "sheet": "生产中未来日报",
                            "station_code": "4",
                            "order_no": "RUNNING-FUTURE-LOG",
                            "status": "RUNNING",
                            "loaded_at": now,
                            "daily_logs": [
                                {
                                    "date": today + timedelta(days=1),
                                    "produced_mold_count": 1,
                                }
                            ],
                        },
                        {
                            "sheet": "完成订单越界日报",
                            "station_code": "5",
                            "order_no": "COMPLETED-OUTSIDE-LOG",
                            "status": "COMPLETED",
                            "loaded_at": now - timedelta(days=4),
                            "unloaded_at": now - timedelta(days=3),
                            "actual_good_quantity": 12,
                            "actual_defective_quantity": 0,
                            "total_material_kg": 0,
                            "labor_cost": 0,
                            "energy_cost": 0,
                            "other_cost": 0,
                            "settlement_notes": "日期边界测试结算",
                            "daily_logs": [
                                {
                                    "date": today - timedelta(days=5),
                                    "produced_mold_count": 1,
                                },
                                {
                                    "date": today - timedelta(days=2),
                                    "produced_mold_count": 1,
                                },
                            ],
                        },
                        {
                            "sheet": "取消订单下机后日报",
                            "station_code": "6",
                            "order_no": "CANCELLED-LATE-LOG",
                            "status": "CANCELLED",
                            "loaded_at": now - timedelta(days=2),
                            "unloaded_at": now - timedelta(days=1),
                            "daily_logs": [
                                {
                                    "date": today,
                                    "produced_mold_count": 1,
                                }
                            ],
                        },
                    ]
                )
            },
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200, preview.content)
        messages = [item["message"] for item in preview.json()["issues"]]
        self.assertTrue(any("不能晚于今天" in item for item in messages))
        self.assertTrue(any("不能早于上模日期" in item for item in messages))
        self.assertGreaterEqual(
            sum("不能晚于下机日期" in item for item in messages), 2
        )

    def test_preview_supports_xlsx_without_worksheet_dimension_metadata(self):
        source = production_workbook_upload(
            [
                {
                    "station_code": "5",
                    "order_no": "NO-DIMENSION-METADATA",
                    "status": "PLANNED",
                }
            ]
        )
        source.seek(0)
        original = io.BytesIO(source.read())
        rebuilt = io.BytesIO()
        removed = 0
        with ZipFile(original, "r") as input_archive, ZipFile(
            rebuilt, "w"
        ) as output_archive:
            for member in input_archive.infolist():
                content = input_archive.read(member.filename)
                if member.filename == "xl/worksheets/sheet1.xml":
                    content, removed = re.subn(
                        rb"<dimension\b[^>]*/>", b"", content, count=1
                    )
                output_archive.writestr(member, content)
        self.assertEqual(removed, 1)
        upload = SimpleUploadedFile(
            "no-dimension.xlsx",
            rebuilt.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        preview = self.client.post(
            "/api/production/imports/preview/",
            {"file": upload},
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200, preview.content)
        self.assertEqual(preview.json()["error_count"], 0, preview.json()["issues"])
        self.assertEqual(preview.json()["total_rows"], 1)
        self.assertEqual(preview.json()["rows"][0]["order_no"], "NO-DIMENSION-METADATA")
