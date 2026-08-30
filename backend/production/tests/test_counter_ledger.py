import io

from django.test import TestCase
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone
from rest_framework.test import APIClient
from openpyxl import load_workbook

from quality.models import QualityOrder
from molds.models import MoldAsset, MoldModel, RackSlot
from molds.services import seed_default_racks
from production.models import ProductionStation

from production.models import ProductionDailyLog, ProductionRun
from production.ledger_imports import create_ledger_template
from production.ocr import _extract_log_rows, _extract_task

from .helpers import ProductionTestMixin


class ProductionCounterLedgerApiTests(ProductionTestMixin, TestCase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.order = QualityOrder.objects.create(
            order_no="LEDGER-ORDER-001",
            item_no="10",
            product_name="密封件",
            specification="20x30",
            material="NBR",
            order_quantity=1000,
            created_by=cls.user,
        )

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def create_task(self, **overrides):
        payload = {
            "order_id": self.order.pk,
            "cavities": 10,
            "estimated_defect_mode": "RATE",
            "estimated_defect_rate": "8.00",
            "is_ledger_only": True,
        }
        payload.update(overrides)
        response = self.client.post("/api/production/runs/", payload, format="json")
        self.assertEqual(response.status_code, 201, response.content)
        return response.json()

    def add_counter(self, run_id, cumulative, **overrides):
        payload = {"cumulative_mold_count": cumulative}
        payload.update(overrides)
        return self.client.post(
            f"/api/production/runs/{run_id}/counter-logs/", payload, format="json"
        )

    def test_minimal_task_uses_unique_order_and_defect_rate_formula(self):
        payload = self.create_task()
        self.assertTrue(payload["is_ledger_only"])
        self.assertIsNone(payload["station"])
        self.assertIsNone(payload["mold"])
        self.assertEqual(payload["order_no"], self.order.order_no)
        self.assertEqual(payload["order_item_no"], "10")
        self.assertEqual(payload["specification"], self.order.specification)
        self.assertEqual(payload["planned_mold_count"], 108)
        self.assertEqual(payload["status"], ProductionRun.Status.PLANNED)

    def test_defect_quantity_formula(self):
        payload = self.create_task(
            estimated_defect_mode="QUANTITY",
            estimated_defect_quantity=85,
            estimated_defect_rate="0",
        )
        self.assertEqual(payload["planned_mold_count"], 109)

    def test_counter_allows_pending_operator_and_date_defaults_or_is_null(self):
        run = self.create_task()
        response = self.add_counter(run["id"], 100)
        self.assertEqual(response.status_code, 201, response.content)
        payload = response.json()
        self.assertEqual(payload["date"], timezone.localdate().isoformat())
        self.assertTrue(payload["operator_pending"])
        self.assertIn(payload["shift"], {"DAY", "NIGHT"})
        self.assertEqual(payload["produced_mold_count"], 100)
        self.assertEqual(payload["theoretical_quantity"], 1000)

        other = self.create_task(segment_no=2)
        undated = self.add_counter(other["id"], 5, date=None, operator="张三")
        self.assertEqual(undated.status_code, 201, undated.content)
        self.assertIsNone(undated.json()["date"])
        self.assertFalse(undated.json()["operator_pending"])

        fully_pending = self.add_counter(
            self.create_task(segment_no=3)["id"],
            7,
            date=None,
            shift="",
        )
        self.assertEqual(fully_pending.status_code, 201, fully_pending.content)
        self.assertIsNone(fully_pending.json()["date"])
        self.assertEqual(fully_pending.json()["shift"], "")
        self.assertTrue(fully_pending.json()["operator_pending"])

    def test_editing_historical_reading_recalculates_following_delta_atomically(self):
        run = self.create_task()
        first = self.add_counter(run["id"], 100, operator="张三").json()
        second = self.add_counter(run["id"], 200, operator="李四").json()
        changed = self.client.patch(
            f"/api/production/runs/{run['id']}/counter-logs/{first['id']}/",
            {"cumulative_mold_count": 120},
            format="json",
        )
        self.assertEqual(changed.status_code, 200, changed.content)
        second_log = ProductionDailyLog.objects.get(pk=second["id"])
        self.assertEqual(second_log.produced_mold_count, 80)

        invalid = self.client.patch(
            f"/api/production/runs/{run['id']}/counter-logs/{first['id']}/",
            {"cumulative_mold_count": 220},
            format="json",
        )
        self.assertEqual(invalid.status_code, 400, invalid.content)
        self.assertEqual(
            ProductionDailyLog.objects.get(pk=first["id"]).cumulative_mold_count,
            120,
        )
        self.assertEqual(
            ProductionDailyLog.objects.get(pk=second["id"]).produced_mold_count,
            80,
        )

    def test_handoff_readings_100_200_280_credit_100_100_80(self):
        run = self.create_task()
        values = []
        for cumulative, operator in ((100, "张三"), (200, "李四"), (280, "王五")):
            response = self.add_counter(run["id"], cumulative, operator=operator)
            self.assertEqual(response.status_code, 201, response.content)
            values.append(response.json()["produced_mold_count"])
        self.assertEqual(values, [100, 100, 80])

    def test_reset_starts_new_counter_segment(self):
        run = self.create_task()
        self.add_counter(run["id"], 200, operator="张三")
        reset = self.client.post(
            f"/api/production/runs/{run['id']}/reset-counter/",
            {"note": "断电后计数归零"},
            format="json",
        )
        self.assertEqual(reset.status_code, 200, reset.content)
        next_log = self.add_counter(run["id"], 30, operator="李四")
        self.assertEqual(next_log.status_code, 201, next_log.content)
        self.assertEqual(next_log.json()["counter_segment"], 2)
        self.assertEqual(next_log.json()["produced_mold_count"], 30)

    def test_cancel_preserves_audit_and_recalculates_next_reading(self):
        run = self.create_task()
        first = self.add_counter(run["id"], 100, operator="张三").json()
        second = self.add_counter(run["id"], 200, operator="李四").json()
        cancelled = self.client.post(
            f"/api/production/runs/{run['id']}/counter-logs/{first['id']}/cancel/",
            {"reason": "补录重复"},
            format="json",
        )
        self.assertEqual(cancelled.status_code, 200, cancelled.content)
        self.assertTrue(cancelled.json()["is_cancelled"])
        self.assertEqual(
            ProductionDailyLog.objects.get(pk=second["id"]).produced_mold_count,
            200,
        )
        audits = self.client.get(
            f"/api/production/runs/{run['id']}/record-audits/"
        )
        self.assertEqual(audits.status_code, 200)
        self.assertIn("CANCELLED", [item["action"] for item in audits.json()])

    def test_complete_ledger_requires_confirmation_below_target(self):
        run = self.create_task(planned_mold_count=100)
        self.add_counter(run["id"], 40)
        blocked = self.client.post(
            f"/api/production/runs/{run['id']}/complete-ledger/", {}, format="json"
        )
        self.assertEqual(blocked.status_code, 400, blocked.content)
        completed = self.client.post(
            f"/api/production/runs/{run['id']}/complete-ledger/",
            {"confirm_below_target": True, "note": "订单改量，人工关闭"},
            format="json",
        )
        self.assertEqual(completed.status_code, 200, completed.content)
        self.assertEqual(completed.json()["status"], ProductionRun.Status.COMPLETED)
        self.order.refresh_from_db()
        self.assertEqual(self.order.status, QualityOrder.Status.OPEN)

    def test_completed_ledger_allows_machine_and_mold_backfill_or_clear(self):
        run = self.create_task(planned_mold_count=10)
        self.add_counter(run["id"], 10)
        completed = self.client.post(
            f"/api/production/runs/{run['id']}/complete-ledger/", {}, format="json"
        )
        self.assertEqual(completed.status_code, 200, completed.content)

        mold = self.create_mold(asset_code="LEDGER-LATE-MOLD", machine_code="1")
        station = ProductionStation.objects.get(code="1")
        backfilled = self.client.patch(
            f"/api/production/runs/{run['id']}/",
            {"station_id": station.pk, "mold_id": mold.pk},
            format="json",
        )
        self.assertEqual(backfilled.status_code, 200, backfilled.content)
        self.assertEqual(backfilled.json()["station"]["id"], station.pk)
        self.assertEqual(backfilled.json()["mold"]["id"], mold.pk)

        cleared = self.client.patch(
            f"/api/production/runs/{run['id']}/",
            {"station_id": None, "mold_id": None},
            format="json",
        )
        self.assertEqual(cleared.status_code, 200, cleared.content)
        self.assertIsNone(cleared.json()["station"])
        self.assertIsNone(cleared.json()["mold"])

    def test_order_progress_accumulates_multiple_tasks_and_keeps_overproduction(self):
        first = self.create_task(planned_mold_count=50)
        second = self.create_task(segment_no=2, planned_mold_count=60)
        self.add_counter(first["id"], 50, defective_quantity=0)
        self.add_counter(second["id"], 60, defective_quantity=20)
        response = self.client.get(
            "/api/production/order-progress/", {"order_id": self.order.pk}
        )
        self.assertEqual(response.status_code, 200, response.content)
        payload = response.json()
        self.assertEqual(payload["production_quantity"], 1080)
        self.assertEqual(payload["overproduction_quantity"], 80)
        self.assertTrue(payload["production_completed"])
        self.assertEqual(payload["run_count"], 2)

    def test_selected_asset_can_save_and_reuse_its_own_default_cavities(self):
        mold = self.create_mold(asset_code="LEDGER-CAVITIES", machine_code="1")
        first = self.create_task(
            mold_id=mold.pk,
            cavities=12,
            save_cavities_as_mold_default=True,
        )
        mold.refresh_from_db()
        self.assertEqual(mold.default_cavities, 12)
        response = self.client.post(
            "/api/production/runs/",
            {
                "order_id": self.order.pk,
                "mold_id": mold.pk,
                "segment_no": 2,
                "estimated_defect_mode": "RATE",
                "estimated_defect_rate": "8.00",
                "is_ledger_only": True,
            },
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.content)
        self.assertEqual(response.json()["cavities"], 12)

    def test_simplified_excel_previews_then_commits_atomically(self):
        workbook = load_workbook(io.BytesIO(create_ledger_template()))
        task_sheet = workbook["生产任务"]
        log_sheet = workbook["交接读数"]
        task_sheet.delete_rows(2, task_sheet.max_row)
        log_sheet.delete_rows(2, log_sheet.max_row)
        task_sheet.append(
            [
                "T1",
                self.order.order_no,
                self.order.item_no,
                "",
                "",
                "",
                "",
                "",
                10,
                "比例",
                8,
                "10x20",
                "",
                "",
                60,
                "历史纸质账",
            ]
        )
        log_sheet.append(["T1", "", "白班", "张三", "", 100, 0, "否", ""])
        output = io.BytesIO()
        workbook.save(output)
        upload = SimpleUploadedFile(
            "simple-ledger.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        preview = self.client.post(
            "/api/production/ledger-imports/preview/",
            {"file": upload},
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200, preview.content)
        self.assertTrue(preview.json()["can_commit"])
        committed = self.client.post(
            "/api/production/ledger-imports/commit/",
            {"token": preview.json()["token"]},
            format="json",
        )
        self.assertEqual(committed.status_code, 200, committed.content)
        run = ProductionRun.objects.get(pk=committed.json()["created_run_ids"][0])
        self.assertTrue(run.is_ledger_only)
        self.assertEqual(run.produced_mold_count, 100)
        self.assertIsNone(run.daily_logs.get().production_date)

    def create_realtime_task(self, station_code="1", mold=None):
        mold = mold or self.create_mold(
            asset_code=f"REALTIME-{station_code}", machine_code=station_code
        )
        response = self.client.post(
            "/api/production/runs/",
            {
                "station_id": ProductionStation.objects.get(code=station_code).pk,
                "mold_id": mold.pk,
                "order_id": self.order.pk,
                "cavities": 10,
                "planned_mold_count": 100,
                "loaded_at": timezone.now().isoformat(),
            },
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.content)
        return response.json(), mold

    def test_pause_on_machine_then_resume_keeps_same_segment_and_mold(self):
        run, mold = self.create_realtime_task()
        self.add_counter(run["id"], 20)
        paused = self.client.post(
            f"/api/production/runs/{run['id']}/pause/",
            {"mode": "ON_MACHINE", "note": "临时停机"},
            format="json",
        )
        self.assertEqual(paused.status_code, 200, paused.content)
        self.assertEqual(paused.json()["status"], ProductionRun.Status.PAUSED_ON_MACHINE)
        mold.refresh_from_db()
        self.assertEqual(mold.status, MoldAsset.Status.ON_MACHINE)
        resumed = self.client.post(
            f"/api/production/runs/{run['id']}/resume/", {}, format="json"
        )
        self.assertEqual(resumed.status_code, 200, resumed.content)
        self.assertEqual(resumed.json()["id"], run["id"])
        self.assertEqual(resumed.json()["segment_no"], 1)
        self.assertEqual(resumed.json()["status"], ProductionRun.Status.RUNNING)
        self.assertEqual(resumed.json()["produced_mold_count"], 20)

    def test_pause_and_unload_then_resume_new_machine_and_mold_uses_remaining_pieces(self):
        seed_default_racks()
        run, first_mold = self.create_realtime_task()
        produced = self.add_counter(run["id"], 50)
        self.assertEqual(produced.status_code, 201, produced.content)
        destination = RackSlot.objects.select_related("zone__level__rack").get(
            zone__level__rack__code="J01",
            zone__level__level_no=1,
            zone__code="A",
            capacity_mode=2,
            position_no=1,
            stack_level=1,
        )
        paused = self.client.post(
            f"/api/production/runs/{run['id']}/pause/",
            {"mode": "UNLOADED", "slot_id": destination.pk, "note": "换急单"},
            format="json",
        )
        self.assertEqual(paused.status_code, 200, paused.content)
        self.assertEqual(paused.json()["status"], ProductionRun.Status.PAUSED_UNLOADED)
        first_mold.refresh_from_db()
        self.assertEqual(first_mold.status, MoldAsset.Status.IN_STOCK)
        self.assertEqual(first_mold.current_slot_id, destination.pk)

        second_slot = RackSlot.objects.select_related("zone__level__rack").get(
            zone__level__rack__code="J01",
            zone__level__level_no=1,
            zone__code="A",
            capacity_mode=2,
            position_no=2,
            stack_level=1,
        )
        second_model = MoldModel.objects.create(
            code="CONTINUATION-MODEL", product_name="密封件"
        )
        second_mold = MoldAsset.objects.create(
            asset_code="CONTINUATION-MOLD",
            mold_model=second_model,
            status=MoldAsset.Status.IN_STOCK,
            current_slot=second_slot,
            default_cavities=20,
        )
        resumed = self.client.post(
            f"/api/production/runs/{run['id']}/resume/",
            {
                "station_id": ProductionStation.objects.get(code="2").pk,
                "mold_id": second_mold.pk,
                "note": "换二号机继续",
            },
            format="json",
        )
        self.assertEqual(resumed.status_code, 200, resumed.content)
        payload = resumed.json()
        self.assertNotEqual(payload["id"], run["id"])
        self.assertEqual(payload["segment_no"], 2)
        self.assertEqual(payload["continuation_of"], run["id"])
        self.assertEqual(payload["cavities"], 20)
        self.assertEqual(payload["planned_mold_count"], 25)
        self.assertEqual(payload["status"], ProductionRun.Status.RUNNING)
        original = ProductionRun.objects.get(pk=run["id"])
        self.assertEqual(original.status, ProductionRun.Status.PAUSED_UNLOADED)
        self.assertEqual(original.produced_mold_count, 50)


class ProductionOcrDraftParserTests(TestCase):
    def test_one_photo_draft_extracts_task_and_multiple_counter_rows(self):
        lines = [
            {"text": "订单号：04-A001-2608210001", "confidence": 96},
            {"text": "规格 27.97X6.99", "confidence": 92},
            {"text": "材质 N7200", "confidence": 92},
            {"text": "模具孔数 10", "confidence": 93},
            {"text": "生产日期 作业员 生产模数 欠模数", "confidence": 94},
            {"text": "2026/08/20 白班 张三 100 980", "confidence": 95},
            {"text": "2026/08/21 夜班 李四 200 880", "confidence": 72},
            {"text": "王五 280 800", "confidence": 91},
        ]
        draft, confidence = _extract_task(lines)
        logs, blockers = _extract_log_rows(lines)
        self.assertEqual(draft["order_no"], "04-A001-2608210001")
        self.assertEqual(draft["cavities"], 10)
        self.assertGreaterEqual(confidence["order_no"], 85)
        self.assertEqual(
            [item["cumulative_mold_count"] for item in logs], [100, 200, 280]
        )
        self.assertEqual(logs[0]["production_date"], "2026-08-20")
        self.assertEqual(logs[0]["shift"], "DAY")
        self.assertEqual(logs[0]["operator"], "张三")
        self.assertFalse(logs[0]["blocking_items"])
        self.assertTrue(logs[1]["blocking_items"])
        self.assertEqual(blockers[0]["source_row"], 2)

    def test_missing_table_is_blocking_and_never_guessed(self):
        logs, blockers = _extract_log_rows(
            [{"text": "一张模糊照片", "confidence": 20}]
        )
        self.assertEqual(logs, [])
        self.assertEqual(blockers[0]["field"], "logs")
