import io

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook

from molds.models import MoldAsset, MoldModel
from production.imports import SETTLEMENT_CELLS, create_production_template
from production.services import seed_default_stations


class ProductionTestMixin:
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.user = get_user_model().objects.create_user(
            username="production-user", password="production-password"
        )
        seed_default_stations()

    @staticmethod
    def create_mold(asset_code="MOLD-001", machine_code="1"):
        from molds.models import Machine

        model = MoldModel.objects.create(
            code=f"MODEL-{asset_code}", product_name=f"产品 {asset_code}"
        )
        return MoldAsset.objects.create(
            asset_code=asset_code,
            mold_model=model,
            status=MoldAsset.Status.ON_MACHINE,
            current_machine=Machine.objects.get(code=machine_code),
        )


def production_workbook_upload(cards, name="production-orders.xlsx"):
    source = io.BytesIO(create_production_template())
    workbook = load_workbook(source)
    template = workbook.active
    sheets = [template]
    for _ in range(1, len(cards)):
        sheets.append(workbook.copy_worksheet(template))
    for index, (card, sheet) in enumerate(zip(cards, sheets)):
        sheet.title = card.get("sheet", f"订单卡-{index + 1:03d}")
        fixed = {
            "B2": card.get("station_code", ""),
            "D2": card.get("order_no", ""),
            "F2": card.get("status", ""),
            "H2": card.get("mold_code", ""),
            "B3": card.get("specification", "密封件"),
            "D3": card.get("material", "NBR"),
            "F3": card.get("order_quantity", 600),
            "H3": card.get("cavities", 6),
            "B4": card.get("estimated_defect_rate", 0),
            "D4": card.get("planned_mold_count", 100),
            "F4": card.get("compound_size", "10x20"),
            "H4": card.get("strip_weight_kg", 1.2),
            "B5": card.get("strips_per_batch", 6),
            "D5": card.get("curing_seconds", 72),
            "F5": card.get("estimated_hours", 2),
            "H5": card.get("operator", "张三"),
            "B6": card.get("loaded_at"),
            "D6": card.get("expected_change_at"),
            "F6": card.get("unloaded_at"),
            "B7": card.get("unit_price", 2),
            "D7": card.get("material_unit_price", 10),
            "F7": card.get("notes", ""),
        }
        for coordinate, value in fixed.items():
            if hasattr(value, "tzinfo") and value.tzinfo is not None:
                value = value.replace(tzinfo=None)
            sheet[coordinate] = value
        daily_logs = card.get("daily_logs", [])
        settlement = {
            "actual_good_quantity": card.get("actual_good_quantity"),
            "actual_defective_quantity": card.get("actual_defective_quantity"),
            "total_material_kg": card.get("total_material_kg"),
            "labor_cost": card.get("labor_cost"),
            "energy_cost": card.get("energy_cost"),
            "other_cost": card.get("other_cost"),
            "settlement_notes": card.get("settlement_notes", ""),
        }
        for field, (_, coordinate, _) in SETTLEMENT_CELLS.items():
            value = settlement[field]
            if value is not None and value != "":
                sheet[coordinate] = value
        for row_no, log in enumerate(daily_logs, 14):
            values = [
                log.get("date"),
                log.get("operator", "张三"),
                log.get("produced_mold_count", 0),
                log.get("notes", ""),
            ]
            for column, value in enumerate(values, 1):
                sheet.cell(row_no, column, value)
    output = io.BytesIO()
    workbook.save(output)
    return SimpleUploadedFile(
        name,
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
