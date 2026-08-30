"""Safe, transaction-first import helpers for the simplified shop-floor ledger."""

import io
import math
import re
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import Max, Q
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from molds.models import MoldAsset
from quality.models import QualityOrder

from .models import (
    ProductionImportBatch,
    ProductionRun,
    ProductionStation,
    normalize_production_station_code,
)
from .services import create_counter_log


TASK_HEADERS = [
    "任务标识",
    "订单号",
    "项次",
    "规格",
    "材质",
    "订单数量",
    "机台编号",
    "模具编号",
    "模具孔数",
    "预估不良方式",
    "预估不良值",
    "胶料尺寸",
    "条重kg",
    "每批条数",
    "硫化时间秒",
    "备注",
]
LOG_HEADERS = [
    "任务标识",
    "生产日期",
    "班次",
    "作业员",
    "协助人员",
    "累计模数",
    "不良数量",
    "计数已清零",
    "备注",
]
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_ROWS = 5000


def _text(value):
    return str(value or "").strip()


def _int(value, *, minimum=0):
    if value in (None, ""):
        return None
    decimal = Decimal(str(value).replace(",", "").strip())
    if decimal != decimal.to_integral_value() or decimal < minimum:
        raise ValueError(f"必须填写不小于{minimum}的整数。")
    return int(decimal)


def _decimal(value, *, minimum=Decimal("0")):
    if value in (None, ""):
        return None
    result = Decimal(str(value).replace(",", "").strip())
    if not result.is_finite() or result < minimum:
        raise ValueError(f"必须填写不小于{minimum}的数字。")
    return result


def _issue(level, message, *, sheet="", row=None, field=""):
    return {
        "level": level,
        "message": message,
        "sheet": sheet,
        "row": row,
        "field": field,
    }


def create_ledger_template():
    workbook = Workbook()
    tasks = workbook.active
    tasks.title = "生产任务"
    logs = workbook.create_sheet("交接读数")
    for sheet, headers in ((tasks, TASK_HEADERS), (logs, LOG_HEADERS)):
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1677FF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{chr(64 + len(headers))}1"
        for index, header in enumerate(headers, 1):
            sheet.column_dimensions[chr(64 + index)].width = max(12, len(header) * 2 + 3)
    tasks.append(
        [
            "任务1",
            "示例订单号",
            "1",
            "示例规格",
            "NBR",
            10000,
            "1",
            "",
            10,
            "比例",
            8,
            "",
            "",
            "",
            "",
            "删除本示例行后填写",
        ]
    )
    logs.append(["任务1", "", "白班", "张三", "", 100, 0, "否", "日期可留空"])
    tasks["A1"].comment = Comment(
        "核心必填：任务标识、订单号＋项次、模具孔数。机台、模具、规格、材质等可留空后补；规格材质和订单数量会从订单台账带出。",
        "ERP",
    )
    logs["A1"].comment = Comment(
        "核心必填：任务标识、累计模数。生产日期、班次、作业员、协助人员和不良数量均可留空后补；日期留空会保留为“日期未记录”。",
        "ERP",
    )
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _load(uploaded_file):
    size = getattr(uploaded_file, "size", None)
    if size is not None and size > MAX_UPLOAD_BYTES:
        raise ValueError("Excel文件不能超过10MB。")
    uploaded_file.seek(0)
    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("无法读取Excel，请确认文件为有效的.xlsx格式。") from exc
    if "生产任务" not in workbook.sheetnames or "交接读数" not in workbook.sheetnames:
        raise ValueError("缺少“生产任务”或“交接读数”工作表，请使用新版简化模板。")
    return workbook


def _rows(sheet, headers, issues):
    actual = [_text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    if actual[: len(headers)] != headers:
        issues.append(_issue("error", "表头与新版简化模板不一致。", sheet=sheet.title, row=1))
        return []
    result = []
    for row_no, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if row_no > MAX_ROWS + 1:
            issues.append(_issue("error", "单个工作表不能超过5000条数据。", sheet=sheet.title))
            break
        values = list(values[: len(headers)]) + [None] * max(0, len(headers) - len(values))
        if not any(value not in (None, "") for value in values):
            continue
        result.append((row_no, values))
    return result


def preview_ledger_workbook(uploaded_file, user):
    workbook = _load(uploaded_file)
    issues = []
    tasks = []
    task_keys = set()
    for row_no, values in _rows(workbook["生产任务"], TASK_HEADERS, issues):
        key, order_no, item_no, specification, material = map(_text, values[:5])
        try:
            order_quantity = _int(values[5], minimum=1)
            cavities = _int(values[8], minimum=1)
            defect_value = _decimal(values[10]) or Decimal("0")
        except (ValueError, InvalidOperation) as exc:
            issues.append(_issue("error", str(exc), sheet="生产任务", row=row_no))
            continue
        if not key:
            issues.append(_issue("error", "任务标识不能为空。", sheet="生产任务", row=row_no, field="任务标识"))
        elif key in task_keys:
            issues.append(_issue("error", "任务标识重复。", sheet="生产任务", row=row_no, field="任务标识"))
        task_keys.add(key)
        if not order_no:
            issues.append(_issue("error", "订单号不能为空。", sheet="生产任务", row=row_no, field="订单号"))
        order_qs = QualityOrder.objects.filter(order_no=order_no, item_no=item_no)
        if order_qs.count() != 1:
            issues.append(
                _issue(
                    "error",
                    "订单号＋项次必须唯一匹配订单台账中的一条记录。",
                    sheet="生产任务",
                    row=row_no,
                    field="订单号/项次",
                )
            )
            order = None
        else:
            order = order_qs.first()
            order_quantity = order.order_quantity
            specification = specification or order.specification
            material = material or order.material
        station_code = normalize_production_station_code(values[6])
        station = ProductionStation.objects.filter(code__iexact=station_code).first() if station_code else None
        if station_code and station is None:
            issues.append(_issue("error", "机台编号不存在。", sheet="生产任务", row=row_no, field="机台编号"))
        mold_code = _text(values[7])
        mold = MoldAsset.objects.filter(asset_code__iexact=mold_code, is_active=True).first() if mold_code else None
        if mold_code and mold is None:
            issues.append(_issue("error", "模具编号不存在或已停用。", sheet="生产任务", row=row_no, field="模具编号"))
        if cavities is None and mold and mold.default_cavities:
            cavities = mold.default_cavities
        if cavities is None:
            issues.append(_issue("error", "未选择有默认孔数的模具时，模具孔数不能为空。", sheet="生产任务", row=row_no, field="模具孔数"))
            cavities = 1
        mode_text = _text(values[9]).upper()
        mode = ProductionRun.DefectEstimateMode.QUANTITY if mode_text in {"件数", "QUANTITY"} else ProductionRun.DefectEstimateMode.RATE
        if mode == ProductionRun.DefectEstimateMode.RATE and defect_value > 100:
            issues.append(_issue("error", "预估不良率不能超过100%。", sheet="生产任务", row=row_no, field="预估不良值"))
        if mode == ProductionRun.DefectEstimateMode.QUANTITY:
            planned = math.ceil((order_quantity + int(defect_value)) / cavities)
        else:
            planned = math.ceil((Decimal(order_quantity) / Decimal(cavities)) * (Decimal("1") + defect_value / Decimal("100")))
        sibling = ProductionRun.objects.filter(order_id=order.pk if order else None)
        if station:
            sibling = sibling.filter(station=station)
        if mold:
            sibling = sibling.filter(mold=mold)
        if order and sibling.exists():
            issues.append(_issue("warning", "系统中已有相似生产任务，请确认不是重复补录。", sheet="生产任务", row=row_no))
        tasks.append(
            {
                "key": key,
                "order_id": order.pk if order else None,
                "order_no": order_no,
                "item_no": item_no,
                "specification": specification,
                "material": material,
                "order_quantity": order_quantity,
                "station_id": station.pk if station else None,
                "mold_id": mold.pk if mold else None,
                "cavities": cavities,
                "estimated_defect_mode": mode,
                "estimated_defect_rate": str(defect_value if mode == ProductionRun.DefectEstimateMode.RATE else Decimal("0")),
                "estimated_defect_quantity": int(defect_value) if mode == ProductionRun.DefectEstimateMode.QUANTITY else 0,
                "planned_mold_count": planned,
                "compound_size": _text(values[11]),
                "strip_weight_kg": str(_decimal(values[12])) if values[12] not in (None, "") else None,
                "strips_per_batch": _int(values[13], minimum=1) if values[13] not in (None, "") else None,
                "curing_seconds": _int(values[14]) or 0,
                "notes": _text(values[15]),
                "logs": [],
            }
        )

    by_key = {task["key"]: task for task in tasks}
    counter_state = {}
    for row_no, values in _rows(workbook["交接读数"], LOG_HEADERS, issues):
        key = _text(values[0])
        task = by_key.get(key)
        if task is None:
            issues.append(_issue("error", "任务标识在生产任务表中不存在。", sheet="交接读数", row=row_no, field="任务标识"))
            continue
        raw_date = values[1]
        production_date = None
        if raw_date not in (None, ""):
            if hasattr(raw_date, "date"):
                production_date = raw_date.date().isoformat()
            else:
                from django.utils.dateparse import parse_date
                parsed = parse_date(_text(raw_date))
                if parsed is None:
                    issues.append(_issue("error", "日期格式应为yyyy-mm-dd或留空。", sheet="交接读数", row=row_no, field="生产日期"))
                else:
                    production_date = parsed.isoformat()
        shift_text = _text(values[2]).upper()
        shift = "DAY" if shift_text in {"白班", "DAY"} else "NIGHT" if shift_text in {"夜班", "NIGHT"} else ""
        operator = _text(values[3])
        try:
            cumulative = _int(values[5], minimum=1)
            defective = _int(values[6]) or 0
        except (ValueError, InvalidOperation) as exc:
            issues.append(_issue("error", str(exc), sheet="交接读数", row=row_no))
            continue
        reset = _text(values[7]).lower() in {"是", "yes", "y", "1", "true"}
        state = counter_state.setdefault(key, {"segment": 1, "previous": 0})
        if reset:
            state["segment"] += 1
            state["previous"] = 0
        produced = cumulative - state["previous"]
        if produced < 1:
            issues.append(_issue("error", f"累计读数必须大于本分段上一读数{state['previous']}。", sheet="交接读数", row=row_no, field="累计模数"))
        if defective > max(produced, 0) * task["cavities"]:
            issues.append(_issue("error", "不良数量超过本次理论产量。", sheet="交接读数", row=row_no, field="不良数量"))
        state["previous"] = cumulative
        task["logs"].append(
            {
                "production_date": production_date,
                "shift": shift,
                "operator": operator,
                "assistants": [name.strip() for name in re.split(r"[,，、]", _text(values[4])) if name.strip()],
                "cumulative_mold_count": cumulative,
                "defective_quantity": defective,
                "reset_before": reset,
                "notes": _text(values[8]),
            }
        )

    errors = [item for item in issues if item["level"] == "error"]
    warnings = [item for item in issues if item["level"] == "warning"]
    batch = ProductionImportBatch.objects.create(
        original_name=getattr(uploaded_file, "name", "production-ledger.xlsx")[:255],
        payload={"format": "LEDGER_V2", "tasks": tasks},
        errors=errors,
        warnings=warnings,
        created_by=user,
    )
    return {
        "token": str(batch.pk),
        "can_commit": not errors,
        "tasks": tasks,
        "errors": errors,
        "warnings": warnings,
    }


@transaction.atomic
def commit_ledger_batch(batch, user, *, confirm_warnings=False):
    batch = ProductionImportBatch.objects.select_for_update().get(pk=batch.pk)
    if batch.status == ProductionImportBatch.Status.COMMITTED:
        return {"token": str(batch.pk), "status": batch.status, "created_run_ids": batch.payload.get("created_run_ids", [])}
    if batch.errors:
        raise ValueError("预检存在阻断错误，不能导入。")
    if batch.warnings and not confirm_warnings:
        raise ValueError("预检存在疑似重复记录；核对后请确认警告再整批导入。")
    if batch.payload.get("format") != "LEDGER_V2":
        raise ValueError("该批次不是新版生产手工账预检结果。")
    batch.status = ProductionImportBatch.Status.COMMITTING
    batch.save(update_fields=["status"])
    created = []
    for task in batch.payload["tasks"]:
        order = QualityOrder.objects.select_for_update().get(pk=task["order_id"])
        station = ProductionStation.objects.filter(pk=task["station_id"]).first() if task["station_id"] else None
        mold = MoldAsset.objects.select_for_update().filter(pk=task["mold_id"]).first() if task["mold_id"] else None
        segment = (ProductionRun.objects.filter(order=order).aggregate(value=Max("segment_no"))["value"] or 0) + 1
        run = ProductionRun.objects.create(
            station=station,
            order=order,
            order_no=order.order_no,
            specification=task["specification"],
            material=task["material"],
            mold=mold,
            order_quantity=order.order_quantity,
            cavities=task["cavities"],
            estimated_defect_mode=task["estimated_defect_mode"],
            estimated_defect_rate=Decimal(task["estimated_defect_rate"]),
            estimated_defect_quantity=task["estimated_defect_quantity"],
            planned_mold_count=task["planned_mold_count"],
            compound_size=task["compound_size"],
            strip_weight_kg=Decimal(task["strip_weight_kg"]) if task["strip_weight_kg"] else None,
            strips_per_batch=task["strips_per_batch"],
            curing_seconds=task["curing_seconds"],
            status=ProductionRun.Status.PLANNED,
            is_ledger_only=True,
            notes=task["notes"],
            segment_no=segment,
            created_by=user,
        )
        for record in task["logs"]:
            if record["reset_before"]:
                run.counter_segment += 1
                run.save(update_fields=["counter_segment", "updated_at"])
            from .models import ProductionEmployee
            assistants = []
            for assistant_name in record["assistants"]:
                employee = ProductionEmployee.objects.filter(name__iexact=assistant_name).first()
                assistants.append(employee or ProductionEmployee.objects.create(name=assistant_name))
            from datetime import date
            create_counter_log(
                run,
                user,
                {
                    "production_date": date.fromisoformat(record["production_date"]) if record["production_date"] else None,
                    "shift": record["shift"],
                    "operator": record["operator"],
                    "cumulative_mold_count": record["cumulative_mold_count"],
                    "cavities_snapshot": task["cavities"],
                    "defective_quantity": record["defective_quantity"],
                    "notes": record["notes"],
                    "assistant_operators": assistants,
                },
            )
        created.append(run.pk)
    payload = dict(batch.payload)
    payload["created_run_ids"] = created
    batch.payload = payload
    batch.status = ProductionImportBatch.Status.COMMITTED
    batch.committed_at = timezone.now()
    batch.save(update_fields=["payload", "status", "committed_at"])
    return {"token": str(batch.pk), "status": batch.status, "created_run_ids": created}
