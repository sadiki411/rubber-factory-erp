import io
import math
import re
from collections import Counter
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from zipfile import BadZipFile, ZipFile

from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import IntegrityError, OperationalError, transaction
from django.db.models import Q
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation

from molds.models import MoldAsset
from orders.models import ProductSpecification
from quality.models import QualityOrder

from .models import (
    ProductionDailyLog,
    ProductionImportBatch,
    ProductionRun,
    ProductionSettlementRevision,
    ProductionStation,
    normalize_operator,
    normalize_production_station_code,
)
from .services import record_settlement_revision


FIELD_CELLS = {
    "station_code": ("A2", "B2", "机台编号"),
    "order_no": ("C2", "D2", "订单号"),
    "status": ("E2", "F2", "状态"),
    "mold_code": ("G2", "H2", "模具编号"),
    "specification": ("A3", "B3", "规格"),
    "material": ("C3", "D3", "材质"),
    "order_quantity": ("E3", "F3", "订单数量"),
    "cavities": ("G3", "H3", "模具孔数"),
    "estimated_defect_rate": ("A4", "B4", "预估不良率(%)"),
    "planned_mold_count": ("C4", "D4", "计划生产模数"),
    "compound_size": ("E4", "F4", "胶料尺寸"),
    "strip_weight_kg": ("G4", "H4", "条重(kg)"),
    "strips_per_batch": ("A5", "B5", "每批条数"),
    "curing_seconds": ("C5", "D5", "硫化时间(秒)"),
    "estimated_hours": ("E5", "F5", "预计生产工时"),
    "operator": ("G5", "H5", "作业员"),
    "loaded_at": ("A6", "B6", "上模时间"),
    "expected_change_at": ("C6", "D6", "预计换模时间"),
    "unloaded_at": ("E6", "F6", "下机时间"),
    "unit_price": ("A7", "B7", "产品单价"),
    "material_unit_price": ("C7", "D7", "材料单价(元/kg)"),
    "notes": ("E7", "F7", "备注"),
}

FIELD_LABEL_ALIASES = {
    "station_code": {"机台编号", "站位编号"},
}

DAILY_HEADERS = [
    ("production_date", "生产日期"),
    ("operator", "作业员"),
    ("produced_mold_count", "生产模数"),
    ("notes", "备注"),
]

LEGACY_DAILY_HEADERS = [
    ("production_date", "生产日期"),
    ("operator", "作业员"),
    ("produced_mold_count", "生产模数"),
    ("good_quantity", "良品数量"),
    ("defective_quantity", "不良数量"),
    ("material_kg", "材料用量(kg)"),
    ("labor_cost", "人工成本"),
    ("energy_cost", "能耗成本"),
    ("other_cost", "其他成本"),
    ("notes", "备注"),
]

SETTLEMENT_CELLS = {
    "actual_good_quantity": ("A9", "B9", "实际良品"),
    "actual_defective_quantity": ("C9", "D9", "实际不良"),
    "total_material_kg": ("E9", "F9", "总材料(kg)"),
    "labor_cost": ("A10", "B10", "人工成本"),
    "energy_cost": ("C10", "D10", "能耗成本"),
    "other_cost": ("E10", "F10", "其他成本"),
    "settlement_notes": ("G9", "H9", "结算备注"),
}

STATUS_ALIASES = {
    "": "",
    "PLANNED": ProductionRun.Status.PLANNED,
    "待上机": ProductionRun.Status.PLANNED,
    "计划": ProductionRun.Status.PLANNED,
    "RUNNING": ProductionRun.Status.RUNNING,
    "生产中": ProductionRun.Status.RUNNING,
    "在产": ProductionRun.Status.RUNNING,
    "COMPLETED": ProductionRun.Status.COMPLETED,
    "已完成": ProductionRun.Status.COMPLETED,
    "完成": ProductionRun.Status.COMPLETED,
    "CANCELLED": ProductionRun.Status.CANCELLED,
    "已取消": ProductionRun.Status.CANCELLED,
    "取消": ProductionRun.Status.CANCELLED,
}

MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_ARCHIVE_FILES = 1000
MAX_WORKSHEETS = 100
MAX_ROWS_PER_SHEET = 2000
MAX_COLUMNS_PER_SHEET = 50
MAX_WORKBOOK_CELLS = 200_000
MAX_POSITIVE_INTEGER = 2_147_483_647
MAX_POSITIVE_SMALL_INTEGER = 32_767


def _text(value):
    return "" if value is None else str(value).strip()


def _text_value(value, issues, sheet, row, field, *, max_length, required=False):
    result = _text(value)
    if required and not result:
        issues.append(
            _issue("error", "该字段不能为空。", sheet=sheet, row=row, field=field)
        )
    if len(result) > max_length:
        issues.append(
            _issue(
                "error",
                f"内容不能超过{max_length}个字符。",
                sheet=sheet,
                row=row,
                field=field,
            )
        )
    return result


def _decimal_shape(value):
    _, digit_tuple, exponent = value.as_tuple()
    if exponent >= 0:
        digits = len(digit_tuple) + exponent
        decimals = 0
    else:
        digits = len(digit_tuple)
        decimals = abs(exponent)
        if decimals > digits:
            digits = decimals
    return digits, decimals


def _validate_workbook_archive(uploaded_file):
    size = getattr(uploaded_file, "size", None)
    if size is not None and size > MAX_UPLOAD_BYTES:
        raise ValueError("Excel文件不能超过10MB。")
    uploaded_file.seek(0)
    try:
        with ZipFile(uploaded_file) as archive:
            members = archive.infolist()
            if len(members) > MAX_ARCHIVE_FILES:
                raise ValueError("Excel内部文件数量过多，无法安全预检。")
            if sum(member.file_size for member in members) > MAX_UNCOMPRESSED_BYTES:
                raise ValueError("Excel解压后内容超过50MB，无法安全预检。")
    except BadZipFile:
        raise
    finally:
        uploaded_file.seek(0)


def _validate_workbook_dimensions(workbook):
    if len(workbook.worksheets) > MAX_WORKSHEETS:
        raise ValueError(f"工作表不能超过{MAX_WORKSHEETS}个。")
    total_cells = 0
    for sheet in workbook.worksheets:
        max_row = sheet.max_row
        max_column = sheet.max_column
        if max_row is None or max_column is None:
            dimension = sheet.calculate_dimension(force=True)
            _, _, calculated_max_column, calculated_max_row = range_boundaries(
                dimension
            )
            max_row = sheet.max_row or calculated_max_row or 0
            max_column = sheet.max_column or calculated_max_column or 0
        if max_row > MAX_ROWS_PER_SHEET:
            raise ValueError(
                f"工作表“{sheet.title}”不能超过{MAX_ROWS_PER_SHEET}行。"
            )
        if max_column > MAX_COLUMNS_PER_SHEET:
            raise ValueError(
                f"工作表“{sheet.title}”不能超过{MAX_COLUMNS_PER_SHEET}列。"
            )
        total_cells += max_row * max_column
        if total_cells > MAX_WORKBOOK_CELLS:
            raise ValueError("工作簿单元格范围过大，无法安全预检。")


def _issue(level, message, *, sheet="", row=None, field=""):
    return {
        "level": level,
        "sheet": sheet,
        "row": row,
        "field": field,
        "message": message,
    }


def _normalized_order_match_value(value):
    return str(value or "").strip().casefold()


def _order_match_key(order_no, specification, material):
    return (
        _normalized_order_match_value(order_no),
        _normalized_order_match_value(specification),
        _normalized_order_match_value(material),
    )


def _resolve_order_links(records, issues, *, for_update=False):
    """Attach only unambiguous order/specification references to import rows.

    The production card values remain the authoritative historical snapshots.
    A master record is used only as a reference and never overwrites those
    imported values.
    """

    order_numbers = {
        str(record.get("order_no") or "").strip()
        for record in records
        if str(record.get("order_no") or "").strip()
    }
    candidate_query = Q()
    for order_no in order_numbers:
        candidate_query |= Q(order_no__iexact=order_no)

    candidates = QualityOrder.objects.none()
    if candidate_query:
        candidates = QualityOrder.objects.filter(candidate_query)
        if for_update:
            candidates = candidates.select_for_update()

    by_key = {}
    for order in candidates:
        key = _order_match_key(order.order_no, order.specification, order.material)
        by_key.setdefault(key, []).append(order)

    for record in records:
        record["order_id"] = None
        record["product_specification_id"] = None
        if not record.get("order_no") or not record.get("specification"):
            continue
        key = _order_match_key(
            record.get("order_no"),
            record.get("specification"),
            record.get("material"),
        )
        matches = by_key.get(key, [])
        if len(matches) == 1:
            order = matches[0]
            record["order_id"] = order.pk
            record["product_specification_id"] = order.product_specification_id
            if order.product_specification_id is None:
                issues.append(
                    _issue(
                        "warning",
                        "已唯一匹配订单明细，但该订单尚未关联产品规格，生产记录的产品规格关联将保持为空。",
                        sheet=record.get("sheet", ""),
                        row=2,
                        field="product_specification_id",
                    )
                )
            continue

        if matches:
            message = (
                f"订单号、规格和材质匹配到{len(matches)}条订单明细，"
                "无法唯一确定，生产记录将保持未关联。"
            )
        else:
            message = (
                "未找到订单号、规格和材质均匹配的订单明细，"
                "生产记录将保持未关联。"
            )
        issues.append(
            _issue(
                "warning",
                message,
                sheet=record.get("sheet", ""),
                row=2,
                field="order_no",
            )
        )


def _apply_card_style(ws):
    dark = "1F4E78"
    medium = "5B9BD5"
    pale = "D9EAF7"
    input_fill = "FFFBE6"
    white = "FFFFFF"
    gray = "666666"
    thin_gray = Side(style="thin", color="B7C9D6")
    medium_blue = Side(style="medium", color=dark)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A14"
    ws.merge_cells("A1:J1")
    ws["A1"] = "生产订单统计表（可直接导入ERP）"
    ws["A1"].font = Font(name="微软雅黑", size=16, bold=True, color=white)
    ws["A1"].fill = PatternFill("solid", fgColor=dark)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for field, (label_cell, value_cell, label) in FIELD_CELLS.items():
        ws[label_cell] = label
        ws[label_cell].font = Font(name="微软雅黑", bold=True, color=dark)
        ws[label_cell].fill = PatternFill("solid", fgColor=pale)
        ws[label_cell].alignment = Alignment(horizontal="center", vertical="center")
        ws[label_cell].border = Border(
            left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray
        )
        ws[value_cell].fill = PatternFill("solid", fgColor=input_fill)
        ws[value_cell].font = Font(name="微软雅黑", color="000000")
        ws[value_cell].alignment = Alignment(
            horizontal="left" if field in {"specification", "material", "compound_size", "notes"} else "center",
            vertical="center",
            wrap_text=True,
        )
        ws[value_cell].border = Border(
            left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray
        )

    ws.merge_cells("F6:H6")
    ws.merge_cells("F7:H7")
    guide_rows = [
        (2, "导入规则"),
        (3, "填写系统中已启用的机台编号"),
        (4, "时间:年-月-日 时:分"),
        (5, "状态可留空自动判断"),
        (6, "每页一张订单"),
        (7, "复制页批量导入"),
    ]
    for row, text in guide_rows:
        ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=10)
        cell = ws.cell(row, 9, text)
        cell.font = Font(
            name="微软雅黑",
            size=10,
            bold=row == 2,
            color=white if row == 2 else dark,
        )
        cell.fill = PatternFill("solid", fgColor=medium if row == 2 else "EEF5FA")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray
        )
    for row in range(2, 8):
        ws.row_dimensions[row].height = 25

    ws.merge_cells("A8:J8")
    ws["A8"] = "订单完工结算（完成下机后统一填写；生产期间无需每日填写成本）"
    ws["A8"].font = Font(name="微软雅黑", size=12, bold=True, color=white)
    ws["A8"].fill = PatternFill("solid", fgColor=medium)
    ws["A8"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[8].height = 24

    for field, (label_cell, value_cell, label) in SETTLEMENT_CELLS.items():
        ws[label_cell] = label
        ws[label_cell].font = Font(name="微软雅黑", bold=True, color=dark)
        ws[label_cell].fill = PatternFill("solid", fgColor=pale)
        ws[label_cell].alignment = Alignment(horizontal="center", vertical="center")
        ws[label_cell].border = Border(
            left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray
        )
        ws[value_cell].fill = PatternFill("solid", fgColor=input_fill)
        ws[value_cell].font = Font(name="微软雅黑", color="000000")
        ws[value_cell].alignment = Alignment(
            horizontal="left" if field == "settlement_notes" else "center",
            vertical="center",
            wrap_text=True,
        )
        ws[value_cell].border = Border(
            left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray
        )
    ws.merge_cells("H9:J10")
    ws.row_dimensions[9].height = 25
    ws.row_dimensions[10].height = 25

    ws.merge_cells("A11:J11")
    ws["A11"] = (
        "填写说明：黄色单元格为可填写项；结算区留空表示尚未结算；如填写结算，六项数字必须全部填写，"
        "实际良品＋实际不良必须等于累计生产模数×模具孔数。"
        "同一订单同一天有多名作业员时，每人单独占一行填写本人实际完成模数；"
        "每一模只能计入一名作业员，禁止重复计算同一模次。"
    )
    ws["A11"].font = Font(name="微软雅黑", size=9, italic=True, color=gray)
    ws["A11"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[11].height = 42

    ws.merge_cells("A12:J12")
    ws["A12"] = "每日人员生产记录"
    ws["A12"].font = Font(name="微软雅黑", size=12, bold=True, color=white)
    ws["A12"].fill = PatternFill("solid", fgColor=medium)
    ws["A12"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[12].height = 24

    for column, (_, label) in enumerate(DAILY_HEADERS, 1):
        cell = ws.cell(13, column, label)
        cell.font = Font(name="微软雅黑", bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=thin_gray, right=thin_gray, top=medium_blue, bottom=medium_blue
        )
    for column in range(5, 11):
        cell = ws.cell(13, column)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.border = Border(
            left=thin_gray, right=thin_gray, top=medium_blue, bottom=medium_blue
        )
    ws.row_dimensions[13].height = 30

    for row in range(14, 44):
        for column in range(1, 11):
            cell = ws.cell(row, column)
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(
                horizontal="left" if column in {2, 4} else "center",
                vertical="center",
                wrap_text=column == 4,
            )
            cell.border = Border(
                left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray
            )
            if row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F4F8FB")
        ws.row_dimensions[row].height = 22
    ws.merge_cells("D13:J13")
    for row in range(14, 44):
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=10)

    widths = {
        "A": 15,
        "B": 16,
        "C": 15,
        "D": 18,
        "E": 14,
        "F": 17,
        "G": 14,
        "H": 15,
        "I": 14,
        "J": 14,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws["B6"].number_format = "yyyy-mm-dd hh:mm"
    ws["D6"].number_format = "yyyy-mm-dd hh:mm"
    ws["F6"].number_format = "yyyy-mm-dd hh:mm"
    for row in range(14, 44):
        ws.cell(row, 1).number_format = "yyyy-mm-dd"
    for cell in ("B4", "H4"):
        ws[cell].number_format = "0.000"
    for cell in ("F5", "B7", "D7"):
        ws[cell].number_format = "#,##0.00"
    ws["F9"].number_format = "0.000"
    for cell in ("B10", "D10", "F10"):
        ws[cell].number_format = "#,##0.00"

    status_validation = DataValidation(
        type="list",
        formula1='"PLANNED,RUNNING,COMPLETED,CANCELLED"',
        allow_blank=True,
    )
    ws.add_data_validation(status_validation)
    status_validation.add(ws["F2"])
    station_validation = DataValidation(
        type="custom",
        formula1='=LEN(TRIM(B2))>0',
        allow_blank=False,
    )
    station_validation.error = "请填写系统中已启用的机台编号。"
    station_validation.errorTitle = "机台编号不能为空"
    station_validation.showErrorMessage = True
    ws.add_data_validation(station_validation)
    station_validation.add(ws["B2"])
    nonnegative = DataValidation(
        type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True
    )
    ws.add_data_validation(nonnegative)
    for coordinate in [
        "F3",
        "B4",
        "D4",
        "H4",
        "B5",
        "D5",
        "F5",
        "B7",
        "D7",
        "B9",
        "D9",
        "F9",
        "B10",
        "D10",
        "F10",
    ]:
        nonnegative.add(ws[coordinate])

    positive_molds = DataValidation(
        type="whole", operator="greaterThanOrEqual", formula1="1", allow_blank=True
    )
    ws.add_data_validation(positive_molds)
    positive_molds.add("C14:C43")

    ws["B2"].comment = Comment(
        "必填。填写系统中当前已启用的机台编号；默认机台为1至6，"
        "也支持D01等自定义编号。旧编号A01/A02/B01/B02/C01/C02仍可兼容导入。",
        "ERP",
    )
    ws["D2"].comment = Comment("必填，建议使用业务订单号且不要重复导入。", "ERP")
    ws["F2"].comment = Comment(
        "可留空：有下机时间自动判定COMPLETED，有上模时间自动判定RUNNING，否则为PLANNED。",
        "ERP",
    )
    ws["H2"].comment = Comment("选填，填写模具实体编号；导入不会改变模具台账状态。", "ERP")
    ws["D4"].comment = Comment("可留空，系统会按订单数量、孔数和预估不良率计算。", "ERP")
    ws["F5"].comment = Comment("可留空，系统会按计划模数×硫化秒数计算。", "ERP")
    ws["C13"].comment = Comment(
        "必填，填写该作业员当天实际完成的模数，必须大于0；每一模只能计入一名作业员。",
        "ERP",
    )
    ws["D13"].comment = Comment("选填，可记录班次、异常、交接等说明。", "ERP")

    ws.auto_filter.ref = "A13:C43"
    ws.print_title_rows = "1:13"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def create_production_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "订单卡-001"
    _apply_card_style(sheet)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _normalize_station_code(value):
    return normalize_production_station_code(_text(value))


def _decimal_value(
    value,
    issues,
    sheet,
    row,
    field,
    *,
    required=False,
    minimum=Decimal("0"),
    max_digits=None,
    decimal_places=None,
):
    if value in (None, ""):
        if required:
            issues.append(_issue("error", "该字段不能为空。", sheet=sheet, row=row, field=field))
        return None
    try:
        result = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError, TypeError):
        issues.append(_issue("error", "必须填写有效数字。", sheet=sheet, row=row, field=field))
        return None
    if not result.is_finite():
        issues.append(
            _issue("error", "数字必须为有限值。", sheet=sheet, row=row, field=field)
        )
        return None
    if result < minimum:
        issues.append(
            _issue("error", f"数值不能小于{minimum}。", sheet=sheet, row=row, field=field)
        )
    digits, decimals = _decimal_shape(result)
    if max_digits is not None and digits > max_digits:
        issues.append(
            _issue(
                "error",
                f"数字总位数不能超过{max_digits}位。",
                sheet=sheet,
                row=row,
                field=field,
            )
        )
    if decimal_places is not None and decimals > decimal_places:
        issues.append(
            _issue(
                "error",
                f"小数位不能超过{decimal_places}位。",
                sheet=sheet,
                row=row,
                field=field,
            )
        )
    return result


def _integer_value(
    value,
    issues,
    sheet,
    row,
    field,
    *,
    required=False,
    minimum=0,
    maximum=MAX_POSITIVE_INTEGER,
):
    decimal_value = _decimal_value(
        value,
        issues,
        sheet,
        row,
        field,
        required=required,
        minimum=Decimal(minimum),
    )
    if decimal_value is None:
        return None
    if decimal_value != decimal_value.to_integral_value():
        issues.append(_issue("error", "必须填写整数。", sheet=sheet, row=row, field=field))
        return None
    if decimal_value > maximum:
        issues.append(
            _issue(
                "error",
                f"数值不能大于{maximum}。",
                sheet=sheet,
                row=row,
                field=field,
            )
        )
        return None
    return int(decimal_value)


def _parse_datetime_value(value, issues, sheet, row, field):
    if value in (None, ""):
        return None
    parsed = None
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time.min)
    else:
        text = _text(value)
        parsed = parse_datetime(text)
        if parsed is None:
            parsed_date = parse_date(text)
            if parsed_date:
                parsed = datetime.combine(parsed_date, time.min)
        if parsed is None:
            for pattern in (
                "%Y/%m/%d %H:%M",
                "%Y-%m-%d %H:%M",
                "%Y/%m/%d %H:%M:%S",
                "%Y年%m月%d日 %H:%M",
            ):
                try:
                    parsed = datetime.strptime(text, pattern)
                    break
                except ValueError:
                    continue
    if parsed is None:
        issues.append(
            _issue(
                "error",
                "日期时间格式无效，建议使用yyyy-mm-dd hh:mm。",
                sheet=sheet,
                row=row,
                field=field,
            )
        )
        return None
    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed


def _parse_date_value(value, issues, sheet, row, field):
    if value in (None, ""):
        issues.append(_issue("error", "生产日期不能为空。", sheet=sheet, row=row, field=field))
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    parsed = parse_date(text)
    if parsed is None:
        for pattern in ("%Y/%m/%d", "%Y年%m月%d日"):
            try:
                parsed = datetime.strptime(text, pattern).date()
                break
            except ValueError:
                continue
    if parsed is None:
        issues.append(
            _issue(
                "error",
                "生产日期格式无效，建议使用yyyy-mm-dd。",
                sheet=sheet,
                row=row,
                field=field,
            )
        )
    return parsed


def _parse_order_card(ws, issues):
    raw = {
        field: ws[value_cell].value
        for field, (_, value_cell, _) in FIELD_CELLS.items()
    }
    is_new_layout = (
        _text(ws["A12"].value) == "每日人员生产记录"
        or _text(ws["A8"].value).startswith("订单完工结算")
    )
    raw_settlement = (
        {
            field: ws[value_cell].value
            for field, (_, value_cell, _) in SETTLEMENT_CELLS.items()
        }
        if is_new_layout
        else {}
    )
    has_daily_values = any(
        ws.cell(row, column).value not in (None, "")
        for row in range(14 if is_new_layout else 11, ws.max_row + 1)
        for column in range(1, 5 if is_new_layout else 11)
    )
    has_settlement_values = any(
        value not in (None, "") for value in raw_settlement.values()
    )
    if (
        not any(value not in (None, "") for value in raw.values())
        and not has_daily_values
        and not has_settlement_values
    ):
        issues.append(
            _issue("warning", "该工作表未填写订单资料，已跳过。", sheet=ws.title)
        )
        return None

    for field, (label_cell, _, expected_label) in FIELD_CELLS.items():
        actual = _text(ws[label_cell].value)
        accepted_labels = FIELD_LABEL_ALIASES.get(field, {expected_label})
        if actual not in accepted_labels:
            issues.append(
                _issue(
                    "error",
                    f"固定字段标题应为“{expected_label}”，请使用系统模板。",
                    sheet=ws.title,
                    row=ws[label_cell].row,
                    field=field,
                )
            )

    if is_new_layout:
        for field, (label_cell, _, expected_label) in SETTLEMENT_CELLS.items():
            actual = _text(ws[label_cell].value)
            if actual != expected_label:
                issues.append(
                    _issue(
                        "error",
                        f"结算字段标题应为“{expected_label}”，请使用系统模板。",
                        sheet=ws.title,
                        row=ws[label_cell].row,
                        field=field,
                    )
                )

    record = {
        "row_key": ws.title,
        "sheet": ws.title,
        "station_code": _normalize_station_code(raw["station_code"]),
        "order_no": _text_value(
            raw["order_no"], issues, ws.title, 2, "order_no", max_length=100
        ),
        "mold_code": _text_value(
            raw["mold_code"], issues, ws.title, 2, "mold_code", max_length=100
        ),
        "specification": _text_value(
            raw["specification"],
            issues,
            ws.title,
            3,
            "specification",
            max_length=200,
        ),
        "material": _text_value(
            raw["material"], issues, ws.title, 3, "material", max_length=100
        ),
        "compound_size": _text_value(
            raw["compound_size"],
            issues,
            ws.title,
            4,
            "compound_size",
            max_length=100,
        ),
        "operator": _text_value(
            raw["operator"], issues, ws.title, 5, "operator", max_length=100
        ),
        "notes": _text(raw["notes"]),
    }
    for field in ["station_code", "order_no", "specification"]:
        if not record[field]:
            _, value_cell, _ = FIELD_CELLS[field]
            issues.append(
                _issue(
                    "error",
                    "该字段不能为空。",
                    sheet=ws.title,
                    row=ws[value_cell].row,
                    field=field,
                )
            )

    record["order_quantity"] = _integer_value(
        raw["order_quantity"], issues, ws.title, 3, "order_quantity", required=True, minimum=1
    )
    record["cavities"] = _integer_value(
        raw["cavities"],
        issues,
        ws.title,
        3,
        "cavities",
        required=True,
        minimum=1,
        maximum=MAX_POSITIVE_SMALL_INTEGER,
    )
    defect_rate = _decimal_value(
        raw["estimated_defect_rate"],
        issues,
        ws.title,
        4,
        "estimated_defect_rate",
        minimum=Decimal("0"),
    )
    if defect_rate is None:
        defect_rate = Decimal("0")
    number_format = str(ws[FIELD_CELLS["estimated_defect_rate"][1]].number_format or "")
    if "%" in number_format and defect_rate <= 1:
        defect_rate *= Decimal("100")
    defect_digits, defect_decimals = _decimal_shape(defect_rate)
    if defect_digits > 5 or defect_decimals > 2:
        issues.append(
            _issue(
                "error",
                "预估不良率最多填写2位小数且总位数不超过5位。",
                sheet=ws.title,
                row=4,
                field="estimated_defect_rate",
            )
        )
    if defect_rate > 100:
        issues.append(
            _issue(
                "error",
                "预估不良率必须在0至100之间。",
                sheet=ws.title,
                row=4,
                field="estimated_defect_rate",
            )
        )
    record["estimated_defect_rate"] = str(defect_rate)

    planned = _integer_value(
        raw["planned_mold_count"],
        issues,
        ws.title,
        4,
        "planned_mold_count",
        minimum=1,
    )
    if planned is None and record["order_quantity"] and record["cavities"]:
        target = Decimal(record["order_quantity"]) * (
            Decimal("1") + defect_rate / Decimal("100")
        )
        planned = max(math.ceil(target / Decimal(record["cavities"])), 1)
    record["planned_mold_count"] = planned

    for field, row, quantum, max_digits, decimal_places_count in [
        ("strip_weight_kg", 4, "0.001", 10, 3),
        ("estimated_hours", 5, "0.01", 10, 2),
        ("unit_price", 7, "0.0001", 14, 4),
        ("material_unit_price", 7, "0.0001", 14, 4),
    ]:
        value = _decimal_value(
            raw[field],
            issues,
            ws.title,
            row,
            field,
            max_digits=max_digits,
            decimal_places=decimal_places_count,
        )
        record[field] = (
            str(value.quantize(Decimal(quantum), rounding=ROUND_HALF_UP))
            if value is not None
            else None
        )
    record["strips_per_batch"] = _integer_value(
        raw["strips_per_batch"], issues, ws.title, 5, "strips_per_batch", minimum=1
    )
    record["curing_seconds"] = _integer_value(
        raw["curing_seconds"], issues, ws.title, 5, "curing_seconds", minimum=0
    )
    if record["curing_seconds"] is None:
        record["curing_seconds"] = 0
    if record["estimated_hours"] is None:
        hours = (
            Decimal(record["planned_mold_count"] or 0)
            * Decimal(record["curing_seconds"])
            / Decimal("3600")
        ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        record["estimated_hours"] = str(hours)
    record["strip_weight_kg"] = record["strip_weight_kg"] or None
    record["unit_price"] = record["unit_price"] or "0.0000"
    record["material_unit_price"] = record["material_unit_price"] or "0.0000"
    record.update(
        {
            "is_settled": False,
            "actual_good_quantity": 0,
            "actual_defective_quantity": 0,
            "total_material_kg": "0.000",
            "labor_cost": "0.00",
            "energy_cost": "0.00",
            "other_cost": "0.00",
            "settlement_notes": "",
        }
    )
    if has_settlement_values:
        record["is_settled"] = True
        for field, row in [
            ("actual_good_quantity", 9),
            ("actual_defective_quantity", 9),
        ]:
            value = _integer_value(
                raw_settlement.get(field),
                issues,
                ws.title,
                row,
                field,
                required=True,
                minimum=0,
            )
            record[field] = value or 0
        for field, row, places, decimals in [
            ("total_material_kg", 9, "0.001", 3),
            ("labor_cost", 10, "0.01", 2),
            ("energy_cost", 10, "0.01", 2),
            ("other_cost", 10, "0.01", 2),
        ]:
            value = _decimal_value(
                raw_settlement.get(field),
                issues,
                ws.title,
                row,
                field,
                required=True,
                max_digits=14,
                decimal_places=decimals,
            )
            record[field] = str(
                (value or Decimal("0")).quantize(
                    Decimal(places), rounding=ROUND_HALF_UP
                )
            )
        record["settlement_notes"] = _text(
            raw_settlement.get("settlement_notes")
        )

    loaded_at = _parse_datetime_value(raw["loaded_at"], issues, ws.title, 6, "loaded_at")
    expected_change_at = _parse_datetime_value(
        raw["expected_change_at"], issues, ws.title, 6, "expected_change_at"
    )
    unloaded_at = _parse_datetime_value(
        raw["unloaded_at"], issues, ws.title, 6, "unloaded_at"
    )
    raw_status = _text(raw["status"]).upper()
    status_value = STATUS_ALIASES.get(raw_status)
    if status_value is None:
        issues.append(
            _issue(
                "error",
                "状态必须为PLANNED、RUNNING、COMPLETED或CANCELLED。",
                sheet=ws.title,
                row=2,
                field="status",
            )
        )
        status_value = ProductionRun.Status.PLANNED
    if not status_value:
        status_value = (
            ProductionRun.Status.COMPLETED
            if unloaded_at
            else ProductionRun.Status.RUNNING
            if loaded_at
            else ProductionRun.Status.PLANNED
        )
    if status_value == ProductionRun.Status.PLANNED:
        if loaded_at:
            issues.append(
                _issue(
                    "error",
                    "待上机订单不能填写上模时间。",
                    sheet=ws.title,
                    row=6,
                    field="loaded_at",
                )
            )
        if expected_change_at:
            issues.append(
                _issue(
                    "error",
                    "待上机订单不能填写预计换模时间。",
                    sheet=ws.title,
                    row=6,
                    field="expected_change_at",
                )
            )
        if unloaded_at:
            issues.append(
                _issue(
                    "error",
                    "待上机订单不能填写下机时间。",
                    sheet=ws.title,
                    row=6,
                    field="unloaded_at",
                )
            )
    elif status_value == ProductionRun.Status.RUNNING:
        if not loaded_at:
            issues.append(
                _issue(
                    "error",
                    "生产中订单必须填写上模时间。",
                    sheet=ws.title,
                    row=6,
                    field="loaded_at",
                )
            )
        if unloaded_at:
            issues.append(
                _issue(
                    "error",
                    "生产中订单不能填写下机时间。",
                    sheet=ws.title,
                    row=6,
                    field="unloaded_at",
                )
            )
    elif status_value == ProductionRun.Status.COMPLETED:
        if not loaded_at:
            issues.append(
                _issue(
                    "error",
                    "完成订单必须填写上模时间。",
                    sheet=ws.title,
                    row=6,
                    field="loaded_at",
                )
            )
        if not unloaded_at:
            issues.append(
                _issue(
                    "error",
                    "完成订单必须填写下机时间。",
                    sheet=ws.title,
                    row=6,
                    field="unloaded_at",
                )
            )
    elif status_value == ProductionRun.Status.CANCELLED:
        if bool(loaded_at) != bool(unloaded_at):
            issues.append(
                _issue(
                    "error",
                    "已开机的取消订单必须同时保留上模和下机时间。",
                    sheet=ws.title,
                    row=6,
                    field="unloaded_at",
                )
            )
        if not loaded_at and expected_change_at:
            issues.append(
                _issue(
                    "error",
                    "未开机的取消订单不能填写预计换模时间。",
                    sheet=ws.title,
                    row=6,
                    field="expected_change_at",
                )
            )
    if loaded_at and unloaded_at and unloaded_at < loaded_at:
        issues.append(
            _issue(
                "error",
                "下机时间不能早于上模时间。",
                sheet=ws.title,
                row=6,
                field="unloaded_at",
            )
        )
    if loaded_at and expected_change_at is None:
        expected_change_at = loaded_at + timedelta(
            seconds=float(Decimal(record["estimated_hours"]) * Decimal("3600"))
        )
    if loaded_at and expected_change_at and expected_change_at < loaded_at:
        issues.append(
            _issue(
                "error",
                "预计换模时间不能早于上模时间。",
                sheet=ws.title,
                row=6,
                field="expected_change_at",
            )
        )
    if not loaded_at:
        expected_change_at = None
    record.update(
        {
            "status": status_value,
            "loaded_at": loaded_at.isoformat() if loaded_at else None,
            "expected_change_at": expected_change_at.isoformat()
            if expected_change_at
            else None,
            "unloaded_at": unloaded_at.isoformat() if unloaded_at else None,
        }
    )

    daily_headers = DAILY_HEADERS if is_new_layout else LEGACY_DAILY_HEADERS
    header_row = 13 if is_new_layout else 10
    first_data_row = 14 if is_new_layout else 11
    for column, (_, expected_header) in enumerate(daily_headers, 1):
        if _text(ws.cell(header_row, column).value) != expected_header:
            issues.append(
                _issue(
                    "error",
                    f"日报表第{column}列标题应为“{expected_header}”。",
                    sheet=ws.title,
                    row=header_row,
                    field=daily_headers[column - 1][0],
                )
            )

    logs = []
    seen_log_keys = set()
    legacy_good = 0
    legacy_defective = 0
    legacy_material = Decimal("0")
    legacy_labor = Decimal("0")
    legacy_energy = Decimal("0")
    legacy_other = Decimal("0")
    for row_no in range(first_data_row, ws.max_row + 1):
        values = [
            ws.cell(row_no, column).value
            for column in range(1, 5 if is_new_layout else 11)
        ]
        if not any(value not in (None, "") for value in values):
            continue
        production_date = _parse_date_value(
            values[0], issues, ws.title, row_no, "production_date"
        )
        operator = normalize_operator(
            _text_value(
                values[1],
                issues,
                ws.title,
                row_no,
                "operator",
                max_length=100,
            )
            or record["operator"]
        )
        if not operator:
            issues.append(
                _issue(
                    "error",
                    "作业员不能为空。",
                    sheet=ws.title,
                    row=row_no,
                    field="operator",
                )
            )
        log_key = (production_date, operator)
        if production_date and operator and log_key in seen_log_keys:
            issues.append(
                _issue(
                    "error",
                    "同一订单中同一作业员当天只能有一条生产记录。",
                    sheet=ws.title,
                    row=row_no,
                    field="operator",
                )
            )
        if production_date and operator:
            seen_log_keys.add(log_key)
            loaded_date = timezone.localtime(loaded_at).date() if loaded_at else None
            unloaded_date = (
                timezone.localtime(unloaded_at).date() if unloaded_at else None
            )
            if loaded_date and production_date < loaded_date:
                issues.append(
                    _issue(
                        "error",
                        "生产日期不能早于上模日期。",
                        sheet=ws.title,
                        row=row_no,
                        field="production_date",
                    )
                )
            if (
                status_value == ProductionRun.Status.RUNNING
                and production_date > timezone.localdate()
            ):
                issues.append(
                    _issue(
                        "error",
                        "生产中订单的日报日期不能晚于今天。",
                        sheet=ws.title,
                        row=row_no,
                        field="production_date",
                    )
                )
            if (
                status_value
                in (ProductionRun.Status.COMPLETED, ProductionRun.Status.CANCELLED)
                and unloaded_date
                and production_date > unloaded_date
            ):
                issues.append(
                    _issue(
                        "error",
                        "生产日期不能晚于下机日期。",
                        sheet=ws.title,
                        row=row_no,
                        field="production_date",
                    )
                )
        produced = _integer_value(
            values[2],
            issues,
            ws.title,
            row_no,
            "produced_mold_count",
            required=True,
            minimum=1,
        )
        notes = _text(values[3] if is_new_layout else values[9])
        if not is_new_layout:
            defective = _integer_value(
                values[4], issues, ws.title, row_no, "defective_quantity"
            ) or 0
            good = _integer_value(
                values[3], issues, ws.title, row_no, "good_quantity"
            )
            if good is None and produced is not None and record["cavities"]:
                good = max(produced * record["cavities"] - defective, 0)
            if produced is not None and record["cavities"]:
                expected_quantity = produced * record["cavities"]
                if (good or 0) + defective != expected_quantity:
                    issues.append(
                        _issue(
                            "error",
                            "旧模板中良品与不良之和必须等于生产模数乘以模具孔数。",
                            sheet=ws.title,
                            row=row_no,
                            field="good_quantity",
                        )
                    )
            legacy_good += good or 0
            legacy_defective += defective
            decimal_values = []
            for index, field, max_digits, decimal_places in [
                (5, "material_kg", 12, 3),
                (6, "labor_cost", 14, 2),
                (7, "energy_cost", 14, 2),
                (8, "other_cost", 14, 2),
            ]:
                decimal_values.append(
                    _decimal_value(
                        values[index],
                        issues,
                        ws.title,
                        row_no,
                        field,
                        max_digits=max_digits,
                        decimal_places=decimal_places,
                    )
                    or Decimal("0")
                )
            legacy_material += decimal_values[0]
            legacy_labor += decimal_values[1]
            legacy_energy += decimal_values[2]
            legacy_other += decimal_values[3]
        logs.append(
            {
                "production_date": production_date.isoformat() if production_date else None,
                "operator": operator,
                "produced_mold_count": produced or 0,
                "notes": notes,
                "row_no": row_no,
            }
        )
    settlement_allowed = status_value == ProductionRun.Status.COMPLETED or (
        status_value == ProductionRun.Status.CANCELLED and loaded_at is not None
    )
    if not is_new_layout and logs and settlement_allowed:
        record.update(
            {
                "is_settled": True,
                "actual_good_quantity": legacy_good,
                "actual_defective_quantity": legacy_defective,
                "total_material_kg": str(
                    legacy_material.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
                ),
                "labor_cost": str(
                    legacy_labor.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                ),
                "energy_cost": str(
                    legacy_energy.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                ),
                "other_cost": str(
                    legacy_other.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                ),
            }
        )
        issues.append(
            _issue(
                "warning",
                "检测到旧版日报成本字段，已按订单汇总为完工结算。",
                sheet=ws.title,
                row=header_row,
                field="settlement",
            )
        )
    if record["is_settled"] and not settlement_allowed:
        issues.append(
            _issue(
                "error",
                "只有已完成或已上模后取消的订单可以填写完工结算。",
                sheet=ws.title,
                row=9 if is_new_layout else header_row,
                field="settlement",
            )
        )
    if record["is_settled"]:
        expected_quantity = sum(
            log["produced_mold_count"] for log in logs
        ) * int(record["cavities"] or 0)
        actual_quantity = (
            record["actual_good_quantity"] + record["actual_defective_quantity"]
        )
        if actual_quantity != expected_quantity:
            issues.append(
                _issue(
                    "error",
                    f"实际良品与实际不良之和必须等于累计生产模数乘以模具孔数（当前应为{expected_quantity}件）。",
                    sheet=ws.title,
                    row=9 if is_new_layout else header_row,
                    field="actual_good_quantity",
                )
            )
    if logs and record["status"] == ProductionRun.Status.PLANNED:
        issues.append(
            _issue(
                "error",
                "待上机订单不能填写生产日报。",
                sheet=ws.title,
            row=first_data_row,
                field="production_date",
            )
        )
    if (
        logs
        and record["status"] == ProductionRun.Status.CANCELLED
        and not loaded_at
    ):
        issues.append(
            _issue(
                "error",
                "未上模即取消的订单不能填写生产日报。",
                sheet=ws.title,
                row=first_data_row,
                field="production_date",
            )
        )
    record["daily_logs"] = logs
    record["daily_log_count"] = len(logs)
    return record


def _record_time_range(record):
    return (
        _payload_datetime(record.get("loaded_at")),
        _payload_datetime(record.get("unloaded_at")),
    )


def _time_ranges_overlap(start_a, end_a, start_b, end_b):
    if not start_a or not start_b:
        return False
    if end_a is not None and end_a <= start_a:
        return False
    if end_b is not None and end_b <= start_b:
        return False
    return (end_a is None or start_b < end_a) and (
        end_b is None or start_a < end_b
    )


def _database_time_overlaps(station, loaded_at, unloaded_at):
    if not station or not loaded_at:
        return ProductionRun.objects.none()
    if unloaded_at is not None and unloaded_at <= loaded_at:
        return ProductionRun.objects.none()
    queryset = ProductionRun.objects.filter(
        station=station,
        loaded_at__isnull=False,
    )
    if unloaded_at is not None:
        queryset = queryset.filter(loaded_at__lt=unloaded_at)
    return queryset.filter(
        Q(unloaded_at__isnull=True) | Q(unloaded_at__gt=loaded_at)
    )


def _preflight_database(records, issues):
    station_codes = {record["station_code"] for record in records if record["station_code"]}
    mold_codes = {record["mold_code"] for record in records if record["mold_code"]}
    stations = {
        station.code: station
        for station in ProductionStation.objects.filter(
            code__in=station_codes, is_active=True
        ).select_related("machine")
    }
    molds = {
        mold.asset_code: mold
        for mold in MoldAsset.objects.filter(
            asset_code__in=mold_codes, is_active=True
        ).select_related("current_machine")
    }

    active_station_sheets = {}
    active_mold_sheets = {}
    timed_station_records = {}
    order_keys = Counter(
        (record["station_code"], record["order_no"])
        for record in records
        if record["station_code"] and record["order_no"]
    )
    for record in records:
        sheet = record["sheet"]
        station_code = record["station_code"]
        mold_code = record["mold_code"]
        station = stations.get(station_code)
        mold = molds.get(mold_code) if mold_code else None
        loaded_at, unloaded_at = _record_time_range(record)
        if station is None:
            issues.append(
                _issue(
                    "error",
                    f"机台不存在或已停用：{station_code}",
                    sheet=sheet,
                    row=2,
                    field="station_code",
                )
            )
        if mold_code and mold is None:
            issues.append(
                _issue(
                    "error",
                    f"模具实体不存在或已停用：{mold_code}",
                    sheet=sheet,
                    row=2,
                    field="mold_code",
                )
            )
        if order_keys[(station_code, record["order_no"])] > 1:
            issues.append(
                _issue(
                    "error",
                    "同一机台和订单号在工作簿中重复。",
                    sheet=sheet,
                    row=2,
                    field="order_no",
                )
            )
        if station and ProductionRun.objects.filter(
            station=station, order_no=record["order_no"]
        ).exists():
            issues.append(
                _issue(
                    "error",
                    "数据库中已存在该机台和订单号，请勿重复导入。",
                    sheet=sheet,
                    row=2,
                    field="order_no",
                )
            )
        if station and loaded_at:
            for other_record, other_loaded, other_unloaded in timed_station_records.get(
                station_code, []
            ):
                if _time_ranges_overlap(
                    loaded_at,
                    unloaded_at,
                    other_loaded,
                    other_unloaded,
                ):
                    issues.append(
                        _issue(
                            "error",
                            f"与工作表“{other_record['sheet']}”的生产时段重叠。",
                            sheet=sheet,
                            row=6,
                            field="loaded_at",
                        )
                    )
            timed_station_records.setdefault(station_code, []).append(
                (record, loaded_at, unloaded_at)
            )
            if _database_time_overlaps(station, loaded_at, unloaded_at).exists():
                issues.append(
                    _issue(
                        "error",
                        "该机台在所填生产时段内已有其他订单。",
                        sheet=sheet,
                        row=6,
                        field="loaded_at",
                    )
                )
        if record["status"] in ProductionRun.ACTIVE_STATUSES:
            if station_code in active_station_sheets:
                issues.append(
                    _issue(
                        "error",
                        f"与工作表“{active_station_sheets[station_code]}”占用同一机台。",
                        sheet=sheet,
                        row=2,
                        field="station_code",
                    )
                )
            else:
                active_station_sheets[station_code] = sheet
            if station and ProductionRun.objects.filter(
                station=station, status__in=ProductionRun.ACTIVE_STATUSES
            ).exists():
                issues.append(
                    _issue(
                        "error",
                        "该机台已有待上机或生产中的订单。",
                        sheet=sheet,
                        row=2,
                        field="station_code",
                    )
                )
            if mold_code:
                if mold_code in active_mold_sheets:
                    issues.append(
                        _issue(
                            "error",
                            f"与工作表“{active_mold_sheets[mold_code]}”占用同一模具。",
                            sheet=sheet,
                            row=2,
                            field="mold_code",
                        )
                    )
                else:
                    active_mold_sheets[mold_code] = sheet
                if mold and ProductionRun.objects.filter(
                    mold=mold, status__in=ProductionRun.ACTIVE_STATUSES
                ).exists():
                    issues.append(
                        _issue(
                            "error",
                            "该模具已用于另一条未结束的生产订单。",
                            sheet=sheet,
                            row=2,
                            field="mold_code",
                        )
                    )
            if mold:
                station_machine_id = station.machine_id if station else None
                if mold.status == MoldAsset.Status.OUTSOURCED:
                    issues.append(
                        _issue(
                            "error",
                            "待上机或生产中的订单不能关联客户收回的模具。",
                            sheet=sheet,
                            row=2,
                            field="mold_code",
                        )
                    )
                elif record["status"] == ProductionRun.Status.RUNNING:
                    if station_machine_id is None:
                        issues.append(
                            _issue(
                                "error",
                                "该生产机台尚未关联模具台账机台，不能登记为生产中。",
                                sheet=sheet,
                                row=2,
                                field="station_code",
                            )
                        )
                    elif (
                        mold.status != MoldAsset.Status.ON_MACHINE
                        or mold.current_machine_id != station_machine_id
                    ):
                        issues.append(
                            _issue(
                                "error",
                                "生产中的模具必须已上到该生产机台关联的模具台账机台。",
                                sheet=sheet,
                                row=2,
                                field="mold_code",
                            )
                        )
                elif (
                    mold.status == MoldAsset.Status.ON_MACHINE
                    and station_machine_id is not None
                    and mold.current_machine_id != station_machine_id
                ):
                    issues.append(
                        _issue(
                            "error",
                            "待上机订单中的模具不能占用其他机台。",
                            sheet=sheet,
                            row=2,
                            field="mold_code",
                        )
                    )


def preview_production_workbook(uploaded_file, user):
    _validate_workbook_archive(uploaded_file)
    workbook = load_workbook(uploaded_file, data_only=True, read_only=True)
    _validate_workbook_dimensions(workbook)
    issues = []
    records = []
    candidate_sheets = [ws for ws in workbook.worksheets if not ws.title.startswith("_")]
    if not candidate_sheets:
        raise ValueError("工作簿中没有可导入的订单卡。")
    for sheet in candidate_sheets:
        record = _parse_order_card(sheet, issues)
        if record:
            records.append(record)
    workbook.close()
    _resolve_order_links(records, issues)
    _preflight_database(records, issues)
    errors = [issue for issue in issues if issue["level"] == "error"]
    warnings = [issue for issue in issues if issue["level"] == "warning"]
    batch = ProductionImportBatch.objects.create(
        original_name=uploaded_file.name[:255],
        created_by=user,
        payload={"source_type": "order_cards", "rows": records},
        errors=errors,
        warnings=warnings,
    )
    preview_rows = [
        {
            key: record.get(key)
            for key in [
                "row_key",
                "sheet",
                "station_code",
                "order_no",
                "status",
                "mold_code",
                "order_id",
                "product_specification_id",
                "specification",
                "material",
                "order_quantity",
                "cavities",
                "planned_mold_count",
                "estimated_hours",
                "loaded_at",
                "expected_change_at",
                "unloaded_at",
                "is_settled",
                "actual_good_quantity",
                "actual_defective_quantity",
                "total_material_kg",
                "labor_cost",
                "energy_cost",
                "other_cost",
                "settlement_notes",
                "daily_log_count",
            ]
        }
        for record in records
    ]
    return {
        "token": str(batch.pk),
        "source_type": "order_cards",
        "sheet_count": len(candidate_sheets),
        "total_rows": len(records),
        "daily_log_count": sum(record["daily_log_count"] for record in records),
        "error_count": len(errors),
        "warning_count": len(warnings),
        "rows": preview_rows,
        "issues": issues,
    }


def _payload_datetime(value):
    if not value:
        return None
    parsed = parse_datetime(value)
    if parsed and timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed


def commit_production_batch(batch, user):
    try:
        with transaction.atomic():
            current = ProductionImportBatch.objects.filter(pk=batch.pk).first()
            if current is None or current.created_by_id != user.pk:
                raise ValueError("无权提交该导入批次。")
            if current.status != ProductionImportBatch.Status.PREVIEWED:
                raise ValueError("该导入批次已经提交或正在提交，不能重复导入。")
            if current.errors:
                raise ValueError("预检存在错误，请修正Excel后重新上传。")
            claimed = ProductionImportBatch.objects.filter(
                pk=batch.pk,
                created_by=user,
                status=ProductionImportBatch.Status.PREVIEWED,
            ).update(status=ProductionImportBatch.Status.COMMITTING)
            if claimed != 1:
                raise ValueError("该导入批次已经提交或正在提交，不能重复导入。")

            locked = ProductionImportBatch.objects.get(pk=batch.pk)
            records = locked.payload.get("rows", [])
            commit_issues = []
            preview_links = [
                (
                    record.get("order_id"),
                    record.get("product_specification_id"),
                )
                for record in records
            ]
            _resolve_order_links(records, commit_issues, for_update=True)
            current_links = [
                (
                    record.get("order_id"),
                    record.get("product_specification_id"),
                )
                for record in records
            ]
            if current_links != preview_links:
                raise ValueError("订单主档匹配结果已变化，请重新上传并预检。")
            _preflight_database(records, commit_issues)
            commit_errors = [
                issue for issue in commit_issues if issue["level"] == "error"
            ]
            if commit_errors:
                raise ValueError(
                    f"提交前复查失败：{commit_errors[0]['message']}请重新预检。"
                )

            imported_count = 0
            log_count = 0
            settled_count = 0
            order_ids = {record.get("order_id") for record in records if record.get("order_id")}
            specification_ids = {
                record.get("product_specification_id")
                for record in records
                if record.get("product_specification_id")
            }
            linked_orders = {
                order.pk: order
                for order in QualityOrder.objects.select_for_update().filter(pk__in=order_ids)
            }
            linked_specifications = {
                specification.pk: specification
                for specification in ProductSpecification.objects.select_for_update().filter(
                    pk__in=specification_ids
                )
            }
            for record in records:
                station = ProductionStation.objects.select_for_update().filter(
                    code=record["station_code"], is_active=True
                ).first()
                if station is None:
                    raise ValueError(
                        f"机台{record['station_code']}不存在或已停用，请重新预检。"
                    )
                mold = None
                if record.get("mold_code"):
                    mold = MoldAsset.objects.select_for_update().filter(
                        asset_code=record["mold_code"], is_active=True
                    ).first()
                    if mold is None:
                        raise ValueError(
                            f"模具{record['mold_code']}不存在或已停用，请重新预检。"
                        )
                run = ProductionRun(
                    station=station,
                    mold=mold,
                    order=linked_orders.get(record.get("order_id")),
                    product_specification=linked_specifications.get(
                        record.get("product_specification_id")
                    ),
                    order_no=record["order_no"],
                    specification=record["specification"],
                    material=record.get("material", ""),
                    order_quantity=record["order_quantity"],
                    cavities=record["cavities"],
                    estimated_defect_rate=Decimal(record["estimated_defect_rate"]),
                    planned_mold_count=record["planned_mold_count"],
                    compound_size=record.get("compound_size", ""),
                    strip_weight_kg=Decimal(record["strip_weight_kg"])
                    if record.get("strip_weight_kg")
                    else None,
                    strips_per_batch=record.get("strips_per_batch"),
                    curing_seconds=record.get("curing_seconds", 0),
                    estimated_hours=Decimal(record["estimated_hours"]),
                    loaded_at=_payload_datetime(record.get("loaded_at")),
                    expected_change_at=_payload_datetime(
                        record.get("expected_change_at")
                    ),
                    unloaded_at=_payload_datetime(record.get("unloaded_at")),
                    status=record["status"],
                    operator=record.get("operator", ""),
                    unit_price=Decimal(record.get("unit_price") or "0"),
                    material_unit_price=Decimal(
                        record.get("material_unit_price") or "0"
                    ),
                    notes=record.get("notes", ""),
                    created_by=user,
                )
                run.full_clean()
                run.save()
                imported_count += 1
                for log_record in record.get("daily_logs", []):
                    log = ProductionDailyLog(
                        run=run,
                        production_date=parse_date(log_record["production_date"]),
                        operator=log_record.get("operator", ""),
                        produced_mold_count=log_record.get("produced_mold_count", 0),
                        notes=log_record.get("notes", ""),
                    )
                    log.save()
                    log_count += 1
                if record.get("is_settled"):
                    run.actual_good_quantity = record.get("actual_good_quantity", 0)
                    run.actual_defective_quantity = record.get(
                        "actual_defective_quantity", 0
                    )
                    run.total_material_kg = Decimal(
                        record.get("total_material_kg") or "0"
                    )
                    run.labor_cost = Decimal(record.get("labor_cost") or "0")
                    run.energy_cost = Decimal(record.get("energy_cost") or "0")
                    run.other_cost = Decimal(record.get("other_cost") or "0")
                    run.settlement_notes = record.get("settlement_notes", "")
                    # Excel has no separate settlement date. Historical imports use
                    # their down-machine time so old finance is not counted today.
                    run.settled_at = run.unloaded_at
                    run.settled_by = user
                    run.save(
                        update_fields=[
                            "actual_good_quantity",
                            "actual_defective_quantity",
                            "total_material_kg",
                            "labor_cost",
                            "energy_cost",
                            "other_cost",
                            "settlement_notes",
                            "settled_at",
                            "settled_by",
                            "updated_at",
                        ]
                    )
                    record_settlement_revision(
                        run, user, ProductionSettlementRevision.Action.SETTLED
                    )
                    settled_count += 1

            committed_at = timezone.now()
            updated = ProductionImportBatch.objects.filter(
                pk=locked.pk, status=ProductionImportBatch.Status.COMMITTING
            ).update(
                status=ProductionImportBatch.Status.COMMITTED,
                committed_at=committed_at,
            )
            if updated != 1:
                raise ValueError("导入批次状态已变化，请刷新后确认结果。")
            return {
                "imported_count": imported_count,
                "log_count": log_count,
                "settled_count": settled_count,
            }
    except ValueError:
        raise
    except DjangoValidationError as exc:
        messages = []
        if hasattr(exc, "message_dict"):
            for field, field_messages in exc.message_dict.items():
                messages.extend(f"{field}：{message}" for message in field_messages)
        else:
            messages.extend(exc.messages)
        raise ValueError(f"提交数据校验失败：{'；'.join(messages)}") from exc
    except IntegrityError as exc:
        raise ValueError("数据已被其他操作占用或重复，请重新预检后再导入。") from exc
    except OperationalError as exc:
        raise ValueError("数据库正忙，导入未执行，请稍后重新提交。") from exc


def _safe_excel_text(value):
    if not isinstance(value, str):
        return value
    if value.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def create_production_error_report(batch):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "生产导入预检问题"
    sheet.append(["级别", "工作表", "行号", "字段", "说明"])
    for issue in [*batch.errors, *batch.warnings]:
        sheet.append(
            [
                _safe_excel_text(issue.get("level", "")),
                _safe_excel_text(issue.get("sheet", "")),
                issue.get("row", ""),
                _safe_excel_text(issue.get("field", "")),
                _safe_excel_text(issue.get("message", "")),
            ]
        )
    dark = "1F4E78"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.alignment = Alignment(horizontal="center")
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == "error":
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="FCE4D6")
        elif row[0].value == "warning":
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
    sheet.freeze_panes = "A2"
    for column, width in {"A": 12, "B": 26, "C": 10, "D": 24, "E": 64}.items():
        sheet.column_dimensions[column].width = width
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
